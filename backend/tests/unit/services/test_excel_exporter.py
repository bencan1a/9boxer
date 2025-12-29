"""Tests for Excel exporter service."""

from collections.abc import Generator
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import pytest

from ninebox.models.employee import Employee, PerformanceLevel, PotentialLevel
from ninebox.services.excel_exporter import ExcelExporter


pytestmark = pytest.mark.unit


@pytest.fixture(autouse=True)
def cleanup_openpyxl_state() -> Generator[None, None, None]:
    """Clean up openpyxl global state before and after each test.

    openpyxl maintains global state in its NumberFormat registry and other
    internal caches. This fixture ensures tests start with a clean state
    and don't pollute other tests.
    """
    import gc

    # Before test: Force garbage collection to clean up any lingering workbook references
    gc.collect()

    yield

    # After test: Force garbage collection again
    gc.collect()


@pytest.fixture
def excel_exporter() -> ExcelExporter:
    """Create exporter instance."""
    return ExcelExporter()


def test_export_when_valid_file_then_creates_excel(
    excel_exporter: ExcelExporter,
    sample_excel_file: Path,
    sample_employees: list[Employee],
    tmp_path: Path,
) -> None:
    """Test exporting creates a valid Excel file."""
    output_path = tmp_path / "output.xlsx"

    excel_exporter.export(sample_excel_file, sample_employees, output_path)

    assert output_path.exists()

    # Verify it's a valid Excel file
    workbook = openpyxl.load_workbook(output_path)
    assert len(workbook.worksheets) >= 2
    workbook.close()  # Prevent openpyxl state pollution


def test_export_when_called_then_adds_modified_column(
    excel_exporter: ExcelExporter,
    sample_excel_file: Path,
    sample_employees: list[Employee],
    tmp_path: Path,
) -> None:
    """Test that 'Modified in Session' column is added."""
    output_path = tmp_path / "output.xlsx"

    excel_exporter.export(sample_excel_file, sample_employees, output_path)

    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook.worksheets[1]

    # Find "Modified in Session" column
    modified_col_found = False
    for col_idx in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, col_idx).value
        if cell_value and "Modified in Session" in str(cell_value):
            modified_col_found = True
            break

    assert modified_col_found
    workbook.close()  # Prevent openpyxl state pollution


def test_export_when_called_then_adds_modification_date_column(
    excel_exporter: ExcelExporter,
    sample_excel_file: Path,
    sample_employees: list[Employee],
    tmp_path: Path,
) -> None:
    """Test that modification date column is added."""
    output_path = tmp_path / "output.xlsx"

    excel_exporter.export(sample_excel_file, sample_employees, output_path)

    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook.worksheets[1]

    # Find "Modification Date" column
    date_col_found = False
    for col_idx in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, col_idx).value
        if cell_value and "Modification Date" in str(cell_value):
            date_col_found = True
            break

    assert date_col_found
    workbook.close()  # Prevent openpyxl state pollution


def test_export_when_employee_modified_then_writes_modified_values(
    excel_exporter: ExcelExporter,
    sample_excel_file: Path,
    sample_employees: list[Employee],
    tmp_path: Path,
) -> None:
    """Test that modified employee values are written."""
    # Mark first employee as modified
    sample_employees[0].modified_in_session = True
    sample_employees[0].last_modified = datetime.now(timezone.utc)
    sample_employees[0].performance = PerformanceLevel.MEDIUM
    sample_employees[0].potential = PerformanceLevel.LOW

    output_path = tmp_path / "output.xlsx"
    excel_exporter.export(sample_excel_file, sample_employees, output_path)

    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook.worksheets[1]

    # Find modified column
    modified_col = None
    for col_idx in range(1, sheet.max_column + 1):
        if sheet.cell(1, col_idx).value and "Modified in Session" in str(
            sheet.cell(1, col_idx).value
        ):
            modified_col = col_idx
            break

    assert modified_col is not None

    # Check first employee row (row 2)
    assert sheet.cell(2, modified_col).value == "Yes"
    workbook.close()  # Prevent openpyxl state pollution


def test_export_when_employee_not_modified_then_marks_as_no(
    excel_exporter: ExcelExporter,
    sample_excel_file: Path,
    sample_employees: list[Employee],
    tmp_path: Path,
) -> None:
    """Test that unmodified employees are marked as 'No'."""
    output_path = tmp_path / "output.xlsx"
    excel_exporter.export(sample_excel_file, sample_employees, output_path)

    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook.worksheets[1]

    # Find modified column
    modified_col = None
    for col_idx in range(1, sheet.max_column + 1):
        if sheet.cell(1, col_idx).value and "Modified in Session" in str(
            sheet.cell(1, col_idx).value
        ):
            modified_col = col_idx
            break

    # All unmodified employees should be "No"
    for row_idx in range(2, sheet.max_row + 1):
        assert sheet.cell(row_idx, modified_col).value == "No"
    workbook.close()  # Prevent openpyxl state pollution


def test_export_when_multiple_employees_modified_then_updates_all(
    excel_exporter: ExcelExporter,
    sample_excel_file: Path,
    sample_employees: list[Employee],
    tmp_path: Path,
) -> None:
    """Test that multiple modified employees are updated correctly."""
    # Mark multiple employees as modified
    sample_employees[0].modified_in_session = True
    sample_employees[0].last_modified = datetime.now(timezone.utc)
    sample_employees[2].modified_in_session = True
    sample_employees[2].last_modified = datetime.now(timezone.utc)

    output_path = tmp_path / "output.xlsx"
    excel_exporter.export(sample_excel_file, sample_employees, output_path)

    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook.worksheets[1]

    # Find modified column
    modified_col = None
    for col_idx in range(1, sheet.max_column + 1):
        if sheet.cell(1, col_idx).value and "Modified in Session" in str(
            sheet.cell(1, col_idx).value
        ):
            modified_col = col_idx
            break

    # Check employees 1 and 3 are marked modified
    assert sheet.cell(2, modified_col).value == "Yes"  # Employee 1
    assert sheet.cell(3, modified_col).value == "No"  # Employee 2
    assert sheet.cell(4, modified_col).value == "Yes"  # Employee 3
    workbook.close()  # Prevent openpyxl state pollution


def test_export_when_file_has_less_than_two_sheets_then_raises_error(
    excel_exporter: ExcelExporter, sample_employees: list[Employee], tmp_path: Path
) -> None:
    """Test error when file has less than 2 sheets."""
    # Create file with only one sheet
    single_sheet_file = tmp_path / "single_sheet.xlsx"
    workbook = openpyxl.Workbook()
    workbook.save(single_sheet_file)
    workbook.close()  # Explicitly close to prevent openpyxl state pollution

    output_path = tmp_path / "output.xlsx"

    with pytest.raises(ValueError, match="Excel file must have at least 2 sheets"):
        excel_exporter.export(single_sheet_file, sample_employees, output_path)


def test_export_when_performance_updated_then_writes_new_value(
    excel_exporter: ExcelExporter,
    sample_excel_file: Path,
    sample_employees: list[Employee],
    tmp_path: Path,
) -> None:
    """Test that updated performance values are written."""
    # Update performance
    sample_employees[0].performance = PerformanceLevel.LOW
    sample_employees[0].potential = PotentialLevel.LOW
    sample_employees[0].grid_position = 1

    output_path = tmp_path / "output.xlsx"
    excel_exporter.export(sample_excel_file, sample_employees, output_path)

    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook.worksheets[1]

    # Find performance column
    perf_col = None
    for col_idx in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, col_idx).value
        if cell_value and "Aug 2025 Talent Assessment Performance" in str(cell_value):
            perf_col = col_idx
            break

    if perf_col:
        assert sheet.cell(2, perf_col).value == "Low"
    workbook.close()  # Prevent openpyxl state pollution


def test_export_when_modification_date_set_then_writes_date(
    excel_exporter: ExcelExporter,
    sample_excel_file: Path,
    sample_employees: list[Employee],
    tmp_path: Path,
) -> None:
    """Test that modification date is written."""
    modification_time = datetime.now(timezone.utc)
    sample_employees[0].modified_in_session = True
    sample_employees[0].last_modified = modification_time

    output_path = tmp_path / "output.xlsx"
    excel_exporter.export(sample_excel_file, sample_employees, output_path)

    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook.worksheets[1]

    # Find modification date column (should be right after modified column)
    modified_col = None
    for col_idx in range(1, sheet.max_column + 1):
        if sheet.cell(1, col_idx).value and "Modified in Session" in str(
            sheet.cell(1, col_idx).value
        ):
            modified_col = col_idx
            break

    date_col = modified_col + 1
    date_value = sheet.cell(2, date_col).value

    # Should have written the ISO format timestamp
    assert date_value is not None
    assert modification_time.isoformat() in str(date_value)
    workbook.close()  # Prevent openpyxl state pollution


def test_find_column_when_column_exists_then_returns_index(
    excel_exporter: ExcelExporter, sample_excel_file: Path
) -> None:
    """Test finding column by name."""
    workbook = openpyxl.load_workbook(sample_excel_file)
    sheet = workbook.worksheets[1]

    col_idx = excel_exporter._find_column(sheet, "Employee ID")

    assert col_idx == 1
    workbook.close()  # Prevent openpyxl state pollution


def test_find_column_when_column_not_exists_and_create_false_then_returns_none(
    excel_exporter: ExcelExporter, sample_excel_file: Path
) -> None:
    """Test finding non-existent column without create flag."""
    workbook = openpyxl.load_workbook(sample_excel_file)
    sheet = workbook.worksheets[1]

    col_idx = excel_exporter._find_column(sheet, "NonExistentColumn", create=False)

    assert col_idx is None
    workbook.close()  # Prevent openpyxl state pollution


def test_find_column_when_column_not_exists_and_create_true_then_returns_next_index(
    excel_exporter: ExcelExporter, sample_excel_file: Path
) -> None:
    """Test finding non-existent column with create flag."""
    workbook = openpyxl.load_workbook(sample_excel_file)
    sheet = workbook.worksheets[1]

    original_max_col = sheet.max_column
    col_idx = excel_exporter._find_column(sheet, "NewColumn", create=True)

    assert col_idx == original_max_col + 1
    workbook.close()  # Prevent openpyxl state pollution


# ========== Donut Mode Tests ==========


def test_export_when_donut_data_exists_then_includes_columns(
    excel_exporter: ExcelExporter,
    sample_excel_file: Path,
    sample_employees: list[Employee],
    tmp_path: Path,
) -> None:
    """Test export includes donut exercise columns when data exists."""
    # Mark employee with donut placement
    sample_employees[0].donut_modified = True
    sample_employees[0].donut_position = 6
    sample_employees[0].donut_notes = "Testing hypothetical placement"

    output_path = tmp_path / "output.xlsx"
    excel_exporter.export(sample_excel_file, sample_employees, output_path)

    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook.worksheets[1]

    # Find donut columns
    donut_position_col_found = False
    donut_label_col_found = False
    donut_description_col_found = False
    donut_notes_col_found = False

    for col_idx in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, col_idx).value
        if cell_value:
            cell_str = str(cell_value)
            if "Donut Exercise Position" in cell_str:
                donut_position_col_found = True
            if "Donut Exercise Label" in cell_str:
                donut_label_col_found = True
            if "Donut Exercise Change Description" in cell_str:
                donut_description_col_found = True
            if "Donut Exercise Notes" in cell_str:
                donut_notes_col_found = True

    assert donut_position_col_found
    assert donut_label_col_found
    assert donut_description_col_found
    assert donut_notes_col_found
    workbook.close()  # Prevent openpyxl state pollution


def test_export_when_donut_data_exists_then_populates_correctly(
    excel_exporter: ExcelExporter,
    sample_excel_file: Path,
    sample_employees: list[Employee],
    tmp_path: Path,
) -> None:
    """Test donut data is populated correctly in export."""
    # Setup session with donut changes
    from ninebox.models.events import DonutMoveEvent  # noqa: PLC0415
    from ninebox.models.session import SessionState  # noqa: PLC0415

    # Mark employee with donut placement (employee 1: moved from H,H to H,M)
    sample_employees[0].donut_modified = True
    sample_employees[0].donut_position = 6
    sample_employees[0].donut_performance = PerformanceLevel.HIGH
    sample_employees[0].donut_potential = PotentialLevel.MEDIUM
    sample_employees[0].donut_notes = "Exploring management track"

    # Create session with donut changes
    session = SessionState(
        session_id="test-session",
        user_id="user1",
        created_at=datetime.now(timezone.utc),
        original_employees=sample_employees,
        current_employees=sample_employees,
        original_filename="test.xlsx",
        original_file_path="/tmp/test.xlsx",
        sheet_name="Employee Data",
        sheet_index=1,
        events=[],
        donut_events=[
            DonutMoveEvent(
                employee_id=1,
                employee_name="Alice Smith",
                timestamp=datetime.now(timezone.utc),
                old_performance=PerformanceLevel.HIGH,
                old_potential=PotentialLevel.HIGH,
                new_performance=PerformanceLevel.HIGH,
                new_potential=PotentialLevel.MEDIUM,
                old_position=9,
                new_position=6,
                notes="Exploring management track",
            )
        ],
    )

    output_path = tmp_path / "output.xlsx"
    excel_exporter.export(sample_excel_file, sample_employees, output_path, session=session)

    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook.worksheets[1]

    # Find donut columns
    donut_position_col = None
    donut_label_col = None
    donut_description_col = None
    donut_notes_col = None

    for col_idx in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, col_idx).value
        if cell_value:
            cell_str = str(cell_value)
            if "Donut Exercise Position" == cell_str:
                donut_position_col = col_idx
            elif "Donut Exercise Label" == cell_str:
                donut_label_col = col_idx
            elif "Donut Exercise Change Description" == cell_str:
                donut_description_col = col_idx
            elif "Donut Exercise Notes" == cell_str:
                donut_notes_col = col_idx

    assert donut_position_col is not None
    assert donut_label_col is not None
    assert donut_description_col is not None
    assert donut_notes_col is not None

    # Check first employee row (row 2) has donut data
    assert sheet.cell(2, donut_position_col).value == 6
    assert sheet.cell(2, donut_label_col).value == "High Impact [H,M]"
    assert "Donut: Moved from Star [H,H] to High Impact [H,M]" in str(
        sheet.cell(2, donut_description_col).value
    )
    assert sheet.cell(2, donut_notes_col).value == "Exploring management track"
    workbook.close()  # Prevent openpyxl state pollution


def test_export_when_no_donut_data_then_empty_cells(
    excel_exporter: ExcelExporter,
    sample_excel_file: Path,
    sample_employees: list[Employee],
    tmp_path: Path,
) -> None:
    """Test employees without donut placements have empty cells."""
    # No donut modifications for any employee
    output_path = tmp_path / "output.xlsx"
    excel_exporter.export(sample_excel_file, sample_employees, output_path)

    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook.worksheets[1]

    # Find donut position column
    donut_position_col = None
    for col_idx in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, col_idx).value
        if cell_value and "Donut Exercise Position" == str(cell_value):
            donut_position_col = col_idx
            break

    assert donut_position_col is not None

    # All employee rows should have empty donut columns (empty cells have value "")
    for row_idx in range(2, sheet.max_row + 1):
        # Donut position column and the next 3 columns should be empty string (written by exporter)
        # or None (not written at all)
        donut_pos_val = sheet.cell(row_idx, donut_position_col).value
        assert donut_pos_val == "" or donut_pos_val is None
        donut_label_val = sheet.cell(row_idx, donut_position_col + 1).value
        assert donut_label_val == "" or donut_label_val is None
        donut_desc_val = sheet.cell(row_idx, donut_position_col + 2).value
        assert donut_desc_val == "" or donut_desc_val is None
        donut_notes_val = sheet.cell(row_idx, donut_position_col + 3).value
        assert donut_notes_val == "" or donut_notes_val is None
    workbook.close()  # Prevent openpyxl state pollution


# ========== Flags Tests ==========


def test_export_when_flags_exist_then_includes_flags_column(
    excel_exporter: ExcelExporter,
    sample_excel_file: Path,
    sample_employees: list[Employee],
    tmp_path: Path,
) -> None:
    """Test export includes Flags column."""
    # Add flags to first employee
    sample_employees[0].flags = ["promotion_ready", "flight_risk"]

    output_path = tmp_path / "output.xlsx"
    excel_exporter.export(sample_excel_file, sample_employees, output_path)

    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook.worksheets[1]

    # Find Flags column
    flags_col_found = False
    for col_idx in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, col_idx).value
        if cell_value and "Flags" == str(cell_value):
            flags_col_found = True
            break

    assert flags_col_found
    workbook.close()  # Prevent openpyxl state pollution


def test_export_when_flags_set_then_writes_comma_separated(
    excel_exporter: ExcelExporter,
    sample_excel_file: Path,
    sample_employees: list[Employee],
    tmp_path: Path,
) -> None:
    """Test flags are written as comma-separated list."""
    # Add flags to first employee
    sample_employees[0].flags = ["promotion_ready", "flight_risk", "new_hire"]

    output_path = tmp_path / "output.xlsx"
    excel_exporter.export(sample_excel_file, sample_employees, output_path)

    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook.worksheets[1]

    # Find Flags column
    flags_col = None
    for col_idx in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, col_idx).value
        if cell_value and "Flags" == str(cell_value):
            flags_col = col_idx
            break

    assert flags_col is not None
    # Check first employee row (row 2)
    flags_value = sheet.cell(2, flags_col).value
    assert flags_value == "promotion_ready, flight_risk, new_hire"
    workbook.close()  # Prevent openpyxl state pollution


def test_export_when_no_flags_then_empty_cell(
    excel_exporter: ExcelExporter,
    sample_excel_file: Path,
    sample_employees: list[Employee],
    tmp_path: Path,
) -> None:
    """Test employees without flags have empty Flags cell."""
    # Clear flags for all employees
    for emp in sample_employees:
        emp.flags = None

    output_path = tmp_path / "output.xlsx"
    excel_exporter.export(sample_excel_file, sample_employees, output_path)

    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook.worksheets[1]

    # Find Flags column
    flags_col = None
    for col_idx in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, col_idx).value
        if cell_value and "Flags" == str(cell_value):
            flags_col = col_idx
            break

    assert flags_col is not None

    # All employee rows should have empty flags (empty string written by exporter, or None if not written)
    for row_idx in range(2, sheet.max_row + 1):
        flags_value = sheet.cell(row_idx, flags_col).value
        assert flags_value == "" or flags_value is None
    workbook.close()  # Prevent openpyxl state pollution


def test_export_when_empty_flags_list_then_empty_cell(
    excel_exporter: ExcelExporter,
    sample_excel_file: Path,
    sample_employees: list[Employee],
    tmp_path: Path,
) -> None:
    """Test employee with empty flags list has empty cell."""
    # Set empty flags list
    sample_employees[0].flags = []

    output_path = tmp_path / "output.xlsx"
    excel_exporter.export(sample_excel_file, sample_employees, output_path)

    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook.worksheets[1]

    # Find Flags column
    flags_col = None
    for col_idx in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, col_idx).value
        if cell_value and "Flags" == str(cell_value):
            flags_col = col_idx
            break

    assert flags_col is not None
    # Check first employee row (row 2)
    flags_value = sheet.cell(2, flags_col).value
    # Excel may return None or empty string for empty cells
    assert flags_value == "" or flags_value is None
    workbook.close()  # Prevent openpyxl state pollution


def test_export_when_single_flag_then_writes_without_comma(
    excel_exporter: ExcelExporter,
    sample_excel_file: Path,
    sample_employees: list[Employee],
    tmp_path: Path,
) -> None:
    """Test single flag is written without comma."""
    # Add single flag
    sample_employees[0].flags = ["promotion_ready"]

    output_path = tmp_path / "output.xlsx"
    excel_exporter.export(sample_excel_file, sample_employees, output_path)

    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook.worksheets[1]

    # Find Flags column
    flags_col = None
    for col_idx in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, col_idx).value
        if cell_value and "Flags" == str(cell_value):
            flags_col = col_idx
            break

    assert flags_col is not None
    # Check first employee row (row 2)
    flags_value = sheet.cell(2, flags_col).value
    assert flags_value == "promotion_ready"
    workbook.close()  # Prevent openpyxl state pollution


# ========== Original Values Tracking Tests ==========


def test_export_when_employee_modified_then_tracks_original_values(
    excel_exporter: ExcelExporter,
    sample_excel_file: Path,
    sample_employees: list[Employee],
    tmp_path: Path,
) -> None:
    """Test that original Performance/Potential values are preserved in tracking columns."""
    from datetime import datetime, timezone

    from ninebox.models.events import GridMoveEvent
    from ninebox.models.session import SessionState

    # Mark first employee as modified with different performance/potential
    original_performance = sample_employees[0].performance
    original_potential = sample_employees[0].potential

    sample_employees[0].modified_in_session = True
    sample_employees[0].last_modified = datetime.now(timezone.utc)
    sample_employees[0].performance = PerformanceLevel.HIGH
    sample_employees[0].potential = PotentialLevel.HIGH

    # Create session with original employees (before changes)
    original_employees = [emp.model_copy() for emp in sample_employees]
    original_employees[0].performance = original_performance
    original_employees[0].potential = original_potential
    original_employees[0].modified_in_session = False

    session = SessionState(
        session_id="test-session",
        user_id="user1",
        created_at=datetime.now(timezone.utc),
        original_employees=original_employees,
        current_employees=sample_employees,
        original_filename="test.xlsx",
        original_file_path=str(sample_excel_file),
        sheet_name="Employee Data",
        sheet_index=1,
        events=[
            GridMoveEvent(
                employee_id=sample_employees[0].employee_id,
                employee_name=sample_employees[0].name,
                timestamp=datetime.now(timezone.utc),
                old_performance=original_performance,
                old_potential=original_potential,
                new_performance=PerformanceLevel.HIGH,
                new_potential=PotentialLevel.HIGH,
                old_position=1,
                new_position=9,
            )
        ],
    )

    output_path = tmp_path / "output.xlsx"
    excel_exporter.export(sample_excel_file, sample_employees, output_path, session=session)

    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook.worksheets[1]

    # Find Original Performance and Original Potential columns
    original_perf_col = None
    original_pot_col = None
    for col_idx in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, col_idx).value
        if cell_value == "Original Performance":
            original_perf_col = col_idx
        elif cell_value == "Original Potential":
            original_pot_col = col_idx

    assert original_perf_col is not None, "Original Performance column not found"
    assert original_pot_col is not None, "Original Potential column not found"

    # Check first employee row (row 2) has original values
    assert sheet.cell(2, original_perf_col).value == original_performance.value
    assert sheet.cell(2, original_pot_col).value == original_potential.value
    workbook.close()


def test_export_when_employee_not_modified_then_no_original_values(
    excel_exporter: ExcelExporter,
    sample_excel_file: Path,
    sample_employees: list[Employee],
    tmp_path: Path,
) -> None:
    """Test that unmodified employees have empty original value tracking columns."""
    output_path = tmp_path / "output.xlsx"
    excel_exporter.export(sample_excel_file, sample_employees, output_path)

    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook.worksheets[1]

    # Find Original Performance and Original Potential columns
    original_perf_col = None
    original_pot_col = None
    for col_idx in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, col_idx).value
        if cell_value == "Original Performance":
            original_perf_col = col_idx
        elif cell_value == "Original Potential":
            original_pot_col = col_idx

    assert original_perf_col is not None
    assert original_pot_col is not None

    # All unmodified employees should have empty original value columns
    for row_idx in range(2, sheet.max_row + 1):
        original_perf_value = sheet.cell(row_idx, original_perf_col).value
        original_pot_value = sheet.cell(row_idx, original_pot_col).value
        assert original_perf_value == "" or original_perf_value is None
        assert original_pot_value == "" or original_pot_value is None
    workbook.close()


def test_export_when_current_values_updated_then_main_columns_have_new_values(
    excel_exporter: ExcelExporter,
    sample_excel_file: Path,
    sample_employees: list[Employee],
    tmp_path: Path,
) -> None:
    """Test that main Performance/Potential columns have the NEW values after export."""
    from datetime import datetime, timezone

    from ninebox.models.session import SessionState

    # Change employee values
    original_performance = sample_employees[0].performance
    original_potential = sample_employees[0].potential

    sample_employees[0].modified_in_session = True
    sample_employees[0].last_modified = datetime.now(timezone.utc)
    sample_employees[0].performance = PerformanceLevel.LOW
    sample_employees[0].potential = PotentialLevel.HIGH

    # Create session with original employees
    original_employees = [emp.model_copy() for emp in sample_employees]
    original_employees[0].performance = original_performance
    original_employees[0].potential = original_potential
    original_employees[0].modified_in_session = False

    session = SessionState(
        session_id="test-session",
        user_id="user1",
        created_at=datetime.now(timezone.utc),
        original_employees=original_employees,
        current_employees=sample_employees,
        original_filename="test.xlsx",
        original_file_path=str(sample_excel_file),
        sheet_name="Employee Data",
        sheet_index=1,
        events=[],
    )

    output_path = tmp_path / "output.xlsx"
    excel_exporter.export(sample_excel_file, sample_employees, output_path, session=session)

    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook.worksheets[1]

    # Find main Performance and Potential columns
    perf_col = None
    pot_col = None
    for col_idx in range(1, sheet.max_column + 1):
        cell_value = str(sheet.cell(1, col_idx).value or "")
        if "Aug 2025 Talent Assessment Performance" in cell_value:
            perf_col = col_idx
        elif "Aug 2025  Talent Assessment Potential" in cell_value:
            pot_col = col_idx

    assert perf_col is not None
    assert pot_col is not None

    # Check first employee row (row 2) has NEW values in main columns
    assert sheet.cell(2, perf_col).value == "Low"
    assert sheet.cell(2, pot_col).value == "High"
    workbook.close()
