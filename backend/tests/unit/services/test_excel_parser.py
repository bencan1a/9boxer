"""Tests for Excel parser service."""

from pathlib import Path

import openpyxl
import pytest

from ninebox.models.employee import PerformanceLevel, PotentialLevel
from ninebox.models.grid_positions import calculate_grid_position, get_position_label
from ninebox.services.excel_parser import ExcelParser


pytestmark = pytest.mark.unit


def test_parse_when_valid_file_then_returns_employees(sample_excel_file: Path) -> None:
    """Test parsing a valid Excel file."""
    parser = ExcelParser()
    result = parser.parse(sample_excel_file)

    assert len(result.employees) == 5
    assert all(emp.employee_id > 0 for emp in result.employees)
    assert all(emp.name for emp in result.employees)
    # Verify metadata is populated
    assert result.metadata.sheet_name is not None
    assert result.metadata.parsed_rows == 5
    assert result.metadata.total_rows >= 5


def test_parse_when_valid_file_then_extracts_employee_data_correctly(
    sample_excel_file: Path,
) -> None:
    """Test that employee data is extracted correctly."""
    parser = ExcelParser()
    result = parser.parse(sample_excel_file)

    # Check first employee
    emp = result.employees[0]
    assert emp.employee_id == 1
    assert emp.name == "Alice Smith"
    assert emp.business_title == "Senior Engineer"
    assert emp.job_level == "MT4"
    assert emp.direct_manager == "Bob Manager"
    assert emp.performance == PerformanceLevel.HIGH
    assert emp.potential == PotentialLevel.HIGH


def test_parse_when_valid_file_then_handles_historical_ratings(sample_excel_file: Path) -> None:
    """Test that historical ratings are parsed correctly."""
    parser = ExcelParser()
    result = parser.parse(sample_excel_file)
    employees = result.employees

    # First employee has 2 ratings
    emp = employees[0]
    assert len(emp.ratings_history) == 2
    assert emp.ratings_history[0].year == 2023
    assert emp.ratings_history[0].rating == "Strong"
    assert emp.ratings_history[1].year == 2024
    assert emp.ratings_history[1].rating == "Leading"

    # Second employee has 1 rating
    emp = employees[1]
    assert len(emp.ratings_history) == 1
    assert emp.ratings_history[0].year == 2024


def test_parse_when_valid_file_then_calculates_grid_positions_correctly(
    sample_excel_file: Path,
) -> None:
    """Test grid position calculation."""
    parser = ExcelParser()
    result = parser.parse(sample_excel_file)
    employees = result.employees

    # H,H = High Performance (3), High Potential (6) = 9
    assert employees[0].grid_position == 9
    # M,M = Medium Performance (2), Medium Potential (3) = 5
    assert employees[1].grid_position == 5
    # L,H = Low Performance (1), High Potential (6) = 7
    assert employees[2].grid_position == 7
    # H,M = High Performance (3), Medium Potential (3) = 6
    assert employees[3].grid_position == 6
    # M,H = Medium Performance (2), High Potential (6) = 8
    assert employees[4].grid_position == 8


def test_parse_when_valid_file_then_handles_optional_fields_gracefully(tmp_path: Path) -> None:
    """Test handling of missing optional fields."""
    # Create minimal Excel file
    file_path = tmp_path / "minimal.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Summary"
    data_sheet = workbook.create_sheet("Employee Data")

    headers = ["Employee ID", "Worker", "Business Title", "Job Level - Primary Position"]
    for col_idx, header in enumerate(headers, start=1):
        data_sheet.cell(1, col_idx, header)

    # Minimal employee data
    data_sheet.cell(2, 1, 1)
    data_sheet.cell(2, 2, "Test User")
    data_sheet.cell(2, 3, "Test Title")
    data_sheet.cell(2, 4, "MT1")

    workbook.save(file_path)
    workbook.close()  # Explicitly close to prevent openpyxl state pollution

    parser = ExcelParser()
    result = parser.parse(file_path)

    assert len(result.employees) == 1
    emp = result.employees[0]
    assert emp.employee_id == 1
    assert emp.name == "Test User"
    assert emp.development_focus is None
    assert emp.notes is None
    # Verify defaulted fields tracking
    assert result.metadata.defaulted_fields.get("Performance", 0) == 1
    assert result.metadata.defaulted_fields.get("Potential", 0) == 1


def test_parse_when_invalid_file_format_then_raises_error(tmp_path: Path) -> None:
    """Test error on invalid file format."""
    # Create a text file instead of Excel
    file_path = tmp_path / "invalid.txt"
    file_path.write_text("This is not an Excel file")

    parser = ExcelParser()
    with pytest.raises(ValueError, match="Failed to read Excel file"):
        parser.parse(file_path)


def test_parse_when_missing_required_columns_then_raises_error(tmp_path: Path) -> None:
    """Test error when required columns are missing."""
    file_path = tmp_path / "missing_columns.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Summary"
    data_sheet = workbook.create_sheet("Employee Data")

    # Only add some headers
    data_sheet.cell(1, 1, "Employee ID")
    data_sheet.cell(1, 2, "Worker")
    # Missing "Business Title" and "Job Level - Primary Position"

    workbook.save(file_path)
    workbook.close()  # Explicitly close to prevent openpyxl state pollution

    parser = ExcelParser()
    with pytest.raises(ValueError, match="Missing required columns"):
        parser.parse(file_path)


def test_parse_when_no_valid_employees_then_raises_error(tmp_path: Path) -> None:
    """Test error when no valid employees can be parsed."""
    file_path = tmp_path / "no_employees.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Summary"
    data_sheet = workbook.create_sheet("Employee Data")

    # Add headers but no data
    headers = ["Employee ID", "Worker", "Business Title", "Job Level - Primary Position"]
    for col_idx, header in enumerate(headers, start=1):
        data_sheet.cell(1, col_idx, header)

    workbook.save(file_path)
    workbook.close()  # Explicitly close to prevent openpyxl state pollution

    parser = ExcelParser()
    with pytest.raises(ValueError, match="No valid employees found"):
        parser.parse(file_path)


def test_calculate_position_when_all_combinations_then_returns_correct_positions() -> None:
    """Test all 9 position calculations.

    Standard 9-box grid layout:
        Performance (columns): Low=1, Medium=2, High=3
        Potential (rows): Low=1-3, Medium=4-6, High=7-9
    """
    # Low performance (column 1)
    assert calculate_grid_position(PerformanceLevel.LOW, PotentialLevel.LOW) == 1
    assert calculate_grid_position(PerformanceLevel.LOW, PotentialLevel.MEDIUM) == 4
    assert calculate_grid_position(PerformanceLevel.LOW, PotentialLevel.HIGH) == 7

    # Medium performance (column 2)
    assert calculate_grid_position(PerformanceLevel.MEDIUM, PotentialLevel.LOW) == 2
    assert calculate_grid_position(PerformanceLevel.MEDIUM, PotentialLevel.MEDIUM) == 5
    assert calculate_grid_position(PerformanceLevel.MEDIUM, PotentialLevel.HIGH) == 8

    # High performance (column 3)
    assert calculate_grid_position(PerformanceLevel.HIGH, PotentialLevel.LOW) == 3
    assert calculate_grid_position(PerformanceLevel.HIGH, PotentialLevel.MEDIUM) == 6
    assert calculate_grid_position(PerformanceLevel.HIGH, PotentialLevel.HIGH) == 9


def test_get_position_label_when_all_combinations_then_returns_correct_labels() -> None:
    """Test position label generation."""
    assert get_position_label(PerformanceLevel.HIGH, PotentialLevel.HIGH) == "Star [H,H]"
    assert get_position_label(PerformanceLevel.HIGH, PotentialLevel.MEDIUM) == "High Impact [H,M]"
    assert get_position_label(PerformanceLevel.HIGH, PotentialLevel.LOW) == "Workhorse [H,L]"
    assert get_position_label(PerformanceLevel.MEDIUM, PotentialLevel.HIGH) == "Growth [M,H]"
    assert get_position_label(PerformanceLevel.MEDIUM, PotentialLevel.MEDIUM) == "Core Talent [M,M]"
    assert get_position_label(PerformanceLevel.MEDIUM, PotentialLevel.LOW) == "Effective Pro [M,L]"
    assert get_position_label(PerformanceLevel.LOW, PotentialLevel.HIGH) == "Enigma [L,H]"
    assert get_position_label(PerformanceLevel.LOW, PotentialLevel.MEDIUM) == "Inconsistent [L,M]"
    assert get_position_label(PerformanceLevel.LOW, PotentialLevel.LOW) == "Underperformer [L,L]"


# ========== Flags Parsing Tests ==========


def test_parse_when_flags_column_exists_then_reads_flags(tmp_path: Path) -> None:
    """Test that flags are parsed from the Flags column."""
    # Create a test Excel file with Flags column
    test_file = tmp_path / "test_flags.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None

    # Headers
    sheet["A1"] = "Employee ID"
    sheet["B1"] = "Worker"
    sheet["C1"] = "Business Title"
    sheet["D1"] = "Job Level - Primary Position"
    sheet["E1"] = "Aug 2025 Talent Assessment Performance"
    sheet["F1"] = "Aug 2025  Talent Assessment Potential"
    sheet["G1"] = "Flags"

    # Data row with flags
    sheet["A2"] = 1
    sheet["B2"] = "Test Employee"
    sheet["C2"] = "Software Engineer"
    sheet["D2"] = "MT4"
    sheet["E2"] = "High"
    sheet["F2"] = "Medium"
    sheet["G2"] = "promotion_ready, flight_risk"

    workbook.save(test_file)
    workbook.close()

    # Parse and verify flags
    parser = ExcelParser()
    result = parser.parse(test_file)

    assert len(result.employees) == 1
    emp = result.employees[0]
    assert emp.flags is not None
    assert set(emp.flags) == {"promotion_ready", "flight_risk"}


def test_parse_when_flags_empty_then_no_flags(tmp_path: Path) -> None:
    """Test that empty Flags column results in None or empty list."""
    test_file = tmp_path / "test_no_flags.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None

    # Headers
    sheet["A1"] = "Employee ID"
    sheet["B1"] = "Worker"
    sheet["C1"] = "Business Title"
    sheet["D1"] = "Job Level - Primary Position"
    sheet["E1"] = "Aug 2025 Talent Assessment Performance"
    sheet["F1"] = "Aug 2025  Talent Assessment Potential"
    sheet["G1"] = "Flags"

    # Data row without flags
    sheet["A2"] = 1
    sheet["B2"] = "Test Employee"
    sheet["C2"] = "Software Engineer"
    sheet["D2"] = "MT4"
    sheet["E2"] = "High"
    sheet["F2"] = "Medium"
    sheet["G2"] = ""  # Empty flags

    workbook.save(test_file)
    workbook.close()

    parser = ExcelParser()
    result = parser.parse(test_file)

    assert len(result.employees) == 1
    emp = result.employees[0]
    assert emp.flags is None or emp.flags == []


# ========== Donut Exercise Parsing Tests ==========


def test_parse_when_donut_data_exists_then_reads_donut_placement(tmp_path: Path) -> None:
    """Test that donut exercise data is parsed from tracking columns."""
    test_file = tmp_path / "test_donut.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None

    # Headers
    sheet["A1"] = "Employee ID"
    sheet["B1"] = "Worker"
    sheet["C1"] = "Business Title"
    sheet["D1"] = "Job Level - Primary Position"
    sheet["E1"] = "Aug 2025 Talent Assessment Performance"
    sheet["F1"] = "Aug 2025  Talent Assessment Potential"
    sheet["G1"] = "Donut Exercise Position"
    sheet["H1"] = "Donut Exercise Notes"

    # Data row with donut placement
    sheet["A2"] = 1
    sheet["B2"] = "Test Employee"
    sheet["C2"] = "Software Engineer"
    sheet["D2"] = "MT4"
    sheet["E2"] = "Medium"  # Main position
    sheet["F2"] = "Medium"  # Main position
    sheet["G2"] = 9  # Donut position (High/High)
    sheet["H2"] = "Hypothetical star placement"

    workbook.save(test_file)
    workbook.close()

    parser = ExcelParser()
    result = parser.parse(test_file)

    assert len(result.employees) == 1
    emp = result.employees[0]
    assert emp.donut_modified is True
    assert emp.donut_position == 9
    assert emp.donut_performance == PerformanceLevel.HIGH
    assert emp.donut_potential == PotentialLevel.HIGH
    assert emp.donut_notes == "Hypothetical star placement"


def test_parse_when_no_donut_data_then_donut_fields_none(tmp_path: Path) -> None:
    """Test that employees without donut data have donut fields set to None/False."""
    test_file = tmp_path / "test_no_donut.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None

    # Headers
    sheet["A1"] = "Employee ID"
    sheet["B1"] = "Worker"
    sheet["C1"] = "Business Title"
    sheet["D1"] = "Job Level - Primary Position"
    sheet["E1"] = "Aug 2025 Talent Assessment Performance"
    sheet["F1"] = "Aug 2025  Talent Assessment Potential"
    sheet["G1"] = "Donut Exercise Position"

    # Data row without donut placement
    sheet["A2"] = 1
    sheet["B2"] = "Test Employee"
    sheet["C2"] = "Software Engineer"
    sheet["D2"] = "MT4"
    sheet["E2"] = "Medium"
    sheet["F2"] = "Medium"
    sheet["G2"] = ""  # Empty donut position

    workbook.save(test_file)
    workbook.close()

    parser = ExcelParser()
    result = parser.parse(test_file)

    assert len(result.employees) == 1
    emp = result.employees[0]
    assert emp.donut_modified is False
    assert emp.donut_position is None
    assert emp.donut_performance is None
    assert emp.donut_potential is None
    assert emp.donut_notes is None


def _write_history_sheet(path: Path, history_headers: list[str], values: list[str | None]) -> None:
    """Write a minimal one-employee sheet with the given history columns."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Employee Data"

    base = [
        "Employee ID",
        "Worker",
        "Business Title",
        "Job Level - Primary Position",
        "Aug 2025 Talent Assessment Performance",
        "Aug 2025  Talent Assessment Potential",
    ]
    for col, header in enumerate(base + history_headers, start=1):
        sheet.cell(1, col, header)
    for col, value in enumerate(
        [1, "Test Employee", "Engineer", "MT4", "High", "High", *values], start=1
    ):
        sheet.cell(2, col, value)

    workbook.save(path)
    workbook.close()


def test_parse_when_history_columns_for_any_year_then_all_are_imported(tmp_path: Path) -> None:
    """Review years beyond 2023/2024 must be imported, not silently dropped."""
    test_file = tmp_path / "any_year_history.xlsx"
    _write_history_sheet(
        test_file,
        [
            "2022 Completed Performance Rating",
            "2025 Completed Performance Rating",
            "2026 Completed Performance Rating",
        ],
        ["Solid", "Strong", "Leading"],
    )

    result = ExcelParser().parse(test_file)

    assert [(r.year, r.rating) for r in result.employees[0].ratings_history] == [
        (2022, "Solid"),
        (2025, "Strong"),
        (2026, "Leading"),
    ]


def test_parse_when_history_columns_out_of_order_then_sorted_ascending(tmp_path: Path) -> None:
    """History is returned oldest-first regardless of column order in the file."""
    test_file = tmp_path / "unordered_history.xlsx"
    _write_history_sheet(
        test_file,
        [
            "2026 Completed Performance Rating",
            "2023 Completed Performance Rating",
            "2025 Completed Performance Rating",
        ],
        ["Leading", "Solid", "Strong"],
    )

    result = ExcelParser().parse(test_file)

    assert [r.year for r in result.employees[0].ratings_history] == [2023, 2025, 2026]


def test_parse_when_history_header_lowercase_then_still_imported(tmp_path: Path) -> None:
    """Year-prefixed history headers are matched case-insensitively."""
    test_file = tmp_path / "lowercase_history.xlsx"
    _write_history_sheet(test_file, ["2026 completed performance rating"], ["Leading"])

    result = ExcelParser().parse(test_file)

    assert [(r.year, r.rating) for r in result.employees[0].ratings_history] == [(2026, "Leading")]


def test_parse_when_history_cell_blank_then_year_omitted(tmp_path: Path) -> None:
    """Empty history cells produce no timeline entry for that year."""
    test_file = tmp_path / "blank_history.xlsx"
    _write_history_sheet(
        test_file,
        [
            "2025 Completed Performance Rating",
            "2026 Completed Performance Rating",
        ],
        [None, "   "],
    )

    result = ExcelParser().parse(test_file)

    assert result.employees[0].ratings_history == []


def _write_prior_position_sheet(path: Path, value: object) -> None:
    """Write a one-employee sheet carrying the given prior calibration cell."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Employee Data"

    headers = [
        "Employee ID",
        "Worker",
        "Business Title",
        "Job Level - Primary Position",
        "Aug 2025 Talent Assessment Performance",
        "Aug 2025  Talent Assessment Potential",
        "Prior Calibration 9-Box Label",
    ]
    for col, header in enumerate(headers, start=1):
        sheet.cell(1, col, header)
    for col, cell in enumerate(
        [1, "Test Employee", "Engineer", "MT4", "High", "High", value], start=1
    ):
        sheet.cell(2, col, cell)

    workbook.save(path)
    workbook.close()


def test_parse_when_prior_calibration_column_present_then_position_imported(
    tmp_path: Path,
) -> None:
    """The previous calibration's box is read for movement comparison."""
    test_file = tmp_path / "prior_position.xlsx"
    _write_prior_position_sheet(test_file, 3)

    result = ExcelParser().parse(test_file)

    assert result.employees[0].prior_grid_position == 3


def test_parse_when_prior_calibration_column_absent_then_none(tmp_path: Path) -> None:
    """Files without the column parse fine and simply have no prior position."""
    test_file = tmp_path / "no_prior.xlsx"
    _write_history_sheet(test_file, [], [])

    result = ExcelParser().parse(test_file)

    assert result.employees[0].prior_grid_position is None


@pytest.mark.parametrize("value", [0, 10, "Top Talent", ""])
def test_parse_when_prior_calibration_invalid_then_none(tmp_path: Path, value: object) -> None:
    """Out-of-range or non-numeric prior positions are ignored, not guessed at."""
    test_file = tmp_path / "bad_prior.xlsx"
    _write_prior_position_sheet(test_file, value)

    result = ExcelParser().parse(test_file)

    assert result.employees[0].prior_grid_position is None


def test_parse_when_talent_indicator_looks_like_a_box_then_not_used_as_prior(
    tmp_path: Path,
) -> None:
    """A talent indicator must not be inferred as the prior calibration position.

    The column's meaning varies by source file - it may hold an older, newer or
    parallel assessment - so guessing would produce confident but wrong flags.
    """
    test_file = tmp_path / "indicator_only.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Employee Data"
    headers = [
        "Employee ID",
        "Worker",
        "Business Title",
        "Job Level - Primary Position",
        "Aug 2025 Talent Assessment Performance",
        "Aug 2025  Talent Assessment Potential",
        "FY25 Talent Indicator",
    ]
    for col, header in enumerate(headers, start=1):
        sheet.cell(1, col, header)
    for col, cell in enumerate(
        [1, "Test Employee", "Engineer", "MT4", "High", "High", "3. Expert Talent"], start=1
    ):
        sheet.cell(2, col, cell)
    workbook.save(test_file)
    workbook.close()

    result = ExcelParser().parse(test_file)

    assert result.employees[0].talent_indicator == "3. Expert Talent"
    assert result.employees[0].prior_grid_position is None
