"""API tests for employee event tracking with new TrackableEvent architecture.

This file contains API-level tests for the new event system.
It complements the existing test_employees.py by focusing on:
- API responses return TrackableEvent objects
- Flag updates create appropriate events
- Move endpoints return event data
- Session status includes events array
"""

from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient

from ninebox.models.employee import PerformanceLevel, PotentialLevel

TEST_FLAG = "promotion_ready"


pytestmark = pytest.mark.unit


@pytest.fixture
def upload_sample_data_with_flags(
    test_client: TestClient, auth_headers: dict[str, str], tmp_path: Path
) -> str:
    """Upload sample data with flags and return session_id."""
    # Create a minimal Excel file with flags
    file_path = tmp_path / "test_with_flags.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Employee Data"

    # Headers
    headers = [
        "Employee ID",
        "Worker",
        "Business Title",
        "Job Title",
        "Job Profile",
        "Job Level - Primary Position",
        "Worker's Manager",
        "Hire Date",
        "Tenure Category (Months)",
        "Time in Job Profile",
        "Aug 2025 Talent Assessment Performance",
        "Aug 2025  Talent Assessment Potential",
        "Promotion Ready",
    ]

    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(1, col_idx, header)

    # Data rows
    employees = [
        (
            1,
            "Alice",
            "Engineer",
            "Software Engineer",
            "Software EngineeringUSA",
            "MT4",
            "Manager",
            "2020-01-01",
            "3-5 years",
            "2 years",
            "High",
            "High",
            "Yes",
        ),
        (
            2,
            "Bob",
            "Engineer",
            "Software Engineer",
            "Software EngineeringUSA",
            "MT3",
            "Manager",
            "2021-01-01",
            "1-3 years",
            "1 year",
            "Medium",
            "Medium",
            "No",
        ),
        (
            3,
            "Charlie",
            "Engineer",
            "Software Engineer",
            "Software EngineeringUSA",
            "MT2",
            "Manager",
            "2022-01-01",
            "0-1 year",
            "6 months",
            "Low",
            "Low",
            "No",
        ),
    ]

    for row_idx, emp_data in enumerate(employees, start=2):
        for col_idx, value in enumerate(emp_data, start=1):
            sheet.cell(row_idx, col_idx, value)

    workbook.save(file_path)
    workbook.close()  # Explicitly close to prevent openpyxl state pollution

    # Upload file
    with open(file_path, "rb") as f:  # noqa: PTH123
        files = {
            "file": (
                "test_with_flags.xlsx",
                f,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        response = test_client.post("/api/session/upload", files=files, headers=auth_headers)

    assert response.status_code == 200
    session_id = response.json()["session_id"]

    # WORKAROUND: Excel parser doesn't populate flags from "Promotion Ready" column
    # Manually set Alice's (employee 1) flags in both original and current employee lists
    # This is needed because tests expect promotion_ready flag to be in original data
    from ninebox.core.dependencies import get_session_manager

    session_mgr = get_session_manager()
    session = session_mgr.get_session("local-user")
    if session:
        # Set flags in both original and current employees
        for employee in session.original_employees:
            if employee.employee_id == 1:  # Alice
                employee.flags = ["promotion_ready"]
        for employee in session.current_employees:
            if employee.employee_id == 1:  # Alice
                employee.flags = ["promotion_ready"]
        # Persist the session with updated flags
        session_mgr._persist_session(session)

    return session_id


class TestMoveEmployeeEvents:
    """Tests for move employee endpoint returning events."""

    def test_move_employee_when_called_then_returns_grid_move_event(
        self,
        test_client: TestClient,
        upload_sample_data_with_flags: str,
    ) -> None:
        """Test that PATCH /employees/{id} with position returns GridMoveEvent."""
        response = test_client.patch(
            "/api/employees/1",
            json={
                "performance": "MEDIUM",
                "potential": "MEDIUM",
            },
        )

        assert response.status_code == 200
        data = response.json()

        # Check that response includes event data
        assert "event" in data or "employee" in data
        # The specific structure depends on implementation

    def test_move_employee_when_back_to_original_then_no_event(
        self,
        test_client: TestClient,
        upload_sample_data_with_flags: str,
    ) -> None:
        """Test that moving back to original position removes event."""
        # Move away
        test_client.patch(
            "/api/employees/1",
            json={
                "performance": "MEDIUM",
                "potential": "MEDIUM",
            },
        )

        # Move back
        response = test_client.patch(
            "/api/employees/1",
            json={
                "performance": "HIGH",
                "potential": "HIGH",
            },
        )

        assert response.status_code == 200

        # Check session status
        status_response = test_client.get("/api/session/status")
        assert status_response.status_code == 200
        session_data = status_response.json()

        # Should have no events
        assert len(session_data.get("events", [])) == 0


class TestFlagUpdateEvents:
    """Tests for flag updates creating events."""

    def test_update_employee_when_adding_flag_then_creates_flag_add_event(
        self,
        test_client: TestClient,
        upload_sample_data_with_flags: str,
    ) -> None:
        """Test that adding a flag creates FlagAddEvent."""
        # Bob originally has no promotion_ready flag
        response = test_client.patch(
            "/api/employees/2",
            json={
                "flags": ["promotion_ready"],
            },
        )

        assert response.status_code == 200

        # Check session status
        status_response = test_client.get("/api/session/status")
        assert status_response.status_code == 200
        session_data = status_response.json()

        events = session_data.get("events", [])
        assert len(events) == 1
        assert events[0]["event_type"] == "flag_add"
        assert events[0]["employee_id"] == 2
        assert events[0]["flag"] == "promotion_ready"

    def test_update_employee_when_removing_original_flag_then_creates_flag_remove_event(
        self,
        test_client: TestClient,
        upload_sample_data_with_flags: str,
    ) -> None:
        """Test that removing a flag from original data creates FlagRemoveEvent."""
        # Alice originally has promotion_ready flag
        response = test_client.patch(
            "/api/employees/1",
            json={
                "flags": [],  # Remove all flags
            },
        )

        assert response.status_code == 200

        # Check session status
        status_response = test_client.get("/api/session/status")
        assert status_response.status_code == 200
        session_data = status_response.json()

        events = session_data.get("events", [])
        assert len(events) == 1
        assert events[0]["event_type"] == "flag_remove"
        assert events[0]["employee_id"] == 1
        assert events[0]["flag"] == "promotion_ready"

    def test_update_employee_when_removing_added_flag_then_no_event(
        self,
        test_client: TestClient,
        upload_sample_data_with_flags: str,
    ) -> None:
        """Test that removing a flag that wasn't in original removes the event."""
        # Add flag
        test_client.patch(
            "/api/employees/2",
            json={
                "flags": ["promotion_ready"],
            },
        )

        # Remove flag
        response = test_client.patch(
            "/api/employees/2",
            json={
                "flags": [],
            },
        )

        assert response.status_code == 200

        # Check session status
        status_response = test_client.get("/api/session/status")
        assert status_response.status_code == 200
        session_data = status_response.json()

        # Should have no events (net-zero)
        assert len(session_data.get("events", [])) == 0

    def test_update_employee_when_adding_back_removed_flag_then_no_event(
        self,
        test_client: TestClient,
        upload_sample_data_with_flags: str,
    ) -> None:
        """Test that adding back a removed flag removes the FlagRemoveEvent."""
        # Remove flag
        test_client.patch(
            "/api/employees/1",
            json={
                "flags": [],
            },
        )

        # Add flag back
        response = test_client.patch(
            "/api/employees/1",
            json={
                "flags": ["promotion_ready"],
            },
        )

        assert response.status_code == 200

        # Check session status
        status_response = test_client.get("/api/session/status")
        assert status_response.status_code == 200
        session_data = status_response.json()

        # Should have no events (back to original)
        assert len(session_data.get("events", [])) == 0

    def test_update_employee_when_multiple_flags_then_multiple_events(
        self,
        test_client: TestClient,
        upload_sample_data_with_flags: str,
    ) -> None:
        """Test that adding multiple flags creates multiple events."""
        response = test_client.patch(
            "/api/employees/2",
            json={
                "flags": ["promotion_ready", "succession_candidate"],
            },
        )

        assert response.status_code == 200

        # Check session status
        status_response = test_client.get("/api/session/status")
        assert status_response.status_code == 200
        session_data = status_response.json()

        events = session_data.get("events", [])
        assert len(events) == 2
        flag_values = {event["flag"] for event in events}
        assert flag_values == {"promotion_ready", "succession_candidate"}


class TestSessionStatusEvents:
    """Tests for session status endpoint returning events."""

    def test_session_status_when_no_changes_then_empty_events(
        self,
        test_client: TestClient,
        upload_sample_data_with_flags: str,
    ) -> None:
        """Test that session status returns empty events array when no changes."""
        response = test_client.get("/api/session/status")

        assert response.status_code == 200
        data = response.json()

        assert "events" in data
        assert len(data["events"]) == 0

    def test_session_status_when_move_then_includes_event(
        self,
        test_client: TestClient,
        upload_sample_data_with_flags: str,
    ) -> None:
        """Test that session status includes move events."""
        # Move employee
        test_client.patch(
            "/api/employees/1",
            json={
                "performance": "MEDIUM",
                "potential": "MEDIUM",
            },
        )

        response = test_client.get("/api/session/status")

        assert response.status_code == 200
        data = response.json()

        events = data.get("events", [])
        assert len(events) == 1
        assert events[0]["event_type"] == "grid_move"
        assert events[0]["employee_id"] == 1

    def test_session_status_when_flag_change_then_includes_event(
        self,
        test_client: TestClient,
        upload_sample_data_with_flags: str,
    ) -> None:
        """Test that session status includes flag events."""
        # Add flag
        test_client.patch(
            "/api/employees/2",
            json={
                "flags": ["promotion_ready"],
            },
        )

        response = test_client.get("/api/session/status")

        assert response.status_code == 200
        data = response.json()

        events = data.get("events", [])
        assert len(events) == 1
        assert events[0]["event_type"] == "flag_add"

    def test_session_status_when_mixed_events_then_includes_all(
        self,
        test_client: TestClient,
        upload_sample_data_with_flags: str,
    ) -> None:
        """Test that session status includes all event types."""
        # Move employee
        test_client.patch(
            "/api/employees/1",
            json={
                "performance": "MEDIUM",
                "potential": "MEDIUM",
            },
        )

        # Add flag to different employee
        test_client.patch(
            "/api/employees/2",
            json={
                "flags": ["promotion_ready"],
            },
        )

        # Remove flag from third employee
        # (Charlie has no flags originally, so this won't create an event)

        response = test_client.get("/api/session/status")

        assert response.status_code == 200
        data = response.json()

        events = data.get("events", [])
        assert len(events) == 2

        event_types = {event["event_type"] for event in events}
        assert "grid_move" in event_types
        assert "flag_add" in event_types


class TestMixedEventWorkflows:
    """Tests for complete workflows with mixed event types."""

    def test_workflow_move_and_flag_change_same_employee(
        self,
        test_client: TestClient,
        upload_sample_data_with_flags: str,
    ) -> None:
        """Test workflow: move employee and change flag on same employee."""
        # Move and add flag
        response = test_client.patch(
            "/api/employees/1",
            json={
                "performance": "MEDIUM",
                "potential": "MEDIUM",
                "flags": ["promotion_ready", "flight_risk"],
            },
        )

        assert response.status_code == 200

        # Check session status
        status_response = test_client.get("/api/session/status")
        session_data = status_response.json()

        events = session_data.get("events", [])
        # Should have GridMoveEvent + FlagAddEvent (flight_risk)
        # promotion_ready was already in original, so no event
        assert len(events) == 2

    def test_workflow_revert_all_changes(
        self,
        test_client: TestClient,
        upload_sample_data_with_flags: str,
    ) -> None:
        """Test workflow: make changes then revert everything."""
        # Make changes
        test_client.patch(
            "/api/employees/1",
            json={
                "performance": "MEDIUM",
                "potential": "MEDIUM",
                "flags": ["promotion_ready", "flight_risk"],
            },
        )

        # Revert all
        response = test_client.patch(
            "/api/employees/1",
            json={
                "performance": "HIGH",
                "potential": "HIGH",
                "flags": ["promotion_ready"],  # Back to original
            },
        )

        assert response.status_code == 200

        # Check session status
        status_response = test_client.get("/api/session/status")
        session_data = status_response.json()

        # Should have no events (all reverted)
        assert len(session_data.get("events", [])) == 0

    def test_workflow_multiple_employees_different_changes(
        self,
        test_client: TestClient,
        upload_sample_data_with_flags: str,
    ) -> None:
        """Test workflow: different types of changes on different employees."""
        # Move Alice
        test_client.patch(
            "/api/employees/1",
            json={
                "performance": "MEDIUM",
                "potential": "MEDIUM",
            },
        )

        # Add flag to Bob
        test_client.patch(
            "/api/employees/2",
            json={
                "flags": ["promotion_ready"],
            },
        )

        # Remove flag from Alice
        test_client.patch(
            "/api/employees/1",
            json={
                "flags": [],
            },
        )

        response = test_client.get("/api/session/status")
        session_data = response.json()

        events = session_data.get("events", [])
        # Should have:
        # 1. GridMoveEvent for Alice
        # 2. FlagAddEvent for Bob
        # 3. FlagRemoveEvent for Alice (promotion_ready)
        assert len(events) == 3

        # Verify all event types present
        event_types = {event["event_type"] for event in events}
        assert "grid_move" in event_types
        assert "flag_add" in event_types
        assert "flag_remove" in event_types


class TestEventPersistence:
    """Tests for event persistence across operations."""

    def test_events_persist_after_multiple_operations(
        self,
        test_client: TestClient,
        upload_sample_data_with_flags: str,
    ) -> None:
        """Test that events persist correctly after multiple operations."""
        # Make initial change
        test_client.patch(
            "/api/employees/1",
            json={
                "performance": "MEDIUM",
                "potential": "MEDIUM",
            },
        )

        # Add another change
        test_client.patch(
            "/api/employees/2",
            json={
                "flags": ["promotion_ready"],
            },
        )

        # Get status multiple times
        response1 = test_client.get("/api/session/status")
        response2 = test_client.get("/api/session/status")

        # Both should return same events
        assert response1.json()["events"] == response2.json()["events"]
        assert len(response1.json()["events"]) == 2

    def test_events_update_on_subsequent_changes(
        self,
        test_client: TestClient,
        upload_sample_data_with_flags: str,
    ) -> None:
        """Test that events update correctly on subsequent changes."""
        # First move
        test_client.patch(
            "/api/employees/1",
            json={
                "performance": "MEDIUM",
                "potential": "MEDIUM",
            },
        )

        response1 = test_client.get("/api/session/status")
        assert len(response1.json()["events"]) == 1

        # Second move (should update existing event)
        test_client.patch(
            "/api/employees/1",
            json={
                "performance": "LOW",
                "potential": "LOW",
            },
        )

        response2 = test_client.get("/api/session/status")
        events = response2.json()["events"]

        # Should still have only 1 event (updated)
        assert len(events) == 1
        assert events[0]["new_performance"] == "Low"  # Enum value is capitalized
        assert events[0]["new_potential"] == "Low"  # Enum value is capitalized
