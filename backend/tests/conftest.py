"""Root conftest for all tests."""

import os
import tempfile
from collections.abc import Generator
from datetime import date, datetime, timezone
from pathlib import Path
from unittest.mock import Mock

import openpyxl
import pytest
from fastapi.testclient import TestClient

from ninebox.models.constants import ALLOWED_FLAGS
from ninebox.models.employee import Employee, HistoricalRating, PerformanceLevel, PotentialLevel
from ninebox.models.grid_positions import get_position_label_by_number


def create_test_employee(
    employee_id: int = 1,
    name: str = "Alice",
    performance: PerformanceLevel = PerformanceLevel.MEDIUM,
    potential: PotentialLevel = PotentialLevel.MEDIUM,
    flags: list[str] | None = None,
    grid_position: int = 5,
) -> Employee:
    """Create a minimal Employee for testing with all required fields.

    DEPRECATED (2025-12-28): This function is an alias for create_simple_test_employee()
    and is kept for backward compatibility with existing imports.

    Migration Guide:
    - For single-employee unit tests → Use create_simple_test_employee()
    - For integration tests with multiple employees → Use sample_employees fixture
    - For larger datasets (100+) → Use rich_sample_employees_medium/large fixtures
    - For intelligence/statistics tests → Use rich_sample_employees_* (includes bias patterns)

    Args:
        employee_id: Employee ID (default: 1)
        name: Employee name (default: "Alice")
        performance: Performance level (default: MEDIUM)
        potential: Potential level (default: MEDIUM)
        flags: Employee flags, must be from ALLOWED_FLAGS (default: None)
        grid_position: Grid position 1-9 (default: 5)

    Returns:
        Employee object with all required fields populated
    """
    return Employee(
        employee_id=employee_id,
        name=name,
        business_title="Test Title",
        job_title="Test Job",
        job_profile="Test ProfileUSA",
        job_level="MT4",
        job_function="Other",
        location="USA",
        direct_manager="Test Manager",
        hire_date=date(2020, 1, 1),
        tenure_category="3-5 years",
        time_in_job_profile="2 years",
        performance=performance,
        potential=potential,
        grid_position=grid_position,
        talent_indicator="Solid Contributor",
        flags=flags,
    )


def create_simple_test_employee(
    employee_id: int = 1,
    name: str = "Alice",
    performance: PerformanceLevel = PerformanceLevel.MEDIUM,
    potential: PotentialLevel = PotentialLevel.MEDIUM,
    flags: list[str] | None = None,
    grid_position: int = 5,
) -> Employee:
    """Create a minimal Employee for testing with all required fields.

    Use this function when you need:
    - Single employee instances for unit tests
    - Full control over employee attributes
    - Minimal test fixtures without organizational hierarchy

    For integration tests or tests requiring multiple employees,
    use the sample_employees fixture instead.

    Args:
        employee_id: Employee ID (default: 1)
        name: Employee name (default: "Alice")
        performance: Performance level (default: MEDIUM)
        potential: Potential level (default: MEDIUM)
        flags: Employee flags, must be from ALLOWED_FLAGS (default: None)
        grid_position: Grid position 1-9 (default: 5)

    Returns:
        Employee object with all required fields populated

    Example:
        >>> employee = create_simple_test_employee(employee_id=99, name="Test User")
        >>> employee.employee_id
        99
    """
    return Employee(
        employee_id=employee_id,
        name=name,
        business_title="Test Title",
        job_title="Test Job",
        job_profile="Test ProfileUSA",
        job_level="MT4",
        job_function="Other",
        location="USA",
        direct_manager="Test Manager",
        hire_date=date(2020, 1, 1),
        tenure_category="3-5 years",
        time_in_job_profile="2 years",
        performance=performance,
        potential=potential,
        grid_position=grid_position,
        talent_indicator="Solid Contributor",
        flags=flags,
    )


@pytest.fixture(scope="session")
def test_db_path(request: pytest.FixtureRequest) -> Generator[str, None, None]:
    """Create a temporary database for testing.

    Supports pytest-xdist by creating separate database files per worker.
    """
    # Get worker ID for parallel test execution (pytest-xdist)
    worker_id = getattr(request.config, "workerinput", {}).get("workerid", "master")

    # Create a temporary directory for this worker's database
    temp_dir = tempfile.mkdtemp(suffix=f"_{worker_id}_ninebox_test")
    db_path = os.path.join(temp_dir, "ninebox.db")

    # Save original environment variables so we can restore them
    # This is critical for VSCode pytest extension communication
    original_app_data_dir = os.environ.get("APP_DATA_DIR")
    original_database_path = os.environ.get("DATABASE_PATH")

    # Set APP_DATA_DIR so get_user_data_dir() returns our test directory
    # This ensures DatabaseManager uses the test database
    os.environ["APP_DATA_DIR"] = temp_dir
    # Also set DATABASE_PATH for any legacy code that might use it
    os.environ["DATABASE_PATH"] = db_path

    yield db_path

    # Restore original environment variables FIRST (important for VSCode extension)
    if original_app_data_dir is None:
        os.environ.pop("APP_DATA_DIR", None)
    else:
        os.environ["APP_DATA_DIR"] = original_app_data_dir

    if original_database_path is None:
        os.environ.pop("DATABASE_PATH", None)
    else:
        os.environ["DATABASE_PATH"] = original_database_path

    # Then cleanup files
    try:
        if os.path.exists(db_path):  # noqa: PTH110
            os.unlink(db_path)  # noqa: PTH108
    except Exception:
        pass  # Ignore cleanup errors

    try:
        if os.path.exists(temp_dir):  # noqa: PTH110
            import shutil  # noqa: PLC0415

            shutil.rmtree(temp_dir)  # Remove directory and any remaining contents
    except Exception:
        pass  # Ignore cleanup errors


@pytest.fixture(autouse=True)
def setup_test_db(request: pytest.FixtureRequest) -> Generator[None, None, None]:
    """Clean up in-memory state and database before and after each test.

    This fixture clears dependency injection caches, in-memory state,
    and database sessions to ensure tests start with a clean slate.

    Note: Integration tests use transaction-based isolation (see integration/conftest.py),
    so this database cleanup only affects unit tests.
    """
    import gc  # noqa: PLC0415

    # BEFORE test: ensure clean state for xdist robustness
    from ninebox.core.dependencies import get_session_manager, get_db_manager  # noqa: PLC0415
    from ninebox.services.database import DatabaseManager  # noqa: PLC0415

    # Check if this is an integration test (skip database cleanup for integration tests)
    # Integration tests use transaction-based isolation which is more robust
    is_integration_test = "integration" in str(request.node.path)

    # Clear database FIRST before test starts (unit tests only)
    if not is_integration_test:
        try:
            temp_db = DatabaseManager()
            with temp_db.get_connection() as conn:
                # Check if tables exist before trying to delete (avoids "no such table" errors)
                cursor = conn.execute("SELECT name FROM sqlite_master WHERE type='table'")
                tables = {row[0] for row in cursor.fetchall()}

                # Clear all tables to ensure full isolation (only if they exist)
                if "historical_ratings" in tables:
                    conn.execute("DELETE FROM historical_ratings")
                if "employees" in tables:
                    conn.execute("DELETE FROM employees")
                if "sessions" in tables:
                    conn.execute("DELETE FROM sessions")
                # No manual commit needed - context manager commits automatically
        except Exception:
            # Ignore errors during cleanup (e.g., if database doesn't exist yet)
            pass

    # Clear caches BEFORE test to ensure fresh instances
    try:
        get_session_manager.cache_clear()
        get_db_manager.cache_clear()
    except Exception:
        pass

    # Clear session manager state
    try:
        mgr = get_session_manager()
        mgr._sessions_loaded = False
        mgr._sessions.clear()
    except Exception:
        pass

    # Force garbage collection before test
    try:
        gc.collect()
    except Exception:
        pass

    yield

    # AFTER test: Clean up openpyxl state to prevent NumberFormat pollution
    try:
        # Force garbage collection to clean up any lingering workbook references
        # This helps prevent openpyxl's NumberFormat registry from getting polluted
        gc.collect()
    except Exception:
        pass

    # AFTER test: Clean up database FIRST, then in-memory state
    # This order is critical: database → memory → caches
    # Clear database sessions (for unit tests that use test_client)
    # Integration tests handle this via transaction rollback (see integration/conftest.py)
    # Skip cleanup for integration tests to avoid interfering with transaction isolation
    from ninebox.services.database import DatabaseManager  # noqa: PLC0415

    try:
        # Check if DatabaseManager.get_connection is patched (integration tests patch this)
        # If patched, skip cleanup since transaction rollback handles it
        is_patched = hasattr(DatabaseManager.get_connection, '__wrapped__') or \
                     hasattr(DatabaseManager.get_connection, '_mock_name')

        if not is_patched:
            # Only clean up for unit tests (not patched)
            temp_db = DatabaseManager()
            with temp_db.get_connection() as conn:
                # Clear all tables to ensure full isolation
                conn.execute("DELETE FROM historical_ratings")
                conn.execute("DELETE FROM employees")
                conn.execute("DELETE FROM sessions")
                conn.commit()  # Ensure changes are committed
    except Exception:
        # Ignore errors during cleanup (e.g., if database doesn't exist)
        pass

    # AFTER database cleanup: Clean up in-memory state from dependency injection cache
    # Reset session manager's in-memory state
    try:
        mgr = get_session_manager()
        mgr._sessions_loaded = False  # Reset lazy loading flag
        mgr._sessions.clear()  # Clear in-memory session cache
    except Exception:
        pass  # If manager doesn't exist yet, that's fine

    # Clear the lru_cache to get fresh instances for next test
    try:
        get_session_manager.cache_clear()
        get_db_manager.cache_clear()
    except Exception:
        pass

    # Final garbage collection after all cleanup
    try:
        gc.collect()
    except Exception:
        pass


@pytest.fixture(autouse=True, scope="function")
def mock_preferences_manager(monkeypatch: pytest.MonkeyPatch) -> Mock:
    """Mock PreferencesManager to prevent race conditions in parallel tests.

    This fixture automatically applies to all tests and mocks the PreferencesManager
    to prevent file system race conditions when tests run in parallel with pytest-xdist.

    The mock returns empty lists for get_recent_files() and None for add_recent_file(),
    allowing tests to focus on API behavior without preferences side effects.

    Args:
        monkeypatch: pytest monkeypatch fixture for patching

    Returns:
        Mock object configured to replace PreferencesManager
    """
    # Create a mock instance
    mock_instance = Mock()
    mock_instance.add_recent_file.return_value = None
    mock_instance.get_recent_files.return_value = []
    mock_instance.get_config.return_value = {}
    mock_instance.update_config.return_value = None

    # Create a mock class that returns our mock instance
    mock_class = Mock(return_value=mock_instance)

    # Patch the PreferencesManager class itself
    monkeypatch.setattr("ninebox.services.preferences_manager.PreferencesManager", mock_class)
    monkeypatch.setattr("ninebox.core.dependencies.PreferencesManager", mock_class)

    # Also clear the lru_cache to ensure fresh instances
    from ninebox.core.dependencies import get_preferences_manager  # noqa: PLC0415

    get_preferences_manager.cache_clear()

    return mock_instance


@pytest.fixture
def sample_employees() -> list[Employee]:
    """Create sample employee data for testing (5 employees).

    Returns a fixed set of 5 employees with known attributes for testing.
    For larger datasets, use rich_sample_employees_small/medium/large fixtures.
    """
    return [
        Employee(
            employee_id=1,
            name="Alice Smith",
            business_title="Senior Engineer",
            job_title="Software Engineer",
            job_profile="Engineering",
            job_level="MT4",
            job_function="Engineering",
            location="USA",
            direct_manager="Bob Manager",
            hire_date=date(2020, 1, 15),
            tenure_category="3-5 years",
            time_in_job_profile="2 years",
            performance=PerformanceLevel.HIGH,
            potential=PotentialLevel.HIGH,
            grid_position=9,
            talent_indicator="High Potential",
            ratings_history=[
                HistoricalRating(year=2023, rating="Strong"),
                HistoricalRating(year=2024, rating="Leading"),
            ],
        ),
        Employee(
            employee_id=2,
            name="Bob Jones",
            business_title="Product Manager",
            job_title="Product Manager",
            job_profile="Product",
            job_level="MT4",
            job_function="Product Manager",
            location="USA",
            direct_manager="Bob Manager",
            hire_date=date(2019, 6, 1),
            tenure_category="5+ years",
            time_in_job_profile="3 years",
            performance=PerformanceLevel.MEDIUM,
            potential=PotentialLevel.MEDIUM,
            grid_position=5,
            talent_indicator="Solid Contributor",
            ratings_history=[
                HistoricalRating(year=2024, rating="Solid"),
            ],
        ),
        Employee(
            employee_id=3,
            name="Charlie Brown",
            business_title="Sales Rep",
            job_title="Sales Representative",
            job_profile="Sales",
            job_level="MT2",
            job_function="Sales",
            location="CAN",
            direct_manager="Alice Manager",
            hire_date=date(2021, 3, 10),
            tenure_category="1-3 years",
            time_in_job_profile="1 year",
            performance=PerformanceLevel.LOW,
            potential=PotentialLevel.HIGH,
            grid_position=7,
            talent_indicator="Development Needed",
        ),
        Employee(
            employee_id=4,
            name="Diana Prince",
            business_title="Marketing Lead",
            job_title="Marketing Manager",
            job_profile="Marketing",
            job_level="MT3",
            job_function="Marketing",
            location="USA",
            direct_manager="Bob Manager",
            hire_date=date(2018, 9, 20),
            tenure_category="5+ years",
            time_in_job_profile="4 years",
            performance=PerformanceLevel.HIGH,
            potential=PotentialLevel.MEDIUM,
            grid_position=6,
            talent_indicator="Solid Contributor",
        ),
        Employee(
            employee_id=5,
            name="Eve Davis",
            business_title="Junior Developer",
            job_title="Software Developer",
            job_profile="Engineering",
            job_level="MT1",
            job_function="Engineering",
            location="CAN",
            direct_manager="Alice Manager",
            hire_date=date(2022, 1, 5),
            tenure_category="<1 year",
            time_in_job_profile="6 months",
            performance=PerformanceLevel.MEDIUM,
            potential=PotentialLevel.HIGH,
            grid_position=8,
            talent_indicator="High Potential",
        ),
    ]


@pytest.fixture
def rich_sample_employees_small() -> list[Employee]:
    """Create a small rich sample dataset (50 employees) for tests.

    This is an alias for sample_employees, providing a more explicit name
    for tests that want to indicate they're using the rich sample generator.

    Use this when:
    - Testing with realistic organizational hierarchies
    - Need complete performance history (3 years)
    - Want to test bias detection patterns
    - Need coverage of all job levels and grid positions

    For minimal single-employee tests, use create_simple_test_employee() instead.
    For larger datasets, use rich_sample_employees_medium or rich_sample_employees_large.

    Returns:
        List of 50 Employee objects with complete organizational data
    """
    from ninebox.services.sample_data_generator import RichDatasetConfig, generate_rich_dataset

    config = RichDatasetConfig(
        size=50,
        include_bias=True,
        seed=42,  # Fixed seed for test reproducibility
        locations=["USA", "CAN", "GBR"],
        job_functions=["Engineering", "Product Manager", "Sales", "Marketing"],
    )

    return generate_rich_dataset(config)


@pytest.fixture
def rich_sample_employees_medium() -> list[Employee]:
    """Create a medium rich sample dataset (100 employees) for tests.

    Uses rich data generator with bias patterns for intelligence testing.
    Provides more employees for statistical significance in analytics tests.

    Use this when:
    - Testing intelligence/statistics features (need 30+ employees)
    - Testing large-scale data operations
    - Need diverse organizational structures
    - Want to validate bias detection with statistical significance

    Returns:
        List of 100 Employee objects with complete organizational data
    """
    from ninebox.services.sample_data_generator import RichDatasetConfig, generate_rich_dataset

    config = RichDatasetConfig(
        size=100,
        include_bias=True,
        seed=42,  # Fixed seed for test reproducibility
        locations=["USA", "CAN", "GBR", "DEU"],
        job_functions=["Engineering", "Product Manager", "Sales", "Marketing", "Operations"],
    )

    return generate_rich_dataset(config)


@pytest.fixture
def rich_sample_employees_large() -> list[Employee]:
    """Create a large rich sample dataset (200 employees) for tests.

    Uses rich data generator with complete features including all locations
    and job functions. Ideal for integration tests and performance validation.

    Use this when:
    - Testing performance with realistic data volumes
    - Need comprehensive coverage of all edge cases
    - Testing export/import with large datasets
    - Validating UI rendering with many employees

    Returns:
        List of 200 Employee objects with complete organizational data
    """
    from ninebox.services.sample_data_generator import RichDatasetConfig, generate_rich_dataset

    config = RichDatasetConfig(
        size=200,
        include_bias=True,
        seed=42,  # Fixed seed for test reproducibility
    )

    return generate_rich_dataset(config)


@pytest.fixture
def sample_excel_file(tmp_path: Path, sample_employees: list[Employee]) -> Path:
    """Create a sample Excel file for testing (5 employees).

    For integration tests requiring more employees, use rich_sample_excel_file.
    """
    file_path = tmp_path / "test_employees.xlsx"

    # Create workbook with two sheets
    workbook = openpyxl.Workbook()

    # First sheet (summary/metadata)
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet["A1"] = "9-Box Talent Review"
    summary_sheet["A2"] = f"Generated: {datetime.now(timezone.utc).isoformat()}"

    # Second sheet (employee data)
    data_sheet = workbook.create_sheet("Employee Data")

    # Headers
    headers = [
        "Employee ID",
        "Worker",
        "Business Title",
        "Job Title",
        "Job Profile",
        "Job Level - Primary Position",
        "Worker's Manager",
        "Management Chain - Level 04",
        "Management Chain - Level 05",
        "Management Chain - Level 06",
        "Hire Date",
        "Tenure Category (Months)",
        "Time in Job Profile",
        "Aug 2025 Talent Assessment Performance",
        "Aug 2025  Talent Assessment Potential",
        "Aug 2025 Talent Assessment 9-Box Label",
        "Talent Mapping Position [Performance vs Potential]",
        "FY25 Talent Indicator",
        "2023 Completed Performance Rating",
        "2024 Completed Performance Rating",
        "Development Focus",
        "Development Action",
        "Notes",
        "Promotion (In-Line,",
    ]

    for col_idx, header in enumerate(headers, start=1):
        data_sheet.cell(1, col_idx, header)

    # Data rows
    for row_idx, emp in enumerate(sample_employees, start=2):
        data_sheet.cell(row_idx, 1, emp.employee_id)
        data_sheet.cell(row_idx, 2, emp.name)
        data_sheet.cell(row_idx, 3, emp.business_title)
        data_sheet.cell(row_idx, 4, emp.job_title)
        data_sheet.cell(row_idx, 5, emp.job_profile)
        data_sheet.cell(row_idx, 6, emp.job_level)
        data_sheet.cell(row_idx, 7, emp.direct_manager)
        data_sheet.cell(row_idx, 8, emp.management_chain_04)
        data_sheet.cell(row_idx, 9, emp.management_chain_05)
        data_sheet.cell(row_idx, 10, emp.management_chain_06)
        data_sheet.cell(row_idx, 11, emp.hire_date.isoformat() if emp.hire_date else None)
        data_sheet.cell(row_idx, 12, emp.tenure_category)
        data_sheet.cell(row_idx, 13, emp.time_in_job_profile)
        data_sheet.cell(row_idx, 14, emp.performance.value)
        data_sheet.cell(row_idx, 15, emp.potential.value)
        data_sheet.cell(row_idx, 16, emp.grid_position)
        data_sheet.cell(row_idx, 17, get_position_label_by_number(emp.grid_position))
        data_sheet.cell(row_idx, 18, emp.talent_indicator)

        # Historical ratings
        if len(emp.ratings_history) > 0 and emp.ratings_history[0].year == 2023:
            data_sheet.cell(row_idx, 19, emp.ratings_history[0].rating)
        if len(emp.ratings_history) > 0:
            for rating in emp.ratings_history:
                if rating.year == 2024:
                    data_sheet.cell(row_idx, 20, rating.rating)

        data_sheet.cell(row_idx, 21, emp.development_focus)
        data_sheet.cell(row_idx, 22, emp.development_action)
        data_sheet.cell(row_idx, 23, emp.notes)
        data_sheet.cell(row_idx, 24, emp.promotion_status)

    workbook.save(file_path)
    workbook.close()  # Explicitly close to prevent openpyxl state pollution
    return file_path


@pytest.fixture
def rich_sample_excel_file(tmp_path: Path, rich_sample_employees_small: list[Employee]) -> Path:
    """Create a rich sample Excel file for testing with 50 employees.

    Uses rich_sample_employees_small fixture to generate realistic test data
    with organizational hierarchies and complete employee records.

    Use this for integration tests that need larger datasets.
    """
    file_path = tmp_path / "test_employees.xlsx"

    # Create workbook with two sheets
    workbook = openpyxl.Workbook()

    # First sheet (summary/metadata)
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet["A1"] = "9-Box Talent Review"
    summary_sheet["A2"] = f"Generated: {datetime.now(timezone.utc).isoformat()}"

    # Second sheet (employee data)
    data_sheet = workbook.create_sheet("Employee Data")

    # Headers
    headers = [
        "Employee ID",
        "Worker",
        "Business Title",
        "Job Title",
        "Job Profile",
        "Job Level - Primary Position",
        "Worker's Manager",
        "Management Chain - Level 04",
        "Management Chain - Level 05",
        "Management Chain - Level 06",
        "Hire Date",
        "Tenure Category (Months)",
        "Time in Job Profile",
        "Aug 2025 Talent Assessment Performance",
        "Aug 2025  Talent Assessment Potential",
        "Aug 2025 Talent Assessment 9-Box Label",
        "Talent Mapping Position [Performance vs Potential]",
        "FY25 Talent Indicator",
        "2023 Completed Performance Rating",
        "2024 Completed Performance Rating",
        "Development Focus",
        "Development Action",
        "Notes",
        "Promotion (In-Line,",
    ]

    for col_idx, header in enumerate(headers, start=1):
        data_sheet.cell(1, col_idx, header)

    # Data rows
    for row_idx, emp in enumerate(rich_sample_employees_small, start=2):
        data_sheet.cell(row_idx, 1, emp.employee_id)
        data_sheet.cell(row_idx, 2, emp.name)
        data_sheet.cell(row_idx, 3, emp.business_title)
        data_sheet.cell(row_idx, 4, emp.job_title)
        data_sheet.cell(row_idx, 5, emp.job_profile)
        data_sheet.cell(row_idx, 6, emp.job_level)
        data_sheet.cell(row_idx, 7, emp.direct_manager)
        data_sheet.cell(row_idx, 8, emp.management_chain_04)
        data_sheet.cell(row_idx, 9, emp.management_chain_05)
        data_sheet.cell(row_idx, 10, emp.management_chain_06)
        data_sheet.cell(row_idx, 11, emp.hire_date.isoformat() if emp.hire_date else None)
        data_sheet.cell(row_idx, 12, emp.tenure_category)
        data_sheet.cell(row_idx, 13, emp.time_in_job_profile)
        data_sheet.cell(row_idx, 14, emp.performance.value)
        data_sheet.cell(row_idx, 15, emp.potential.value)
        data_sheet.cell(row_idx, 16, emp.grid_position)
        data_sheet.cell(row_idx, 17, get_position_label_by_number(emp.grid_position))
        data_sheet.cell(row_idx, 18, emp.talent_indicator)

        # Historical ratings
        if len(emp.ratings_history) > 0 and emp.ratings_history[0].year == 2023:
            data_sheet.cell(row_idx, 19, emp.ratings_history[0].rating)
        if len(emp.ratings_history) > 0:
            for rating in emp.ratings_history:
                if rating.year == 2024:
                    data_sheet.cell(row_idx, 20, rating.rating)

        data_sheet.cell(row_idx, 21, emp.development_focus)
        data_sheet.cell(row_idx, 22, emp.development_action)
        data_sheet.cell(row_idx, 23, emp.notes)
        data_sheet.cell(row_idx, 24, emp.promotion_status)

    workbook.save(file_path)
    workbook.close()  # Explicitly close to prevent openpyxl state pollution
    return file_path


@pytest.fixture
def test_client(test_db_path: str) -> TestClient:
    """Create a FastAPI test client."""
    # Import app after database path is set
    # Patch database path for testing
    import ninebox.core.database as db_module  # noqa: PLC0415
    from ninebox.main import app  # noqa: PLC0415

    original_get_db_path = db_module.get_db_path

    def mock_get_db_path() -> Path:
        return Path(test_db_path)

    db_module.get_db_path = mock_get_db_path

    client = TestClient(app)

    yield client

    # Restore original function
    db_module.get_db_path = original_get_db_path


@pytest.fixture
def auth_headers() -> dict[str, str]:
    """Get authentication headers for testing.

    Note: This app is local-only without authentication.
    This fixture returns empty headers for backward compatibility.
    """
    return {}


@pytest.fixture
def session_with_employees(test_client: TestClient, auth_headers: dict[str, str]) -> dict[str, str]:
    """Create session with employees via sample data API (no file upload).

    This fixture uses the /api/employees/generate-sample endpoint to create
    a session with employee data, bypassing the file upload flow. This is
    faster than session_with_data while properly setting up cookies and
    session state.

    Use this fixture for:
    - Statistics API tests
    - Intelligence API tests
    - Employee CRUD tests
    - Filter/search tests
    - Any test that needs employee data but doesn't test file upload

    Use session_with_data (file upload) ONLY for:
    - File upload/download tests
    - Excel parsing tests
    - Session management integration tests

    Args:
        test_client: FastAPI test client
        auth_headers: Authentication headers (empty dict for local-only app)

    Returns:
        Headers dict (includes session cookie set by API)

    Example:
        >>> def test_get_employees(test_client, session_with_employees):
        ...     response = test_client.get("/api/employees", headers=session_with_employees)
        ...     assert response.status_code == 200
        ...     employees = response.json()["employees"]
        ...     assert len(employees) > 0
    """
    # Use the generate-sample endpoint to create a session with sample data
    # This sets up cookies and session state properly while avoiding file I/O
    # Note: generate-sample requires size >= 50
    response = test_client.post(
        "/api/employees/generate-sample",
        json={"size": 50, "include_bias": False, "seed": 42},
        headers=auth_headers,
    )
    assert response.status_code == 200

    return auth_headers
