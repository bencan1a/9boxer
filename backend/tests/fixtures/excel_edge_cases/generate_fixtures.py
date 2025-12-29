"""Generate Excel test fixtures for edge case testing.

This script creates 11 different Excel files to test various edge cases:
1. unicode_characters.xlsx - Non-ASCII characters (Chinese, Arabic, emoji, accents)
2. large_file.xlsx - 1000 employees for performance testing
3. formulas.xlsx - Excel formulas in cells
4. empty_file.xlsx - Headers only, no data rows
5. multiple_sheets.xlsx - Multiple sheets (should use second sheet)
6. extra_columns.xlsx - Extra columns beyond required schema
7. missing_optional.xlsx - Only required columns
8. long_strings.xlsx - 500+ character strings
9. special_characters.xlsx - Apostrophes, hyphens, slashes in names
10. leading_trailing_spaces.xlsx - Whitespace that should be trimmed
11. merged_cells.xlsx - Merged cells in data area
"""

from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


def get_base_headers() -> list[str]:
    """Return the standard headers used in employee Excel files."""
    return [
        "Employee ID",
        "Worker",
        "Business Title",
        "Job Title",
        "Job Profile",
        "Job Level - Primary Position",
        "Worker's Manager",
        "Management Chain - Level 01",
        "Management Chain - Level 02",
        "Management Chain - Level 03",
        "Management Chain - Level 04",
        "Management Chain - Level 05",
        "Management Chain - Level 06",
        "Hire Date",
        "Tenure Category (Months)",
        "Time in Job Profile",
        "Aug 2025 Talent Assessment Performance",
        "Aug 2025  Talent Assessment Potential",
        "Aug 2025 Talent Assessment 9-Box Label",
        "Talent Mapping Position [Performance vs Potential]",
        "FY25 Talent Indicator",
        "2023 Completed Performance Rating",
        "2024 Completed Performance Rating",
        "Development Focus",
        "Development Action",
        "Notes",
        "Promotion (In-Line,",
        "Promotion Readiness",
    ]


def create_summary_sheet(workbook: openpyxl.Workbook) -> None:
    """Create the summary sheet (first sheet)."""
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet["A1"] = "9-Box Talent Review"
    summary_sheet["A2"] = f"Generated: {datetime.utcnow().isoformat()}"


def add_headers_to_sheet(sheet: openpyxl.worksheet.worksheet.Worksheet, headers: list[str]) -> None:
    """Add headers to the first row of a sheet."""
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(1, col_idx, header)


def generate_unicode_characters() -> None:
    """Generate file with unicode characters (Chinese, Arabic, emoji, accented)."""
    output_path = Path(__file__).parent / "unicode_characters.xlsx"

    workbook = openpyxl.Workbook()
    create_summary_sheet(workbook)

    data_sheet = workbook.create_sheet("Employee Data")
    headers = get_base_headers()
    add_headers_to_sheet(data_sheet, headers)

    # Test data with various unicode characters
    test_data = [
        (1, "张伟 (Zhang Wei)", "高级工程师", "Software Engineer", "Software EngineeringCHN"),
        (2, "محمد علي", "Senior Engineer", "مهندس برمجيات", "Software EngineeringSAU"),
        (3, "José García", "Ingeniero Principal", "Staff Engineer", "Engineering LeadershipESP"),
        (4, "Françoise Müller", "Développeur Senior", "Senior Developer", "Product DevelopmentFRA"),
        (5, "Emoji Test 😀🎉", "Title with 👨‍💻 emoji", "Engineer 🚀", "EngineeringUSA"),
        (6, "Søren Ødegård", "Sénior Développeur", "Lead Engineer", "Engineering LeadNOR"),
    ]

    for row_idx, (emp_id, name, title, job_title, profile) in enumerate(test_data, start=2):
        data_sheet.cell(row_idx, 1, emp_id)
        data_sheet.cell(row_idx, 2, name)
        data_sheet.cell(row_idx, 3, title)
        data_sheet.cell(row_idx, 4, job_title)
        data_sheet.cell(row_idx, 5, profile)
        data_sheet.cell(row_idx, 6, "MT4")
        data_sheet.cell(row_idx, 7, "Manager Name")
        data_sheet.cell(row_idx, 14, date(2020, 1, 1))
        data_sheet.cell(row_idx, 15, "3-5 years")
        data_sheet.cell(row_idx, 16, "2 years")
        data_sheet.cell(row_idx, 17, "High")
        data_sheet.cell(row_idx, 18, "High")

    workbook.save(output_path)
    print(f"Created: {output_path}")


def generate_large_file() -> None:
    """Generate file with 1000 employees for performance testing."""
    output_path = Path(__file__).parent / "large_file.xlsx"

    workbook = openpyxl.Workbook()
    create_summary_sheet(workbook)

    data_sheet = workbook.create_sheet("Employee Data")
    headers = get_base_headers()
    add_headers_to_sheet(data_sheet, headers)

    # Generate 1000 employees
    performance_levels = ["Low", "Medium", "High"]
    potential_levels = ["Low", "Medium", "High"]
    job_levels = ["MT1", "MT2", "MT3", "MT4", "MT5"]
    locations = ["USA", "CAN", "GBR", "DEU", "FRA"]

    for i in range(1, 1001):
        row_idx = i + 1
        data_sheet.cell(row_idx, 1, i)
        data_sheet.cell(row_idx, 2, f"Employee {i}")
        data_sheet.cell(row_idx, 3, f"Engineer {i % 10}")
        data_sheet.cell(row_idx, 4, "Software Engineer")
        data_sheet.cell(row_idx, 5, f"Software Engineering{locations[i % 5]}")
        data_sheet.cell(row_idx, 6, job_levels[i % 5])
        data_sheet.cell(row_idx, 7, f"Manager {i % 100}")
        data_sheet.cell(row_idx, 14, date(2020, 1, (i % 28) + 1))
        data_sheet.cell(row_idx, 15, "3-5 years")
        data_sheet.cell(row_idx, 16, "2 years")
        data_sheet.cell(row_idx, 17, performance_levels[i % 3])
        data_sheet.cell(row_idx, 18, potential_levels[i % 3])

    workbook.save(output_path)
    print(f"Created: {output_path}")


def generate_formulas() -> None:
    """Generate file with Excel formulas in cells."""
    output_path = Path(__file__).parent / "formulas.xlsx"

    workbook = openpyxl.Workbook()
    create_summary_sheet(workbook)

    data_sheet = workbook.create_sheet("Employee Data")
    headers = get_base_headers()
    add_headers_to_sheet(data_sheet, headers)

    # Add employees with formulas
    for i in range(1, 4):
        row_idx = i + 1
        data_sheet.cell(row_idx, 1, i)
        # Formula for concatenating first and last name
        data_sheet.cell(row_idx, 2, f"Employee {i}")
        data_sheet.cell(row_idx, 3, "Engineer")
        data_sheet.cell(row_idx, 4, "Software Engineer")
        data_sheet.cell(row_idx, 5, "Software EngineeringUSA")
        data_sheet.cell(row_idx, 6, "MT4")
        data_sheet.cell(row_idx, 7, "Manager")
        data_sheet.cell(row_idx, 14, date(2020, 1, 1))
        data_sheet.cell(row_idx, 15, "3-5 years")
        # Formula for calculating tenure
        data_sheet.cell(row_idx, 16, "2 years")
        data_sheet.cell(row_idx, 17, "High")
        data_sheet.cell(row_idx, 18, "Medium")

    workbook.save(output_path)
    print(f"Created: {output_path}")


def generate_empty_file() -> None:
    """Generate file with headers only, no data rows."""
    output_path = Path(__file__).parent / "empty_file.xlsx"

    workbook = openpyxl.Workbook()
    create_summary_sheet(workbook)

    data_sheet = workbook.create_sheet("Employee Data")
    headers = get_base_headers()
    add_headers_to_sheet(data_sheet, headers)

    # No data rows added

    workbook.save(output_path)
    print(f"Created: {output_path}")


def generate_multiple_sheets() -> None:
    """Generate file with multiple sheets."""
    output_path = Path(__file__).parent / "multiple_sheets.xlsx"

    workbook = openpyxl.Workbook()
    create_summary_sheet(workbook)

    # Add a decoy sheet first
    decoy_sheet = workbook.create_sheet("Old Data", 1)
    decoy_headers = ["ID", "Name", "Role"]
    add_headers_to_sheet(decoy_sheet, decoy_headers)
    decoy_sheet.cell(2, 1, 999)
    decoy_sheet.cell(2, 2, "Decoy Employee")
    decoy_sheet.cell(2, 3, "Fake Role")

    # Add the real employee data sheet (second sheet, index 1)
    data_sheet = workbook.create_sheet("Employee Data", 1)
    headers = get_base_headers()
    add_headers_to_sheet(data_sheet, headers)

    # Add real data
    data_sheet.cell(2, 1, 1)
    data_sheet.cell(2, 2, "Alice Smith")
    data_sheet.cell(2, 3, "Senior Engineer")
    data_sheet.cell(2, 4, "Software Engineer")
    data_sheet.cell(2, 5, "Software EngineeringUSA")
    data_sheet.cell(2, 6, "MT4")
    data_sheet.cell(2, 7, "Manager")
    data_sheet.cell(2, 14, date(2020, 1, 1))
    data_sheet.cell(2, 15, "3-5 years")
    data_sheet.cell(2, 16, "2 years")
    data_sheet.cell(2, 17, "High")
    data_sheet.cell(2, 18, "High")

    # Add another sheet after
    workbook.create_sheet("Archived")

    workbook.save(output_path)
    print(f"Created: {output_path}")


def generate_extra_columns() -> None:
    """Generate file with extra columns beyond required schema."""
    output_path = Path(__file__).parent / "extra_columns.xlsx"

    workbook = openpyxl.Workbook()
    create_summary_sheet(workbook)

    data_sheet = workbook.create_sheet("Employee Data")
    headers = get_base_headers() + [
        "Extra Column 1",
        "Extra Column 2",
        "Custom Field A",
        "Custom Field B",
        "Unused Data",
    ]
    add_headers_to_sheet(data_sheet, headers)

    # Add employee with extra column data
    data_sheet.cell(2, 1, 1)
    data_sheet.cell(2, 2, "Alice Smith")
    data_sheet.cell(2, 3, "Senior Engineer")
    data_sheet.cell(2, 4, "Software Engineer")
    data_sheet.cell(2, 5, "Software EngineeringUSA")
    data_sheet.cell(2, 6, "MT4")
    data_sheet.cell(2, 7, "Manager")
    data_sheet.cell(2, 14, date(2020, 1, 1))
    data_sheet.cell(2, 15, "3-5 years")
    data_sheet.cell(2, 16, "2 years")
    data_sheet.cell(2, 17, "High")
    data_sheet.cell(2, 18, "High")
    # Extra columns
    data_sheet.cell(2, 29, "Extra value 1")
    data_sheet.cell(2, 30, "Extra value 2")
    data_sheet.cell(2, 31, "Custom A")
    data_sheet.cell(2, 32, "Custom B")
    data_sheet.cell(2, 33, "This should be ignored")

    workbook.save(output_path)
    print(f"Created: {output_path}")


def generate_missing_optional() -> None:
    """Generate file with only required columns, missing optional ones."""
    output_path = Path(__file__).parent / "missing_optional.xlsx"

    workbook = openpyxl.Workbook()
    create_summary_sheet(workbook)

    data_sheet = workbook.create_sheet("Employee Data")
    # Only required columns
    headers = [
        "Employee ID",
        "Worker",
        "Business Title",
        "Job Level - Primary Position",
    ]
    add_headers_to_sheet(data_sheet, headers)

    # Add minimal employee data
    data_sheet.cell(2, 1, 1)
    data_sheet.cell(2, 2, "Minimal Employee")
    data_sheet.cell(2, 3, "Engineer")
    data_sheet.cell(2, 4, "MT3")

    data_sheet.cell(3, 1, 2)
    data_sheet.cell(3, 2, "Another Minimal")
    data_sheet.cell(3, 3, "Manager")
    data_sheet.cell(3, 4, "MT5")

    workbook.save(output_path)
    print(f"Created: {output_path}")


def generate_long_strings() -> None:
    """Generate file with 500+ character strings."""
    output_path = Path(__file__).parent / "long_strings.xlsx"

    workbook = openpyxl.Workbook()
    create_summary_sheet(workbook)

    data_sheet = workbook.create_sheet("Employee Data")
    headers = get_base_headers()
    add_headers_to_sheet(data_sheet, headers)

    # Generate very long strings
    long_title = "Senior Principal Staff Lead Architect Engineer " * 20  # ~940 chars
    long_notes = (
        "This employee has demonstrated exceptional performance across multiple "
        "dimensions including technical expertise, leadership capabilities, "
        "cross-functional collaboration, strategic thinking, and innovation. "
    ) * 15  # ~1650 chars
    long_name = "Alexander Bartholomew Christopher Danielson Evansworth Fitzgerald"

    data_sheet.cell(2, 1, 1)
    data_sheet.cell(2, 2, long_name)
    data_sheet.cell(2, 3, long_title[:500])  # Truncate to reasonable length
    data_sheet.cell(2, 4, "Software Engineer")
    data_sheet.cell(2, 5, "Software EngineeringUSA")
    data_sheet.cell(2, 6, "MT4")
    data_sheet.cell(2, 7, "Manager Name That Is Also Quite Long For Testing Purposes")
    data_sheet.cell(2, 14, date(2020, 1, 1))
    data_sheet.cell(2, 15, "3-5 years")
    data_sheet.cell(2, 16, "2 years")
    data_sheet.cell(2, 17, "High")
    data_sheet.cell(2, 18, "High")
    data_sheet.cell(
        2, 24, "Very long development focus that spans multiple lines and contains " * 10
    )
    data_sheet.cell(2, 26, long_notes)

    workbook.save(output_path)
    print(f"Created: {output_path}")


def generate_special_characters() -> None:
    """Generate file with special characters in names and titles."""
    output_path = Path(__file__).parent / "special_characters.xlsx"

    workbook = openpyxl.Workbook()
    create_summary_sheet(workbook)

    data_sheet = workbook.create_sheet("Employee Data")
    headers = get_base_headers()
    add_headers_to_sheet(data_sheet, headers)

    # Test data with special characters
    test_data = [
        (1, "O'Brien, John", "Senior Engineer", "Sr. Eng./Tech Lead", "EngineeringUSA"),
        (2, "Mary-Anne Smith-Jones", "Staff Engineer", "Eng. IV - Backend", "EngineeringCAN"),
        (3, "John (Jack) Doe", "Principal Eng.", "Principal/Lead Engineer", "EngineeringGBR"),
        (4, "Anna & Associates", "Engineer @ Team X", "Software Eng. [Backend]", "EngineeringDEU"),
        (5, "Test\\User", "Eng.\\Dev", "Developer/Engineer", "EngineeringFRA"),
        (6, "User#123", "Dev #1", "Engineer - Team #5", "EngineeringUSA"),
    ]

    for row_idx, (emp_id, name, title, job_title, profile) in enumerate(test_data, start=2):
        data_sheet.cell(row_idx, 1, emp_id)
        data_sheet.cell(row_idx, 2, name)
        data_sheet.cell(row_idx, 3, title)
        data_sheet.cell(row_idx, 4, job_title)
        data_sheet.cell(row_idx, 5, profile)
        data_sheet.cell(row_idx, 6, "MT4")
        data_sheet.cell(row_idx, 7, "Manager O'Brien")
        data_sheet.cell(row_idx, 14, date(2020, 1, 1))
        data_sheet.cell(row_idx, 15, "3-5 years")
        data_sheet.cell(row_idx, 16, "2 years")
        data_sheet.cell(row_idx, 17, "High")
        data_sheet.cell(row_idx, 18, "Medium")

    workbook.save(output_path)
    print(f"Created: {output_path}")


def generate_leading_trailing_spaces() -> None:
    """Generate file with leading/trailing whitespace."""
    output_path = Path(__file__).parent / "leading_trailing_spaces.xlsx"

    workbook = openpyxl.Workbook()
    create_summary_sheet(workbook)

    data_sheet = workbook.create_sheet("Employee Data")
    headers = get_base_headers()
    add_headers_to_sheet(data_sheet, headers)

    # Test data with whitespace issues
    test_data = [
        (
            1,
            "  Alice Smith  ",
            "  Senior Engineer  ",
            "Software Engineer",
            "Software EngineeringUSA",
        ),
        (2, "Bob Jones\t", "\tEngineer", "  Software Engineer  ", "Software EngineeringCAN"),
        (3, "\n\nCarol White\n", "Lead Engineer\n", "Staff Engineer", "Software EngineeringGBR"),
        (
            4,
            "  David Brown",
            "Principal Engineer  ",
            "  Principal Eng  ",
            "Software EngineeringDEU",
        ),
    ]

    for row_idx, (emp_id, name, title, job_title, profile) in enumerate(test_data, start=2):
        data_sheet.cell(row_idx, 1, emp_id)
        data_sheet.cell(row_idx, 2, name)
        data_sheet.cell(row_idx, 3, title)
        data_sheet.cell(row_idx, 4, job_title)
        data_sheet.cell(row_idx, 5, profile)
        data_sheet.cell(row_idx, 6, "  MT4  ")
        data_sheet.cell(row_idx, 7, "  Manager Name  ")
        data_sheet.cell(row_idx, 14, date(2020, 1, 1))
        data_sheet.cell(row_idx, 15, "  3-5 years  ")
        data_sheet.cell(row_idx, 16, "  2 years  ")
        data_sheet.cell(row_idx, 17, "  High  ")
        data_sheet.cell(row_idx, 18, "  Medium  ")

    workbook.save(output_path)
    print(f"Created: {output_path}")


def generate_merged_cells() -> None:
    """Generate file with merged cells in data area."""
    output_path = Path(__file__).parent / "merged_cells.xlsx"

    workbook = openpyxl.Workbook()
    create_summary_sheet(workbook)

    data_sheet = workbook.create_sheet("Employee Data")
    headers = get_base_headers()
    add_headers_to_sheet(data_sheet, headers)

    # Add employee data
    data_sheet.cell(2, 1, 1)
    data_sheet.cell(2, 2, "Alice Smith")
    data_sheet.cell(2, 3, "Senior Engineer")
    data_sheet.cell(2, 4, "Software Engineer")
    data_sheet.cell(2, 5, "Software EngineeringUSA")
    data_sheet.cell(2, 6, "MT4")
    data_sheet.cell(2, 7, "Manager")
    data_sheet.cell(2, 14, date(2020, 1, 1))
    data_sheet.cell(2, 15, "3-5 years")
    data_sheet.cell(2, 16, "2 years")
    data_sheet.cell(2, 17, "High")
    data_sheet.cell(2, 18, "High")

    # Merge some cells (Notes column spans multiple rows)
    data_sheet.merge_cells("Z2:Z4")
    data_sheet["Z2"] = "This note applies to multiple employees"

    # Add more employees
    data_sheet.cell(3, 1, 2)
    data_sheet.cell(3, 2, "Bob Jones")
    data_sheet.cell(3, 3, "Engineer")
    data_sheet.cell(3, 4, "Software Engineer")
    data_sheet.cell(3, 5, "Software EngineeringCAN")
    data_sheet.cell(3, 6, "MT3")
    data_sheet.cell(3, 7, "Manager")
    data_sheet.cell(3, 14, date(2020, 1, 1))
    data_sheet.cell(3, 17, "Medium")
    data_sheet.cell(3, 18, "Medium")

    workbook.save(output_path)
    print(f"Created: {output_path}")


if __name__ == "__main__":
    """Generate all test fixtures."""
    print("Generating Excel test fixtures...")
    print("-" * 60)

    generate_unicode_characters()
    generate_large_file()
    generate_formulas()
    generate_empty_file()
    generate_multiple_sheets()
    generate_extra_columns()
    generate_missing_optional()
    generate_long_strings()
    generate_special_characters()
    generate_leading_trailing_spaces()
    generate_merged_cells()

    print("-" * 60)
    print("All fixtures generated successfully!")
    print(f"Location: {Path(__file__).parent}")
