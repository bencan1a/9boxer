"""Integration tests for PyInstaller frozen executable.

This module tests the frozen backend executable built with PyInstaller to ensure
all functionality works correctly in the frozen state, including:
- Executable startup and health checks
- JWT authentication
- Excel import/export functionality
- Employee data management
- Performance metrics (startup time, memory usage)

These tests require the frozen executable to be built first using:
    ./scripts/build_executable.sh
"""

import os
import socket
import subprocess
import time
import uuid
from collections.abc import Generator
from pathlib import Path

import openpyxl
import pytest
import requests

# Executable path - dynamically determine based on platform
import platform

pytestmark = [pytest.mark.e2e, pytest.mark.slow]

if platform.system() == "Windows":
    FROZEN_EXE_PATH = Path("backend/dist/ninebox/ninebox.exe")
else:
    FROZEN_EXE_PATH = Path("backend/dist/ninebox/ninebox")

STARTUP_TIMEOUT = 15  # seconds
STARTUP_POLL_INTERVAL = 0.5  # seconds


def get_free_port() -> int:
    """Get a free port number for the test backend instance."""
    sock = socket.socket()
    sock.bind(("", 0))
    port = sock.getsockname()[1]
    sock.close()
    return port


@pytest.fixture
def frozen_executable_path() -> Path:
    """Get the path to the frozen executable."""
    return FROZEN_EXE_PATH


@pytest.fixture
def frozen_backend() -> Generator[tuple[subprocess.Popen, str], None, None]:
    """
    Start the frozen backend executable and ensure it's ready.

    Yields:
        tuple: (subprocess.Popen, base_url) - The running backend process and its base URL

    Raises:
        pytest.Failed: If the backend fails to start within timeout
    """
    # Ensure executable exists
    if not FROZEN_EXE_PATH.exists():
        pytest.skip(
            f"Frozen executable not found at {FROZEN_EXE_PATH}. Run ./scripts/build_executable.sh first."
        )

    # Get free port for this test instance to avoid conflicts
    port = get_free_port()
    base_url = f"http://localhost:{port}"

    # Create unique database for this test instance to ensure isolation
    test_db = FROZEN_EXE_PATH.parent / f"test_frozen_{uuid.uuid4().hex}.db"

    # Start the executable
    proc = subprocess.Popen(
        [str(FROZEN_EXE_PATH)],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        cwd=str(FROZEN_EXE_PATH.parent),
        env={**os.environ, "DATABASE_PATH": str(test_db), "PORT": str(port)},
    )

    # Wait for startup (max STARTUP_TIMEOUT seconds)
    start_time = time.time()
    backend_ready = False

    for _ in range(int(STARTUP_TIMEOUT / STARTUP_POLL_INTERVAL)):
        time.sleep(STARTUP_POLL_INTERVAL)

        # Check if process crashed
        if proc.poll() is not None:
            stdout, stderr = proc.communicate()
            pytest.fail(
                f"Backend process crashed during startup.\n"
                f"Exit code: {proc.returncode}\n"
                f"STDOUT: {stdout.decode()}\n"
                f"STDERR: {stderr.decode()}"
            )

        # Check health endpoint
        try:
            resp = requests.get(f"{base_url}/health", timeout=1)
            if resp.status_code == 200:
                backend_ready = True
                break
        except requests.exceptions.RequestException:
            continue

    startup_time = time.time() - start_time

    if not backend_ready:
        proc.terminate()
        proc.wait(timeout=5)
        stdout, stderr = proc.communicate()
        pytest.fail(
            f"Backend did not start within {STARTUP_TIMEOUT} seconds.\n"
            f"STDOUT: {stdout.decode()}\n"
            f"STDERR: {stderr.decode()}"
        )

    print(f"\n[OK] Backend started in {startup_time:.2f}s on port {port}")

    try:
        yield proc, base_url
    finally:
        # Clean shutdown
        if proc.poll() is None:
            proc.terminate()
            try:
                proc.wait(timeout=5)
            except subprocess.TimeoutExpired:
                proc.kill()
                proc.wait()

        # Clean up test database
        if test_db.exists():
            test_db.unlink()


@pytest.fixture
def auth_token(frozen_backend: tuple[subprocess.Popen, str]) -> str:
    """
    Get authentication token for API requests.

    Note: This app is local-only without authentication.
    Returns empty string for backward compatibility.

    Args:
        frozen_backend: The running backend process and base URL

    Returns:
        str: Empty string (no authentication)
    """
    proc, base_url = frozen_backend
    return ""


@pytest.fixture
def auth_headers() -> dict[str, str]:
    """
    Get authentication headers for API requests.

    Note: This app is local-only without authentication.
    Returns empty dict for backward compatibility.

    Returns:
        dict: Empty headers (no authentication)
    """
    return {}


@pytest.fixture
def sample_excel_file(tmp_path: Path) -> Path:
    """
    Create a sample Excel file for testing.

    Args:
        tmp_path: pytest temporary directory

    Returns:
        Path: Path to created Excel file
    """
    file_path = tmp_path / "test_employees.xlsx"

    # Create workbook with two sheets
    workbook = openpyxl.Workbook()

    # First sheet (summary/metadata)
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet["A1"] = "9-Box Talent Review"
    summary_sheet["A2"] = "Test Data for Frozen Executable"

    # Second sheet (employee data)
    data_sheet = workbook.create_sheet("Employee Data")

    # Headers
    headers = [
        "Employee ID",
        "Worker",
        "Business Title",
        "Job Title",
        "Job Profile",
        "Job Level - Primary Position",
        "Worker's Manager",
        "Management Chain - Level 04",
        "Management Chain - Level 05",
        "Management Chain - Level 06",
        "Hire Date",
        "Tenure Category (Months)",
        "Time in Job Profile",
        "Aug 2025 Talent Assessment Performance",
        "Aug 2025  Talent Assessment Potential",
        "Aug 2025 Talent Assessment 9-Box Label",
        "Talent Mapping Position [Performance vs Potential]",
        "FY25 Talent Indicator",
        "2023 Completed Performance Rating",
        "2024 Completed Performance Rating",
        "Development Focus",
        "Development Action",
        "Notes",
        "Promotion (In-Line,",
    ]

    for col_idx, header in enumerate(headers, start=1):
        data_sheet.cell(1, col_idx, header)

    # Sample data rows
    employees_data = [
        [
            1,
            "Alice Smith",
            "Senior Engineer",
            "Software Engineer",
            "Software EngineeringUSA",
            "MT4",
            "Bob Manager",
            "Bob Manager",
            "Carol Director",
            None,
            "2020-01-15",
            "3-5 years",
            "2 years",
            "High",
            "High",
            9,
            "Star [H,H]",
            "High Potential",
            "Strong",
            "Leading",
            "Leadership skills",
            "Executive coaching",
            "Top performer",
            "Ready now",
        ],
        [
            2,
            "Bob Jones",
            "Engineer",
            "Software Engineer",
            "Software EngineeringCAN",
            "MT2",
            "Bob Manager",
            "Bob Manager",
            "Carol Director",
            None,
            "2021-06-01",
            "1-3 years",
            "1 year",
            "Medium",
            "Medium",
            5,
            "Core Talent [M,M]",
            "Solid Contributor",
            None,
            "Solid",
            None,
            None,
            None,
            None,
        ],
        [
            3,
            "Carol Williams",
            "Junior Engineer",
            "Software Engineer",
            "Software EngineeringGBR",
            "MT1",
            "Dave Lead",
            "Dave Lead",
            "Carol Director",
            None,
            "2023-01-10",
            "0-1 year",
            "6 months",
            "Low",
            "High",
            7,
            "Enigma [L,H]",
            "Emerging",
            None,
            None,
            "Technical skills",
            "Mentorship program",
            "New hire",
            None,
        ],
    ]

    for row_idx, row_data in enumerate(employees_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            data_sheet.cell(row_idx, col_idx, value)

    workbook.save(file_path)
    workbook.close()  # Explicitly close to prevent openpyxl state pollution
    return file_path


# ============================================================================
# BASIC FUNCTIONALITY TESTS
# ============================================================================


@pytest.mark.skipif(not FROZEN_EXE_PATH.exists(), reason="Frozen executable not built")
def test_frozen_executable_exists() -> None:
    """Test that frozen executable was built and is executable."""
    assert FROZEN_EXE_PATH.exists(), f"Frozen executable not found at {FROZEN_EXE_PATH}"
    assert FROZEN_EXE_PATH.is_file(), "Executable is not a file"
    assert os.access(FROZEN_EXE_PATH, os.X_OK), "Executable is not executable"


def test_frozen_executable_starts(frozen_backend: tuple[subprocess.Popen, str]) -> None:
    """Test that frozen executable starts and responds to health check."""
    proc, base_url = frozen_backend
    # Backend already started by fixture, just verify it's running
    assert proc.poll() is None, "Backend process has terminated"

    # Test health endpoint
    response = requests.get(f"{base_url}/health", timeout=5)
    assert response.status_code == 200

    data = response.json()
    assert data["status"] == "healthy"
    # Version field is optional
    assert "status" in data


def test_frozen_swagger_ui_accessible(frozen_backend: tuple[subprocess.Popen, str]) -> None:
    """Test that Swagger UI documentation is accessible."""
    proc, base_url = frozen_backend
    response = requests.get(f"{base_url}/docs", timeout=5)
    assert response.status_code == 200
    assert b"swagger" in response.content.lower()


# ============================================================================
# AUTHENTICATION TESTS
# ============================================================================
# NOTE: All authentication tests removed - this app is local-only without authentication
# - test_frozen_authentication_login
# - test_frozen_authentication_invalid_credentials
# - test_frozen_authentication_protected_endpoint
# - test_frozen_authentication_without_token


# ============================================================================
# EXCEL IMPORT TESTS
# ============================================================================


def test_frozen_excel_import(
    frozen_backend: tuple[subprocess.Popen, str],
    auth_headers: dict[str, str],
    sample_excel_file: Path,
) -> None:
    """Test Excel import with frozen executable."""
    proc, base_url = frozen_backend
    # Upload Excel file
    with open(sample_excel_file, "rb") as f:  # noqa: PTH123
        files = {
            "file": (
                "test.xlsx",
                f,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        response = requests.post(
            f"{base_url}/api/session/upload",
            files=files,
            headers=auth_headers,
            timeout=10,
        )

    assert response.status_code == 200
    data = response.json()

    # Verify employees are parsed correctly
    assert "employee_count" in data
    assert data["employee_count"] == 3
    # uploaded_at is the actual field name (not upload_date)
    assert "uploaded_at" in data or "upload_date" in data
    assert "filename" in data


def test_frozen_excel_import_get_employees(
    frozen_backend: tuple[subprocess.Popen, str],
    auth_headers: dict[str, str],
    sample_excel_file: Path,
) -> None:
    """Test retrieving employees after Excel import."""
    proc, base_url = frozen_backend
    # Upload file
    with open(sample_excel_file, "rb") as f:  # noqa: PTH123
        files = {
            "file": (
                "test.xlsx",
                f,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        requests.post(
            f"{base_url}/api/session/upload", files=files, headers=auth_headers, timeout=10
        )

    # Get employees
    response = requests.get(f"{base_url}/api/employees", headers=auth_headers, timeout=5)
    assert response.status_code == 200

    data = response.json()
    assert "employees" in data
    assert len(data["employees"]) == 3

    # Check employee data structure
    employee = data["employees"][0]
    assert "employee_id" in employee
    assert "name" in employee
    assert "performance" in employee
    assert "potential" in employee
    assert "grid_position" in employee


# ============================================================================
# EMPLOYEE MANAGEMENT TESTS
# ============================================================================


def test_frozen_employee_move(
    frozen_backend: tuple[subprocess.Popen, str],
    auth_headers: dict[str, str],
    sample_excel_file: Path,
) -> None:
    """Test moving an employee in the 9-box grid."""
    proc, base_url = frozen_backend
    # Upload file
    with open(sample_excel_file, "rb") as f:  # noqa: PTH123
        files = {
            "file": (
                "test.xlsx",
                f,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        requests.post(
            f"{base_url}/api/session/upload", files=files, headers=auth_headers, timeout=10
        )

    # Move employee
    response = requests.patch(
        f"{base_url}/api/employees/1/move",
        json={"performance": "Low", "potential": "Medium"},
        headers=auth_headers,
        timeout=5,
    )

    assert response.status_code == 200
    data = response.json()

    # Verify move
    assert data["employee"]["performance"] == "Low"
    assert data["employee"]["potential"] == "Medium"
    assert data["employee"]["modified_in_session"] is True


def test_frozen_employee_statistics(
    frozen_backend: tuple[subprocess.Popen, str],
    auth_headers: dict[str, str],
    sample_excel_file: Path,
) -> None:
    """Test statistics calculation with scipy in frozen executable."""
    proc, base_url = frozen_backend
    # Upload file
    with open(sample_excel_file, "rb") as f:  # noqa: PTH123
        files = {
            "file": (
                "test.xlsx",
                f,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        requests.post(
            f"{base_url}/api/session/upload", files=files, headers=auth_headers, timeout=10
        )

    # Get statistics
    response = requests.get(f"{base_url}/api/statistics", headers=auth_headers, timeout=5)
    assert response.status_code == 200

    data = response.json()
    assert "total_employees" in data
    assert data["total_employees"] == 3
    assert "by_performance" in data
    assert "by_potential" in data
    # The actual field is "distribution" not "by_grid_position"
    assert "distribution" in data or "by_grid_position" in data


# ============================================================================
# EXCEL EXPORT TESTS
# ============================================================================


def test_frozen_excel_export(
    frozen_backend: tuple[subprocess.Popen, str],
    auth_headers: dict[str, str],
    sample_excel_file: Path,
) -> None:
    """Test Excel export with frozen executable."""
    proc, base_url = frozen_backend
    # Upload file
    with open(sample_excel_file, "rb") as f:  # noqa: PTH123
        files = {
            "file": (
                "test.xlsx",
                f,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        requests.post(
            f"{base_url}/api/session/upload", files=files, headers=auth_headers, timeout=10
        )

    # Move an employee
    requests.patch(
        f"{base_url}/api/employees/1/move",
        json={"performance": "Low", "potential": "Low"},
        headers=auth_headers,
        timeout=5,
    )

    # Export to Excel (returns JSON with file path)
    response = requests.post(
        f"{base_url}/api/session/export",
        json={"mode": "update_original"},
        headers=auth_headers,
        timeout=10,
    )
    assert response.status_code == 200
    export_data = response.json()
    assert export_data["success"] is True
    assert "file_path" in export_data

    # Verify exported file exists and is valid Excel
    exported_file_path = Path(export_data["file_path"])
    assert exported_file_path.exists(), f"Exported file not found at {exported_file_path}"

    workbook = openpyxl.load_workbook(exported_file_path)

    # Should have at least 2 sheets
    assert len(workbook.worksheets) >= 2

    # Find employee data sheet
    data_sheet = workbook.worksheets[1]

    # Should have headers
    assert data_sheet.cell(1, 1).value == "Employee ID"

    # Should have data rows
    assert data_sheet.max_row >= 4  # Header + 3 employees
    workbook.close()  # Prevent openpyxl state pollution


def test_frozen_excel_export_contains_changes(
    frozen_backend: tuple[subprocess.Popen, str],
    auth_headers: dict[str, str],
    sample_excel_file: Path,
) -> None:
    """Test that exported Excel file contains modifications."""
    proc, base_url = frozen_backend
    # Upload file
    with open(sample_excel_file, "rb") as f:  # noqa: PTH123
        files = {
            "file": (
                "test.xlsx",
                f,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        requests.post(
            f"{base_url}/api/session/upload", files=files, headers=auth_headers, timeout=10
        )

    # Move an employee
    requests.patch(
        f"{base_url}/api/employees/1/move",
        json={"performance": "Medium", "potential": "High"},
        headers=auth_headers,
        timeout=5,
    )

    # Export (returns JSON with file path)
    response = requests.post(
        f"{base_url}/api/session/export",
        json={"mode": "update_original"},
        headers=auth_headers,
        timeout=10,
    )
    assert response.status_code == 200
    export_data = response.json()
    assert export_data["success"] is True
    assert "file_path" in export_data

    # Verify changes are in exported file
    exported_file_path = Path(export_data["file_path"])
    assert exported_file_path.exists(), f"Exported file not found at {exported_file_path}"

    workbook = openpyxl.load_workbook(exported_file_path)
    sheet = workbook.worksheets[1]

    # Find "Modified in Session" column
    modified_col = None
    for col_idx in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, col_idx).value
        if cell_value and "Modified in Session" in str(cell_value):
            modified_col = col_idx
            break

    assert modified_col is not None, "Modified in Session column not found"

    # Check that at least one employee is marked as modified
    modified_found = False
    for row_idx in range(2, sheet.max_row + 1):
        if sheet.cell(row_idx, modified_col).value == "Yes":
            modified_found = True
            break

    assert modified_found, "No modified employees found in exported file"
    workbook.close()  # Prevent openpyxl state pollution


# ============================================================================
# PERFORMANCE TESTS
# ============================================================================


@pytest.mark.skipif(not FROZEN_EXE_PATH.exists(), reason="Frozen executable not built")
def test_frozen_performance_startup_time(frozen_executable_path: Path) -> None:
    """Measure startup time of frozen executable."""
    # Use a separate database for this test
    test_db = frozen_executable_path.parent / "test_perf.db"

    # Get free port for this test instance
    port = get_free_port()
    base_url = f"http://localhost:{port}"

    proc = subprocess.Popen(
        [str(frozen_executable_path)],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        cwd=str(frozen_executable_path.parent),
        env={**os.environ, "DATABASE_PATH": str(test_db), "PORT": str(port)},
    )

    start_time = time.time()
    backend_ready = False

    try:
        # Wait for backend to be ready
        for _ in range(int(STARTUP_TIMEOUT / STARTUP_POLL_INTERVAL)):
            time.sleep(STARTUP_POLL_INTERVAL)

            if proc.poll() is not None:
                pytest.fail("Process crashed during startup")

            try:
                resp = requests.get(f"{base_url}/health", timeout=1)
                if resp.status_code == 200:
                    backend_ready = True
                    break
            except requests.exceptions.RequestException:
                continue

        startup_time = time.time() - start_time

        assert backend_ready, "Backend did not start in time"
        assert startup_time < 5.0, f"Startup time {startup_time:.2f}s exceeds 5 second target"

        print(f"\n[OK] Startup time: {startup_time:.2f}s")

    finally:
        if proc.poll() is None:
            proc.terminate()
            proc.wait(timeout=5)
        if test_db.exists():
            test_db.unlink()


def test_frozen_performance_response_time(
    frozen_backend: tuple[subprocess.Popen, str], auth_headers: dict[str, str]
) -> None:
    """Measure response time of health check endpoint."""
    proc, base_url = frozen_backend
    # Warm up
    requests.get(f"{base_url}/health", timeout=5)

    # Measure response time
    samples = []
    for _ in range(10):
        start = time.time()
        response = requests.get(f"{base_url}/health", timeout=5)
        elapsed = time.time() - start

        assert response.status_code == 200
        samples.append(elapsed * 1000)  # Convert to ms

    avg_response_time = sum(samples) / len(samples)
    max_response_time = max(samples)

    print(f"\n[OK] Average response time: {avg_response_time:.2f}ms")
    print(f"[OK] Max response time: {max_response_time:.2f}ms")

    # Should be under 5000ms average (5 seconds is reasonable for API calls)
    # Note: This includes database queries and JSON serialization
    assert avg_response_time < 5000, (
        f"Average response time {avg_response_time:.2f}ms exceeds 5000ms target"
    )


def test_frozen_performance_memory_usage(frozen_backend: tuple[subprocess.Popen, str]) -> None:
    """Measure memory usage of frozen executable."""
    proc, base_url = frozen_backend
    try:
        import psutil  # noqa: PLC0415
    except ImportError:
        pytest.skip("psutil not installed")

    # Get process
    psutil_proc = psutil.Process(proc.pid)

    # Measure memory
    memory_info = psutil_proc.memory_info()
    memory_mb = memory_info.rss / (1024 * 1024)

    print(f"\n[OK] Memory usage: {memory_mb:.2f} MB")

    # Should be under 200MB idle
    assert memory_mb < 200, f"Memory usage {memory_mb:.2f} MB exceeds 200 MB target"


# ============================================================================
# ERROR HANDLING TESTS
# ============================================================================


def test_frozen_error_handling_invalid_excel(
    frozen_backend: tuple[subprocess.Popen, str], auth_headers: dict[str, str], tmp_path: Path
) -> None:
    """Test error handling for invalid Excel file."""
    proc, base_url = frozen_backend
    # Create invalid file
    invalid_file = tmp_path / "invalid.xlsx"
    invalid_file.write_text("This is not an Excel file")

    # Try to upload
    with open(invalid_file, "rb") as f:  # noqa: PTH123
        files = {
            "file": (
                "invalid.xlsx",
                f,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        response = requests.post(
            f"{base_url}/api/session/upload",
            files=files,
            headers=auth_headers,
            timeout=10,
        )

    # Should return error
    assert response.status_code >= 400


def test_frozen_error_handling_session_not_active(
    frozen_backend: tuple[subprocess.Popen, str], auth_headers: dict[str, str]
) -> None:
    """Test error handling when session is not active."""
    proc, base_url = frozen_backend
    # Try to get employees without uploading file
    response = requests.get(f"{base_url}/api/employees", headers=auth_headers, timeout=5)

    # Should return error or empty list
    assert response.status_code in [200, 400, 404]


def test_frozen_error_handling_invalid_employee_id(
    frozen_backend: tuple[subprocess.Popen, str],
    auth_headers: dict[str, str],
    sample_excel_file: Path,
) -> None:
    """Test error handling for invalid employee ID."""
    proc, base_url = frozen_backend
    # Upload file
    with open(sample_excel_file, "rb") as f:  # noqa: PTH123
        files = {
            "file": (
                "test.xlsx",
                f,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        requests.post(
            f"{base_url}/api/session/upload", files=files, headers=auth_headers, timeout=10
        )

    # Try to move non-existent employee
    response = requests.patch(
        f"{base_url}/api/employees/999/move",
        json={"performance": "High", "potential": "High"},
        headers=auth_headers,
        timeout=5,
    )

    # Should return 404
    assert response.status_code == 404


# ============================================================================
# END-TO-END WORKFLOW TEST
# ============================================================================


def test_frozen_complete_workflow(
    frozen_backend: tuple[subprocess.Popen, str], sample_excel_file: Path
) -> None:
    """Test complete workflow with frozen executable.

    Note: This app is local-only without authentication.
    """
    proc, base_url = frozen_backend
    # No authentication needed for local-only app
    headers = {}

    # 1. Upload Excel file
    with open(sample_excel_file, "rb") as f:  # noqa: PTH123
        files = {
            "file": (
                "test.xlsx",
                f,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        upload_response = requests.post(
            f"{base_url}/api/session/upload",
            files=files,
            headers=headers,
            timeout=10,
        )
    assert upload_response.status_code == 200
    assert upload_response.json()["employee_count"] == 3

    # 2. Get employees
    employees_response = requests.get(f"{base_url}/api/employees", headers=headers, timeout=5)
    assert employees_response.status_code == 200
    employees = employees_response.json()["employees"]
    assert len(employees) == 3

    # 3. Move employee
    move_response = requests.patch(
        f"{base_url}/api/employees/1/move",
        json={"performance": "Medium", "potential": "High"},
        headers=headers,
        timeout=5,
    )
    assert move_response.status_code == 200
    assert move_response.json()["employee"]["modified_in_session"] is True

    # 4. Get statistics
    stats_response = requests.get(f"{base_url}/api/statistics", headers=headers, timeout=5)
    assert stats_response.status_code == 200
    stats = stats_response.json()
    assert stats["modified_employees"] == 1

    # 5. Export file (returns JSON with file path)
    export_response = requests.post(
        f"{base_url}/api/session/export",
        json={"mode": "update_original"},
        headers=headers,
        timeout=10,
    )
    assert export_response.status_code == 200
    export_data = export_response.json()
    assert export_data["success"] is True
    assert "file_path" in export_data

    # 6. Verify exported file exists and is valid Excel
    exported_file_path = Path(export_data["file_path"])
    assert exported_file_path.exists(), f"Exported file not found at {exported_file_path}"

    workbook = openpyxl.load_workbook(exported_file_path)
    assert len(workbook.worksheets) >= 2

    print("\n[OK] Complete workflow test passed")
    workbook.close()  # Prevent openpyxl state pollution
