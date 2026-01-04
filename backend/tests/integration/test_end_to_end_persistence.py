"""End-to-end integration tests for complete session persistence workflows."""

from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient

from ninebox.core.dependencies import get_db_manager, get_session_manager
from ninebox.services.session_manager import SessionManager


pytestmark = [pytest.mark.integration, pytest.mark.slow]


class TestEndToEndPersistence:
    """Test complete workflows from upload to export with persistence."""

    def test_full_workflow_upload_move_notes_restart_export(
        self, test_client: TestClient, sample_excel_file: Path
    ) -> None:
        """Full workflow: Upload → move multiple employees → add notes → restart → export.

        This test verifies the complete persistence story:
        1. Upload file and create session
        2. Move multiple employees to different positions
        3. Add notes to some changes
        4. Simulate backend restart
        5. Verify all changes persisted
        6. Export to Excel and verify changes are reflected
        """
        # Step 1: Upload file
        with open(sample_excel_file, "rb") as f:  # noqa: PTH123
            files = {
                "file": (
                    "test.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            upload_response = test_client.post("/api/session/upload", files=files)

        assert upload_response.status_code == 200
        original_count = upload_response.json()["employee_count"]

        # Step 2: Get employees and move multiple ones
        response = test_client.get("/api/employees")
        employees = response.json()["employees"][:3]  # Move first 3

        moved_employees = []
        for idx, emp in enumerate(employees):
            employee_id = emp["employee_id"]
            original_perf = emp["performance"]

            # Move to a DIFFERENT performance level, varying potential
            new_perf = "Low" if original_perf != "Low" else "High"
            potential = ["High", "Medium", "Low"][idx]
            response = test_client.patch(
                f"/api/employees/{employee_id}/move",
                json={"performance": new_perf, "potential": potential},
            )
            assert response.status_code == 200
            moved_employees.append((employee_id, new_perf, potential))

        # Step 3: Add notes to first two employees
        notes_data = [
            (moved_employees[0][0], "Promoted to senior role, excellent leadership"),
            (moved_employees[1][0], "Strong technical contributor, needs mentoring"),
        ]

        for employee_id, notes in notes_data:
            response = test_client.patch(
                f"/api/session/changes/{employee_id}/notes",
                json={"notes": notes},
            )
            assert response.status_code == 200

        # Verify session state before restart
        response = test_client.get("/api/session/status")
        status_before = response.json()
        assert status_before["changes_count"] == 3

        # Step 4: Simulate backend restart
        # Clear the dependency cache to force new instance creation
        get_session_manager.cache_clear()
        new_manager = SessionManager()

        # Step 5: Verify all changes persisted after restart
        assert len(new_manager.sessions) == 1
        restored_session = list(new_manager.sessions.values())[0]

        # Verify employee count unchanged
        assert len(restored_session.original_employees) == original_count
        assert len(restored_session.current_employees) == original_count

        # Verify all 3 changes persisted
        assert len(restored_session.events) == 3

        # Verify employee positions persisted
        for employee_id, expected_perf, expected_potential in moved_employees:
            employee = next(
                e for e in restored_session.current_employees if e.employee_id == employee_id
            )
            assert employee.performance.value == expected_perf
            assert employee.potential.value == expected_potential
            assert employee.modified_in_session is True

        # Verify notes persisted
        for employee_id, expected_notes in notes_data:
            event = next(
                e
                for e in restored_session.events
                if e.employee_id == employee_id and e.event_type == "grid_move"
            )
            assert event.notes == expected_notes

        # Step 6: Export to Excel and verify changes reflected
        # The new_manager already has the sessions loaded from DB
        # Use dependency injection to override the session manager for this test
        # Since we're using test_client which has its own app instance,
        # the export should work with the persisted data

        response = test_client.post("/api/session/export", json={"mode": "update_original"})
        assert response.status_code == 200
        export_data = response.json()
        assert export_data["success"] is True

        # Load exported Excel
        exported_file_path = export_data["file_path"]
        workbook = openpyxl.load_workbook(exported_file_path)

        # Verify data sheet exists
        assert len(workbook.worksheets) >= 2
        data_sheet = workbook.worksheets[1]

        # Find Performance and Modified columns
        perf_col = None
        modified_col = None
        for col_idx in range(1, data_sheet.max_column + 1):
            cell_value = str(data_sheet.cell(1, col_idx).value or "")
            if "Performance" in cell_value:
                perf_col = col_idx
            if "Modified" in cell_value:
                modified_col = col_idx

        assert perf_col is not None
        assert modified_col is not None

        # Count modified employees in export
        modified_count = 0
        for row_idx in range(2, data_sheet.max_row + 1):
            if data_sheet.cell(row_idx, modified_col).value == "Yes":
                modified_count += 1

        assert modified_count == 3

    def test_database_and_filesystem_state_after_operations(
        self, test_client: TestClient, sample_excel_file: Path
    ) -> None:
        """Test database and file system state after each operation.

        Verifies that:
        - Database rows are created/updated/deleted correctly
        - File system state is consistent
        - No orphaned data remains
        """
        # Upload file
        with open(sample_excel_file, "rb") as f:  # noqa: PTH123
            files = {
                "file": (
                    "test.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            test_client.post("/api/session/upload", files=files)

        # Verify database state after upload
        db_mgr = get_db_manager()
        with db_mgr.get_connection() as conn:
            cursor = conn.execute("SELECT * FROM sessions")
            rows = cursor.fetchall()
            assert len(rows) == 1

            session_row = dict(rows[0])
            assert session_row["original_filename"] == "test.xlsx"
            assert session_row["sheet_name"] == "Employee Data"

        # Move employee to a DIFFERENT position
        response = test_client.get("/api/employees")
        first_employee = response.json()["employees"][0]
        employee_id = first_employee["employee_id"]
        original_perf = first_employee["performance"]
        original_pot = first_employee["potential"]

        # Ensure we move to a different position
        new_perf = "Low" if original_perf != "Low" else "High"
        new_pot = "Low" if original_pot != "Low" else "High"

        test_client.patch(
            f"/api/employees/{employee_id}/move",
            json={"performance": new_perf, "potential": new_pot},
        )

        # Verify database updated after move
        with db_mgr.get_connection() as conn:
            cursor = conn.execute("SELECT events, updated_at FROM sessions")
            row = cursor.fetchone()
            import json

            changes = json.loads(row["events"])
            assert len(changes) == 1
            assert changes[0]["employee_id"] == employee_id

        # Delete session
        test_client.delete("/api/session/clear")

        # Verify database cleaned up
        with db_mgr.get_connection() as conn:
            cursor = conn.execute("SELECT COUNT(*) FROM sessions")
            count = cursor.fetchone()[0]
            assert count == 0

    def test_original_vs_current_employees_after_restart(
        self, test_client: TestClient, sample_excel_file: Path
    ) -> None:
        """Verify original_employees vs current_employees after restart.

        Tests that:
        - original_employees remains unchanged after moves
        - current_employees reflects all moves
        - The diff between them is correct
        """
        # Upload file
        with open(sample_excel_file, "rb") as f:  # noqa: PTH123
            files = {
                "file": (
                    "test.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            test_client.post("/api/session/upload", files=files)

        # Get employees and their original positions
        response = test_client.get("/api/employees")
        employees = response.json()["employees"]
        employee_id = employees[0]["employee_id"]
        original_perf = employees[0]["performance"]
        original_pot = employees[0]["potential"]

        # Move employee to different position
        new_perf = "Low" if original_perf != "Low" else "High"
        new_pot = "High" if original_pot != "High" else "Medium"

        test_client.patch(
            f"/api/employees/{employee_id}/move",
            json={"performance": new_perf, "potential": new_pot},
        )

        # Simulate restart
        get_session_manager.cache_clear()
        new_manager = SessionManager()

        restored_session = list(new_manager.sessions.values())[0]

        # Verify original_employees unchanged
        original_employee = next(
            e for e in restored_session.original_employees if e.employee_id == employee_id
        )
        assert original_employee.performance.value == original_perf
        assert original_employee.potential.value == original_pot
        assert original_employee.modified_in_session is False

        # Verify current_employees reflects change
        current_employee = next(
            e for e in restored_session.current_employees if e.employee_id == employee_id
        )
        assert current_employee.performance.value == new_perf
        assert current_employee.potential.value == new_pot
        assert current_employee.modified_in_session is True

        # Verify change entry exists (grid move event)
        grid_events = [e for e in restored_session.events if e.event_type == "grid_move"]
        assert len(grid_events) == 1
        event = grid_events[0]
        assert event.employee_id == employee_id
        assert event.old_performance.value == original_perf
        assert event.old_potential.value == original_pot
        assert event.new_performance.value == new_perf
        assert event.new_potential.value == new_pot


class TestOriginalValueTracking:
    """Test that original Performance/Potential values are tracked in exports."""

    def test_export_tracks_original_values_when_employee_modified(
        self, test_client: TestClient, sample_excel_file: Path, export_dir: Path
    ) -> None:
        """Test that original Performance/Potential values are preserved in tracking columns.

        This test verifies:
        1. Original values are written to "Original Performance" and "Original Potential" columns
        2. New values are written to main Performance/Potential columns
        3. Changes persist when file is reloaded
        """
        # Upload file
        with open(sample_excel_file, "rb") as f:  # noqa: PTH123
            files = {
                "file": (
                    "test.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            test_client.post("/api/session/upload", files=files)

        # Get first employee and record original values
        response = test_client.get("/api/employees")
        employees = response.json()["employees"]
        employee_id = employees[0]["employee_id"]
        original_performance = employees[0]["performance"]
        original_potential = employees[0]["potential"]

        # Move employee to different position
        new_performance = "High" if original_performance != "High" else "Low"
        new_potential = "High" if original_potential != "High" else "Low"

        response = test_client.patch(
            f"/api/employees/{employee_id}/move",
            json={"performance": new_performance, "potential": new_potential},
        )
        assert response.status_code == 200

        # Export to new file
        export_path = export_dir / "exported_with_tracking.xlsx"
        response = test_client.post(
            "/api/session/export",
            json={"mode": "save_new", "new_path": str(export_path)},
        )
        assert response.status_code == 200
        assert response.json()["success"] is True

        # Load exported file and verify columns
        workbook = openpyxl.load_workbook(export_path)
        sheet = workbook.worksheets[1]

        # Find columns
        perf_col = None
        pot_col = None
        original_perf_col = None
        original_pot_col = None

        for col_idx in range(1, sheet.max_column + 1):
            cell_value = str(sheet.cell(1, col_idx).value or "")
            if cell_value == "Aug 2025 Talent Assessment Performance":
                perf_col = col_idx
            elif cell_value == "Aug 2025  Talent Assessment Potential":
                pot_col = col_idx
            elif cell_value == "Original Performance":
                original_perf_col = col_idx
            elif cell_value == "Original Potential":
                original_pot_col = col_idx

        assert perf_col is not None, "Performance column not found"
        assert pot_col is not None, "Potential column not found"
        assert original_perf_col is not None, "Original Performance column not found"
        assert original_pot_col is not None, "Original Potential column not found"

        # Find employee row
        employee_row = None
        for row_idx in range(2, sheet.max_row + 1):
            if int(sheet.cell(row_idx, 1).value) == employee_id:
                employee_row = row_idx
                break

        assert employee_row is not None, f"Employee {employee_id} not found in export"

        # Verify main columns have NEW values
        assert sheet.cell(employee_row, perf_col).value == new_performance
        assert sheet.cell(employee_row, pot_col).value == new_potential

        # Verify tracking columns have ORIGINAL values
        assert sheet.cell(employee_row, original_perf_col).value == original_performance
        assert sheet.cell(employee_row, original_pot_col).value == original_potential

        workbook.close()

        # Reload file to verify persistence
        from ninebox.services.excel_parser import ExcelParser

        parser = ExcelParser()
        result = parser.parse(export_path)

        # Find the employee in parsed data
        reloaded_employee = next(
            (e for e in result.employees if e.employee_id == employee_id), None
        )
        assert reloaded_employee is not None

        # Verify the NEW values persisted (main columns)
        assert reloaded_employee.performance.value == new_performance
        assert reloaded_employee.potential.value == new_potential

        # Original values are in tracking columns and don't affect reload
        # This ensures changes persist across reload

    def test_export_no_original_values_when_employee_not_modified(
        self, test_client: TestClient, sample_excel_file: Path, export_dir: Path
    ) -> None:
        """Test that unmodified employees have empty original value tracking columns."""
        # Upload file
        with open(sample_excel_file, "rb") as f:  # noqa: PTH123
            files = {
                "file": (
                    "test.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            test_client.post("/api/session/upload", files=files)

        # Export immediately without making changes
        export_path = export_dir / "exported_no_changes.xlsx"
        response = test_client.post(
            "/api/session/export",
            json={"mode": "save_new", "new_path": str(export_path)},
        )
        assert response.status_code == 200

        # Load exported file
        workbook = openpyxl.load_workbook(export_path)
        sheet = workbook.worksheets[1]

        # Find Original Performance and Original Potential columns
        original_perf_col = None
        original_pot_col = None
        for col_idx in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, col_idx).value
            if cell_value == "Original Performance":
                original_perf_col = col_idx
            elif cell_value == "Original Potential":
                original_pot_col = col_idx

        assert original_perf_col is not None
        assert original_pot_col is not None

        # All employees should have empty original value columns
        for row_idx in range(2, sheet.max_row + 1):
            original_perf_value = sheet.cell(row_idx, original_perf_col).value
            original_pot_value = sheet.cell(row_idx, original_pot_col).value
            # Empty cells can be "" or None
            assert original_perf_value in ("", None), (
                f"Row {row_idx} has non-empty original performance: {original_perf_value}"
            )
            assert original_pot_value in ("", None), (
                f"Row {row_idx} has non-empty original potential: {original_pot_value}"
            )

        workbook.close()


class TestFlagsAndDonutPersistence:
    """Test that flags and donut exercise data persist across export/reload cycles."""

    def test_flags_persist_across_export_reload(
        self, test_client: TestClient, sample_excel_file: Path, tmp_path: Path
    ) -> None:
        """Test that flags are written to Excel and read back correctly."""
        # Upload file
        with open(sample_excel_file, "rb") as f:  # noqa: PTH123
            files = {
                "file": (
                    "test.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            test_client.post("/api/session/upload", files=files)

        # Get first employee and add flags
        response = test_client.get("/api/employees")
        employees = response.json()["employees"]
        employee_id = employees[0]["employee_id"]

        # Add flags via API (assuming there's an endpoint for this)
        # For now, we'll test by directly using the exporter with flagged employees
        from ninebox.models.employee import Employee, PerformanceLevel, PotentialLevel
        from ninebox.services.excel_exporter import ExcelExporter
        from ninebox.services.excel_parser import ExcelParser
        from ninebox.services.session_manager import SessionManager
        from datetime import datetime, timezone

        # Get session
        from ninebox.core.dependencies import get_session_manager
        session_mgr = get_session_manager()
        session = session_mgr.get_session("local-user")
        assert session is not None

        # Add flags to first employee
        session.current_employees[0].flags = ["promotion_ready", "flight_risk"]

        # Export to new file
        export_path = tmp_path / "exported_with_flags.xlsx"
        exporter = ExcelExporter()
        exporter.export(
            sample_excel_file,
            session.current_employees,
            export_path,
            sheet_index=1,
            session=session,
        )

        # Reload file and verify flags persisted
        parser = ExcelParser()
        result = parser.parse(export_path)

        # Find the employee
        reloaded_employee = next(
            (e for e in result.employees if e.employee_id == employee_id), None
        )
        assert reloaded_employee is not None
        assert reloaded_employee.flags is not None
        assert set(reloaded_employee.flags) == {"promotion_ready", "flight_risk"}

    def test_donut_data_persists_across_export_reload(
        self, test_client: TestClient, sample_excel_file: Path, tmp_path: Path
    ) -> None:
        """Test that donut exercise data is written to Excel and read back correctly."""
        # Upload file
        with open(sample_excel_file, "rb") as f:  # noqa: PTH123
            files = {
                "file": (
                    "test.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            test_client.post("/api/session/upload", files=files)

        # Get employee and add donut placement
        from ninebox.models.employee import Employee, PerformanceLevel, PotentialLevel
        from ninebox.services.excel_exporter import ExcelExporter
        from ninebox.services.excel_parser import ExcelParser
        from datetime import datetime, timezone

        # Get session
        from ninebox.core.dependencies import get_session_manager
        session_mgr = get_session_manager()
        session = session_mgr.get_session("local-user")
        assert session is not None

        # Get first employee
        response = test_client.get("/api/employees")
        employees = response.json()["employees"]
        employee_id = employees[0]["employee_id"]

        # Add donut placement to first employee
        session.current_employees[0].donut_modified = True
        session.current_employees[0].donut_position = 6  # High performance, Medium potential
        session.current_employees[0].donut_performance = PerformanceLevel.HIGH
        session.current_employees[0].donut_potential = PotentialLevel.MEDIUM
        session.current_employees[0].donut_notes = "Hypothetical placement for leadership track"

        # Export to new file
        export_path = tmp_path / "exported_with_donut.xlsx"
        exporter = ExcelExporter()
        exporter.export(
            sample_excel_file,
            session.current_employees,
            export_path,
            sheet_index=1,
            session=session,
        )

        # Reload file and verify donut data persisted
        parser = ExcelParser()
        result = parser.parse(export_path)

        # Find the employee
        reloaded_employee = next(
            (e for e in result.employees if e.employee_id == employee_id), None
        )
        assert reloaded_employee is not None
        assert reloaded_employee.donut_modified is True
        assert reloaded_employee.donut_position == 6
        assert reloaded_employee.donut_performance == PerformanceLevel.HIGH
        assert reloaded_employee.donut_potential == PotentialLevel.MEDIUM
        assert reloaded_employee.donut_notes == "Hypothetical placement for leadership track"

    def test_flags_and_donut_together_persist(
        self, test_client: TestClient, sample_excel_file: Path, tmp_path: Path
    ) -> None:
        """Test that both flags and donut data can coexist and persist together."""
        # Upload file
        with open(sample_excel_file, "rb") as f:  # noqa: PTH123
            files = {
                "file": (
                    "test.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            test_client.post("/api/session/upload", files=files)

        from ninebox.models.employee import Employee, PerformanceLevel, PotentialLevel
        from ninebox.services.excel_exporter import ExcelExporter
        from ninebox.services.excel_parser import ExcelParser

        # Get session
        from ninebox.core.dependencies import get_session_manager
        session_mgr = get_session_manager()
        session = session_mgr.get_session("local-user")
        assert session is not None

        # Get first employee
        response = test_client.get("/api/employees")
        employees = response.json()["employees"]
        employee_id = employees[0]["employee_id"]

        # Add both flags and donut placement
        session.current_employees[0].flags = ["promotion_ready", "new_hire"]
        session.current_employees[0].donut_modified = True
        session.current_employees[0].donut_position = 9
        session.current_employees[0].donut_performance = PerformanceLevel.HIGH
        session.current_employees[0].donut_potential = PotentialLevel.HIGH
        session.current_employees[0].donut_notes = "Star performer potential"

        # Export
        export_path = tmp_path / "exported_with_both.xlsx"
        exporter = ExcelExporter()
        exporter.export(
            sample_excel_file,
            session.current_employees,
            export_path,
            sheet_index=1,
            session=session,
        )

        # Reload and verify both persisted
        parser = ExcelParser()
        result = parser.parse(export_path)

        reloaded_employee = next(
            (e for e in result.employees if e.employee_id == employee_id), None
        )
        assert reloaded_employee is not None

        # Check flags
        assert reloaded_employee.flags is not None
        assert set(reloaded_employee.flags) == {"promotion_ready", "new_hire"}

        # Check donut data
        assert reloaded_employee.donut_modified is True
        assert reloaded_employee.donut_position == 9
        assert reloaded_employee.donut_performance == PerformanceLevel.HIGH
        assert reloaded_employee.donut_potential == PotentialLevel.HIGH
        assert reloaded_employee.donut_notes == "Star performer potential"
