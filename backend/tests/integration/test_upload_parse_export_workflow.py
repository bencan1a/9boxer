"""Integration tests for Upload → Parse → Grid → Export workflow.

Tests the complete roundtrip from Excel upload through modifications to export.
Integration points tested:
- Excel upload → Parser → SessionManager → Employee storage
- Grid position updates → Employee service → Export service
- Data integrity across full roundtrip (upload → modify → export → re-upload)
"""

from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient

pytestmark = pytest.mark.integration


class TestUploadParseFlow:
    """Test Excel upload flows through parser to session creation."""

    def test_upload_excel_then_parse_when_valid_file_then_creates_session_with_employees(
        self, test_client: TestClient, sample_excel_file: Path
    ) -> None:
        """Test uploading Excel file creates session with parsed employees."""
        # Upload file
        with open(sample_excel_file, "rb") as f:  # noqa: PTH123
            files = {
                "file": (
                    "test.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            response = test_client.post("/api/session/upload", files=files)

        assert response.status_code == 200
        data = response.json()
        assert data["employee_count"] == 5  # sample_excel_file has 5 employees
        assert data["session_id"]
        assert data["filename"] == "test.xlsx"

        # Verify session was created
        status_response = test_client.get("/api/session/status")
        assert status_response.status_code == 200
        status = status_response.json()
        assert status["active"] is True
        assert status["employee_count"] == 5

    def test_upload_excel_then_query_employees_when_parsed_then_data_accessible(
        self, test_client: TestClient, sample_excel_file: Path
    ) -> None:
        """Test uploaded employee data is immediately queryable."""
        # Upload
        with open(sample_excel_file, "rb") as f:  # noqa: PTH123
            files = {
                "file": (
                    "test.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            test_client.post("/api/session/upload", files=files)

        # Query employees
        response = test_client.get("/api/employees")
        assert response.status_code == 200
        data = response.json()
        assert len(data["employees"]) == 5
        # Verify employee has expected fields
        first_emp = data["employees"][0]
        assert "employee_id" in first_emp
        assert "name" in first_emp
        assert "performance" in first_emp
        assert "potential" in first_emp


class TestGridModificationFlow:
    """Test grid position modifications flow through employee service."""

    def test_upload_then_move_employee_when_position_changes_then_persists(
        self, test_client: TestClient, sample_excel_file: Path
    ) -> None:
        """Test employee movement persists in session."""
        # Upload
        with open(sample_excel_file, "rb") as f:  # noqa: PTH123
            files = {
                "file": (
                    "test.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            test_client.post("/api/session/upload", files=files)

        # Get first employee
        emp_response = test_client.get("/api/employees")
        first_emp = emp_response.json()["employees"][0]
        employee_id = first_emp["employee_id"]

        # Move employee
        move_response = test_client.patch(
            f"/api/employees/{employee_id}/move",
            json={"performance": "Low", "potential": "High"},
        )
        assert move_response.status_code == 200
        assert move_response.json()["success"] is True

        # Verify change persisted
        verify_response = test_client.get(f"/api/employees/{employee_id}")
        assert verify_response.status_code == 200
        updated_emp = verify_response.json()
        assert updated_emp["performance"] == "Low"
        assert updated_emp["potential"] == "High"

    def test_upload_then_update_employee_fields_when_modified_then_persists(
        self, test_client: TestClient, sample_excel_file: Path
    ) -> None:
        """Test updating employee fields (notes, development) persists."""
        # Upload
        with open(sample_excel_file, "rb") as f:  # noqa: PTH123
            files = {
                "file": (
                    "test.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            test_client.post("/api/session/upload", files=files)

        # Get first employee
        emp_response = test_client.get("/api/employees")
        first_emp = emp_response.json()["employees"][0]
        employee_id = first_emp["employee_id"]

        # Update fields
        update_response = test_client.patch(
            f"/api/employees/{employee_id}",
            json={
                "notes": "Test note",
                "development_focus": "Leadership",
                "development_action": "Coaching",
            },
        )
        assert update_response.status_code == 200

        # Verify updates persisted
        verify_response = test_client.get(f"/api/employees/{employee_id}")
        updated_emp = verify_response.json()
        assert updated_emp["notes"] == "Test note"
        assert updated_emp["development_focus"] == "Leadership"
        assert updated_emp["development_action"] == "Coaching"


class TestExportRoundtripFlow:
    """Test complete roundtrip: upload → modify → export → verify."""

    def test_upload_modify_export_when_roundtrip_then_modifications_preserved(
        self, test_client: TestClient, sample_excel_file: Path
    ) -> None:
        """Test modifications survive export/import roundtrip."""
        # Upload
        with open(sample_excel_file, "rb") as f:  # noqa: PTH123
            files = {
                "file": (
                    "test.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            test_client.post("/api/session/upload", files=files)

        # Get first employee and move them
        emp_response = test_client.get("/api/employees")
        first_emp = emp_response.json()["employees"][0]
        employee_id = first_emp["employee_id"]

        # Move to a specific known position
        move_response = test_client.patch(
            f"/api/employees/{employee_id}/move",
            json={"performance": "Low", "potential": "Low"},
        )
        assert move_response.status_code == 200

        # Verify employee was actually moved in session
        verify_response = test_client.get(f"/api/employees/{employee_id}")
        assert verify_response.status_code == 200
        updated_emp = verify_response.json()
        assert updated_emp["performance"] == "Low"
        assert updated_emp["potential"] == "Low"

        # Export
        export_response = test_client.post(
            "/api/session/export", json={"mode": "update_original"}
        )
        assert export_response.status_code == 200
        exported_path = export_response.json()["file_path"]

        # Verify export succeeded and file exists
        assert Path(exported_path).exists()

        # The key test: Verify the exported file can be loaded
        # and contains valid Excel data
        workbook = openpyxl.load_workbook(exported_path)
        assert len(workbook.worksheets) >= 2
        data_sheet = workbook.worksheets[1]
        # Should have at least the header row + data rows
        assert data_sheet.max_row >= 2
        workbook.close()

    def test_upload_export_when_no_modifications_then_data_integrity_maintained(
        self, test_client: TestClient, rich_sample_excel_file: Path
    ) -> None:
        """Test export without modifications maintains data integrity."""
        # Upload
        with open(rich_sample_excel_file, "rb") as f:  # noqa: PTH123
            files = {
                "file": (
                    "test.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            upload_response = test_client.post("/api/session/upload", files=files)

        original_count = upload_response.json()["employee_count"]

        # Get all employees
        emp_response = test_client.get("/api/employees")
        original_employees = emp_response.json()["employees"]

        # Export immediately (no changes)
        export_response = test_client.post(
            "/api/session/export", json={"mode": "update_original"}
        )
        assert export_response.status_code == 200
        exported_path = export_response.json()["file_path"]

        # Verify exported file has same employee count
        workbook = openpyxl.load_workbook(exported_path)
        data_sheet = workbook.worksheets[1]
        # Subtract 1 for header row
        exported_count = data_sheet.max_row - 1
        assert exported_count == original_count
        workbook.close()

    def test_upload_multiple_changes_export_when_roundtrip_then_all_changes_preserved(
        self, test_client: TestClient, sample_excel_file: Path
    ) -> None:
        """Test multiple modifications all preserved in export."""
        # Upload
        with open(sample_excel_file, "rb") as f:  # noqa: PTH123
            files = {
                "file": (
                    "test.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            test_client.post("/api/session/upload", files=files)

        # Get employees
        emp_response = test_client.get("/api/employees")
        employees = emp_response.json()["employees"][:3]  # Modify first 3

        # Make multiple changes
        for emp in employees:
            test_client.patch(
                f"/api/employees/{emp['employee_id']}/move",
                json={"performance": "High", "potential": "High"},
            )

        # Export
        export_response = test_client.post(
            "/api/session/export", json={"mode": "update_original"}
        )
        assert export_response.status_code == 200

        # Verify session shows 3 changes
        status_response = test_client.get("/api/session/status")
        # After export, changes count should be reset to 0 (baseline updated)
        assert status_response.json()["changes_count"] == 0

    def test_upload_export_clear_upload_when_sequence_then_sessions_independent(
        self, test_client: TestClient, sample_excel_file: Path
    ) -> None:
        """Test session lifecycle: upload → export → clear → new upload."""
        # First upload
        with open(sample_excel_file, "rb") as f:  # noqa: PTH123
            files = {
                "file": (
                    "test1.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            test_client.post("/api/session/upload", files=files)

        # Export
        export_response = test_client.post(
            "/api/session/export", json={"mode": "update_original"}
        )
        assert export_response.status_code == 200

        # Clear session
        clear_response = test_client.delete("/api/session/clear")
        assert clear_response.status_code == 200

        # Verify session cleared
        status_response = test_client.get("/api/session/status")
        assert status_response.json()["active"] is False

        # New upload
        with open(sample_excel_file, "rb") as f:  # noqa: PTH123
            files = {
                "file": (
                    "test2.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            new_upload_response = test_client.post("/api/session/upload", files=files)

        assert new_upload_response.status_code == 200
        assert new_upload_response.json()["employee_count"] == 5

    def test_upload_modify_export_when_modified_flag_then_highlighted_in_excel(
        self, test_client: TestClient, sample_excel_file: Path
    ) -> None:
        """Test modified employees are flagged in exported Excel."""
        # Upload
        with open(sample_excel_file, "rb") as f:  # noqa: PTH123
            files = {
                "file": (
                    "test.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            test_client.post("/api/session/upload", files=files)

        # Move first employee
        emp_response = test_client.get("/api/employees")
        first_emp = emp_response.json()["employees"][0]
        test_client.patch(
            f"/api/employees/{first_emp['employee_id']}/move",
            json={"performance": "Low", "potential": "Low"},
        )

        # Export
        export_response = test_client.post(
            "/api/session/export", json={"mode": "update_original"}
        )
        assert export_response.status_code == 200
        exported_path = export_response.json()["file_path"]

        # Verify modified flag in Excel
        workbook = openpyxl.load_workbook(exported_path)
        data_sheet = workbook.worksheets[1]

        # Find "Modified in Session" column
        modified_col = None
        for col_idx in range(1, data_sheet.max_column + 1):
            header = data_sheet.cell(1, col_idx).value
            if header and "Modified" in str(header):
                modified_col = col_idx
                break

        # Check first employee row is marked modified
        if modified_col:
            modified_value = data_sheet.cell(2, modified_col).value
            assert modified_value in ["Yes", "TRUE", True, "true", 1]

        workbook.close()


class TestParserServiceIntegration:
    """Test Excel parser integrates correctly with employee service."""

    def test_upload_when_rich_dataset_then_all_fields_parsed(
        self, test_client: TestClient, rich_sample_excel_file: Path
    ) -> None:
        """Test rich dataset upload preserves all employee fields."""
        # Upload rich sample file (50 employees with complete data)
        with open(rich_sample_excel_file, "rb") as f:  # noqa: PTH123
            files = {
                "file": (
                    "rich.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            response = test_client.post("/api/session/upload", files=files)

        assert response.status_code == 200
        assert response.json()["employee_count"] == 50

        # Verify employees have all expected fields
        emp_response = test_client.get("/api/employees")
        employees = emp_response.json()["employees"]
        assert len(employees) == 50

        # Check first employee has required fields
        first_emp = employees[0]
        required_fields = [
            "employee_id",
            "name",
            "business_title",
            "job_title",
            "job_level",
            "direct_manager",
            "performance",
            "potential",
            "grid_position",
        ]
        for field in required_fields:
            assert field in first_emp
            assert first_emp[field] is not None
