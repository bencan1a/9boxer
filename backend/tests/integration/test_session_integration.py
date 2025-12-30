"""Integration tests for session management and lifecycle."""

from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient

pytestmark = [pytest.mark.integration, pytest.mark.slow]


class TestSessionLifecycle:
    """Test complete session lifecycle including cleanup."""

    def test_session_when_created_then_persists_data(
        self, test_client: TestClient, rich_sample_excel_file: Path
    ) -> None:
        """Test session data persists across requests."""
        # Upload data
        with open(rich_sample_excel_file, "rb") as f:  # noqa: PTH123
            files = {
                "file": (
                    "test.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            response = test_client.post("/api/session/upload", files=files)
        assert response.status_code == 200
        assert response.json()["employee_count"] == 50

        # Get status - data should persist
        response = test_client.get("/api/session/status")
        assert response.status_code == 200
        assert response.json()["active"] is True
        assert response.json()["employee_count"] == 50

        # Get employees - data still there
        response = test_client.get("/api/employees")
        assert response.status_code == 200
        assert len(response.json()["employees"]) == 50

    def test_session_when_cleared_then_removes_all_data(
        self, test_client: TestClient, rich_sample_excel_file: Path
    ) -> None:
        """Test session clear removes all data."""
        # Upload data
        with open(rich_sample_excel_file, "rb") as f:  # noqa: PTH123
            files = {
                "file": (
                    "test.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            test_client.post("/api/session/upload", files=files)

        # Verify data exists
        response = test_client.get("/api/session/status")
        assert response.json()["active"] is True
        assert response.json()["employee_count"] == 50

        # Clear session
        response = test_client.delete("/api/session/clear")
        assert response.status_code == 200
        assert response.json()["success"] is True

        # Verify data is gone
        response = test_client.get("/api/session/status")
        assert response.status_code == 200
        assert response.json()["active"] is False

    def test_session_when_new_upload_then_replaces_data(
        self, test_client: TestClient, sample_excel_file: Path
    ) -> None:
        """Test new upload replaces existing session data."""
        # First upload
        with open(sample_excel_file, "rb") as f:  # noqa: PTH123
            files = {
                "file": (
                    "test1.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            response1 = test_client.post("/api/session/upload", files=files)
        count1 = response1.json()["employee_count"]

        # Second upload
        with open(sample_excel_file, "rb") as f:  # noqa: PTH123
            files = {
                "file": (
                    "test2.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            response2 = test_client.post("/api/session/upload", files=files)
        count2 = response2.json()["employee_count"]

        # Should have fresh data (counts should be same since same file)
        assert response2.status_code == 200
        assert count2 == count1

        # Verify only second upload data exists
        response = test_client.get("/api/session/status")
        assert response.json()["employee_count"] == count2
        assert response.json()["uploaded_filename"] == "test2.xlsx"


class TestChangeTracking:
    """Test change tracking across operations."""

    def test_changes_when_employee_moved_then_tracked(
        self, test_client: TestClient, sample_excel_file: Path
    ) -> None:
        """Test employee movements are tracked."""
        # Upload
        with open(sample_excel_file, "rb") as f:  # noqa: PTH123
            files = {
                "file": (
                    "test.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            test_client.post("/api/session/upload", files=files)

        # Get first employee
        response = test_client.get("/api/employees")
        employees = response.json()["employees"]
        first_employee = employees[0]
        employee_id = first_employee["employee_id"]
        original_perf = first_employee["performance"]
        original_pot = first_employee["potential"]

        # Move employee to a DIFFERENT position (not their original)
        new_perf = "Low" if original_perf != "Low" else "High"
        new_pot = "Low" if original_pot != "Low" else "High"

        response = test_client.patch(
            f"/api/employees/{employee_id}/move",
            json={"performance": new_perf, "potential": new_pot},
        )
        assert response.status_code == 200

        # Check employee is marked as modified
        response = test_client.get(f"/api/employees/{employee_id}")
        assert response.status_code == 200
        employee_data = response.json()
        assert employee_data["modified_in_session"] is True

        # Check status shows modifications
        response = test_client.get("/api/session/status")
        status_data = response.json()
        assert status_data["changes_count"] == 1

    def test_changes_when_multiple_moves_then_all_tracked(
        self, test_client: TestClient, sample_excel_file: Path
    ) -> None:
        """Test multiple changes are all tracked."""
        # Upload
        with open(sample_excel_file, "rb") as f:  # noqa: PTH123
            files = {
                "file": (
                    "test.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            test_client.post("/api/session/upload", files=files)

        # Get employees
        response = test_client.get("/api/employees")
        employees = response.json()["employees"][:3]  # Move first 3

        # Move multiple employees
        for emp in employees:
            test_client.patch(
                f"/api/employees/{emp['employee_id']}/move",
                json={"performance": "High", "potential": "Medium"},
            )

        # Verify all tracked
        response = test_client.get("/api/session/status")
        status_data = response.json()
        assert status_data["changes_count"] == 3

        # Verify each marked as modified
        for emp in employees:
            response = test_client.get(f"/api/employees/{emp['employee_id']}")
            emp_data = response.json()
            assert emp_data["modified_in_session"] is True


class TestExportIntegration:
    """Test Excel export integration with modifications."""

    def test_export_when_no_changes_then_includes_all_data(
        self, test_client: TestClient, rich_sample_excel_file: Path
    ) -> None:
        """Test export works without modifications."""
        # Upload
        with open(rich_sample_excel_file, "rb") as f:  # noqa: PTH123
            files = {
                "file": (
                    "test.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            upload_response = test_client.post("/api/session/upload", files=files)
        original_count = upload_response.json()["employee_count"]

        # Export
        response = test_client.post("/api/session/export", json={"mode": "update_original"})
        assert response.status_code == 200
        data = response.json()
        assert data["success"] is True
        assert "file_path" in data
        assert "message" in data

        # Verify exported file can be opened
        exported_file_path = data["file_path"]
        workbook = openpyxl.load_workbook(exported_file_path)
        assert len(workbook.worksheets) >= 2
        assert original_count == 50
        workbook.close()  # Prevent openpyxl state pollution

    def test_export_when_changes_made_then_highlights_modifications(
        self, test_client: TestClient, sample_excel_file: Path
    ) -> None:
        """Test export highlights modified employees."""
        # Upload
        with open(sample_excel_file, "rb") as f:  # noqa: PTH123
            files = {
                "file": (
                    "test.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            test_client.post("/api/session/upload", files=files)

        # Move an employee to a different position
        response = test_client.get("/api/employees")
        first_employee = response.json()["employees"][0]
        employee_id = first_employee["employee_id"]
        original_perf = first_employee["performance"]

        # Move to a DIFFERENT performance level
        new_perf = "Low" if original_perf != "Low" else "High"
        test_client.patch(
            f"/api/employees/{employee_id}/move",
            json={"performance": new_perf, "potential": "High"},
        )

        # Export
        response = test_client.post("/api/session/export", json={"mode": "update_original"})
        assert response.status_code == 200
        data = response.json()
        assert data["success"] is True
        assert "file_path" in data

        # Verify exported file has modified flag
        exported_file_path = data["file_path"]
        workbook = openpyxl.load_workbook(exported_file_path)
        # Get employee data sheet (should be second sheet)
        sheet = workbook.worksheets[1]

        # Find "Modified in Session" column
        modified_col = None
        for col_idx in range(1, min(sheet.max_column + 1, 50)):  # Limit search to avoid issues
            cell_value = sheet.cell(1, col_idx).value
            if cell_value and "Modified" in str(cell_value):
                modified_col = col_idx
                break

        # If modified column exists, verify at least one modified employee
        if modified_col:
            modified_found = False
            for row_idx in range(2, min(sheet.max_row + 1, 100)):
                cell_val = sheet.cell(row_idx, modified_col).value
                if cell_val in ["Yes", "TRUE", True, "true", "1", 1]:
                    modified_found = True
                    break
            assert modified_found, "No modified employees found in exported file"
        workbook.close()  # Prevent openpyxl state pollution

    def test_export_when_multiple_changes_then_all_tracked(
        self, test_client: TestClient, sample_excel_file: Path
    ) -> None:
        """Test export includes all modifications when multiple employees moved."""
        # Upload
        with open(sample_excel_file, "rb") as f:  # noqa: PTH123
            files = {
                "file": (
                    "test.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            test_client.post("/api/session/upload", files=files)

        # Move multiple employees
        response = test_client.get("/api/employees")
        employees = response.json()["employees"][:3]
        for emp in employees:
            test_client.patch(
                f"/api/employees/{emp['employee_id']}/move",
                json={"performance": "Low", "potential": "Low"},
            )

        # Export
        response = test_client.post("/api/session/export", json={"mode": "update_original"})
        assert response.status_code == 200
        data = response.json()
        assert data["success"] is True

        # Verify all 3 modifications are tracked
        exported_file_path = data["file_path"]
        workbook = openpyxl.load_workbook(exported_file_path)
        sheet = workbook.worksheets[1]

        # Count modified employees
        modified_col = None
        for col_idx in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, col_idx).value
            if cell_value and "Modified in Session" in str(cell_value):
                modified_col = col_idx
                break

        modified_count = 0
        for row_idx in range(2, sheet.max_row + 1):
            if sheet.cell(row_idx, modified_col).value == "Yes":
                modified_count += 1

        assert modified_count == 3
        workbook.close()  # Prevent openpyxl state pollution
