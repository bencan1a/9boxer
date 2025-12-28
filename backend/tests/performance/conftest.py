"""Fixtures for performance tests."""

import tempfile
from datetime import date, datetime, timezone
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient

from ninebox.models.employee import Employee, HistoricalRating, PerformanceLevel, PotentialLevel
from ninebox.models.grid_positions import get_position_label_by_number


def pytest_configure(config: pytest.Config) -> None:
    """Configure pytest to disable xdist for performance tests.

    pytest-benchmark requires serial execution for accurate benchmarking.
    This hook automatically disables xdist (parallel test execution) when
    running tests from the performance test directory.
    """
    # Check if we're running performance tests
    if config.getoption("file_or_dir"):
        test_paths = [str(path) for path in config.getoption("file_or_dir")]
        if any("performance" in path for path in test_paths):
            # Disable xdist by setting workers to 0
            config.option.numprocesses = 0
            config.option.dist = "no"


def create_test_employee(
    employee_id: int,
    job_function: str = "Other",
    performance: PerformanceLevel = PerformanceLevel.MEDIUM,
    potential: PotentialLevel = PotentialLevel.MEDIUM,
) -> Employee:
    """Create a test employee with given parameters."""
    # Vary attributes based on employee_id for diversity
    perf_levels = [PerformanceLevel.LOW, PerformanceLevel.MEDIUM, PerformanceLevel.HIGH]
    pot_levels = [PotentialLevel.LOW, PotentialLevel.MEDIUM, PotentialLevel.HIGH]
    functions = ["Engineering", "Product Management", "Sales", "Marketing", "Operations", "Other"]

    perf = perf_levels[employee_id % 3] if performance == PerformanceLevel.MEDIUM else performance
    pot = pot_levels[employee_id % 3] if potential == PotentialLevel.MEDIUM else potential
    func = functions[employee_id % len(functions)] if job_function == "Other" else job_function

    # Calculate grid position based on performance and potential
    perf_val = {"Low": 1, "Medium": 2, "High": 3}[perf.value]
    pot_val = {"Low": 1, "Medium": 2, "High": 3}[pot.value]
    grid_position = (pot_val - 1) * 3 + perf_val

    return Employee(
        employee_id=employee_id,
        name=f"Employee {employee_id}",
        business_title=f"Title {employee_id}",
        job_title=f"Job Title {employee_id}",
        job_profile=f"{func}USA",
        job_level=f"MT{(employee_id % 5) + 1}",
        job_function=func,
        location="USA" if employee_id % 2 == 0 else "CAN",
        manager=f"Manager {employee_id % 10}",
        management_chain_04=f"Manager {employee_id % 10}",
        management_chain_05=f"Director {employee_id % 5}",
        management_chain_06=None,
        hire_date=date(2020 + (employee_id % 5), 1, 1),
        tenure_category=f"{employee_id % 5}-{employee_id % 5 + 1} years",
        time_in_job_profile=f"{employee_id % 3} years",
        performance=perf,
        potential=pot,
        grid_position=grid_position,
        talent_indicator="High Potential" if pot == PotentialLevel.HIGH else "Solid Contributor",
        ratings_history=[
            HistoricalRating(year=2023, rating="Solid"),
            HistoricalRating(year=2024, rating="Strong"),
        ] if employee_id % 2 == 0 else [],
        development_focus="Leadership" if employee_id % 3 == 0 else None,
        development_action="Coaching" if employee_id % 3 == 0 else None,
        notes=f"Notes for employee {employee_id}" if employee_id % 4 == 0 else None,
        promotion_status="Ready now" if employee_id % 5 == 0 else None,
        modified_in_session=False,
    )


def create_excel_file(employees: list[Employee], file_path: Path) -> None:
    """Create an Excel file with employee data."""
    workbook = openpyxl.Workbook()

    # Summary sheet
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet["A1"] = "9-Box Talent Review"
    summary_sheet["A2"] = f"Generated: {datetime.now(timezone.utc).isoformat()}"

    # Employee data sheet
    data_sheet = workbook.create_sheet("Employee Data")

    # Headers
    headers = [
        "Employee ID",
        "Worker",
        "Business Title",
        "Job Title",
        "Job Profile",
        "Job Level - Primary Position",
        "Worker's Manager",
        "Management Chain - Level 04",
        "Management Chain - Level 05",
        "Management Chain - Level 06",
        "Hire Date",
        "Tenure Category (Months)",
        "Time in Job Profile",
        "Aug 2025 Talent Assessment Performance",
        "Aug 2025  Talent Assessment Potential",
        "Aug 2025 Talent Assessment 9-Box Label",
        "Talent Mapping Position [Performance vs Potential]",
        "FY25 Talent Indicator",
        "2023 Completed Performance Rating",
        "2024 Completed Performance Rating",
        "Development Focus",
        "Development Action",
        "Notes",
        "Promotion (In-Line,",
    ]

    for col_idx, header in enumerate(headers, start=1):
        data_sheet.cell(1, col_idx, header)

    # Data rows
    for row_idx, emp in enumerate(employees, start=2):
        data_sheet.cell(row_idx, 1, emp.employee_id)
        data_sheet.cell(row_idx, 2, emp.name)
        data_sheet.cell(row_idx, 3, emp.business_title)
        data_sheet.cell(row_idx, 4, emp.job_title)
        data_sheet.cell(row_idx, 5, emp.job_profile)
        data_sheet.cell(row_idx, 6, emp.job_level)
        data_sheet.cell(row_idx, 7, emp.manager)
        data_sheet.cell(row_idx, 8, emp.management_chain_04)
        data_sheet.cell(row_idx, 9, emp.management_chain_05)
        data_sheet.cell(row_idx, 10, emp.management_chain_06)
        data_sheet.cell(row_idx, 11, emp.hire_date)
        data_sheet.cell(row_idx, 12, emp.tenure_category)
        data_sheet.cell(row_idx, 13, emp.time_in_job_profile)
        data_sheet.cell(row_idx, 14, emp.performance.value)
        data_sheet.cell(row_idx, 15, emp.potential.value)
        data_sheet.cell(row_idx, 16, emp.grid_position)
        data_sheet.cell(row_idx, 17, get_position_label_by_number(emp.grid_position))
        data_sheet.cell(row_idx, 18, emp.talent_indicator)

        # Historical ratings
        if len(emp.ratings_history) > 0:
            for rating in emp.ratings_history:
                if rating.year == 2023:
                    data_sheet.cell(row_idx, 19, rating.rating)
                elif rating.year == 2024:
                    data_sheet.cell(row_idx, 20, rating.rating)

        data_sheet.cell(row_idx, 21, emp.development_focus)
        data_sheet.cell(row_idx, 22, emp.development_action)
        data_sheet.cell(row_idx, 23, emp.notes)
        data_sheet.cell(row_idx, 24, emp.promotion_status)

    workbook.save(file_path)
    workbook.close()  # Explicitly close to prevent openpyxl state pollution


@pytest.fixture
def small_excel_file(tmp_path: Path, sample_employees: list[Employee]) -> Path:
    """Create a small Excel file with 5 employees for performance testing."""
    file_path = tmp_path / "small_test.xlsx"
    create_excel_file(sample_employees, file_path)
    return file_path


@pytest.fixture
def large_employee_dataset() -> list[Employee]:
    """Create a large dataset of 1000 employees for performance testing."""
    return [create_test_employee(i) for i in range(1, 1001)]


@pytest.fixture
def large_excel_file(tmp_path: Path, large_employee_dataset: list[Employee]) -> Path:
    """Create a large Excel file with 1000 employees for performance testing."""
    file_path = tmp_path / "large_test.xlsx"
    create_excel_file(large_employee_dataset, file_path)
    return file_path


@pytest.fixture
def test_client_with_session(
    test_client: TestClient,
    small_excel_file: Path,
) -> TestClient:
    """Create a test client with an active session (small dataset)."""
    # Upload file to create session
    with small_excel_file.open("rb") as f:
        files = {"file": (small_excel_file.name, f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        response = test_client.post("/api/session/upload", files=files)
        assert response.status_code == 200

    return test_client


@pytest.fixture
def test_client_with_large_session(
    test_client: TestClient,
    large_excel_file: Path,
) -> TestClient:
    """Create a test client with an active session (large dataset)."""
    # Upload file to create session
    with large_excel_file.open("rb") as f:
        files = {"file": (large_excel_file.name, f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        response = test_client.post("/api/session/upload", files=files)
        assert response.status_code == 200

    return test_client
