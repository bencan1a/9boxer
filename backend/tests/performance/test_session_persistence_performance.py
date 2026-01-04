"""Performance tests for session persistence operations.

Performance targets documented in each test docstring.
Use --benchmark-compare to track performance over time.
"""

from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient

from ninebox.core.dependencies import get_session_manager
from ninebox.models.employee import Employee, PerformanceLevel, PotentialLevel
from ninebox.models.session import SessionState
from ninebox.services.session_manager import SessionManager
from ninebox.services.session_serializer import SessionSerializer

# Check if pytest-benchmark is available
try:
    import pytest_benchmark  # noqa: F401

    BENCHMARK_AVAILABLE = True
except ImportError:
    BENCHMARK_AVAILABLE = False

pytestmark = [
    pytest.mark.performance,
    pytest.mark.slow,
    pytest.mark.skipif(
        not BENCHMARK_AVAILABLE,
        reason="pytest-benchmark not installed. Install with: pip install pytest-benchmark",
    ),
]


@pytest.fixture
def large_dataset(tmp_path: Path) -> Path:
    """Create a large dataset for performance testing (100 employees)."""
    file_path = tmp_path / "large_dataset.xlsx"
    workbook = openpyxl.Workbook()

    # Summary sheet
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet["A1"] = "Large Dataset Test"

    # Employee data sheet
    data_sheet = workbook.create_sheet("Employee Data")

    # Headers
    headers = [
        "Employee ID",
        "Worker",
        "Business Title",
        "Job Title",
        "Job Profile",
        "Job Level - Primary Position",
        "Worker's Manager",
        "Management Chain - Level 04",
        "Management Chain - Level 05",
        "Management Chain - Level 06",
        "Hire Date",
        "Tenure Category (Months)",
        "Time in Job Profile",
        "Aug 2025 Talent Assessment Performance",
        "Aug 2025  Talent Assessment Potential",
        "Aug 2025 Talent Assessment 9-Box Label",
        "Talent Mapping Position [Performance vs Potential]",
    ]

    for col_idx, header in enumerate(headers, start=1):
        data_sheet.cell(1, col_idx, header)

    # Generate 100 employees
    performances = ["Low", "Medium", "High"]
    potentials = ["Low", "Medium", "High"]

    for i in range(100):
        row_idx = i + 2
        data_sheet.cell(row_idx, 1, i + 1)  # Employee ID
        data_sheet.cell(row_idx, 2, f"Employee {i + 1}")  # Name
        data_sheet.cell(row_idx, 3, f"Title {i % 10}")  # Business Title
        data_sheet.cell(row_idx, 4, "Software Engineer")  # Job Title
        data_sheet.cell(row_idx, 5, "EngineeringUSA")  # Job Profile
        data_sheet.cell(row_idx, 6, f"MT{(i % 5) + 1}")  # Job Level
        data_sheet.cell(row_idx, 7, f"Manager {i % 20}")  # Manager
        data_sheet.cell(row_idx, 8, f"Manager {i % 20}")  # MC 04
        data_sheet.cell(row_idx, 9, "Director")  # MC 05
        data_sheet.cell(row_idx, 10, None)  # MC 06
        data_sheet.cell(row_idx, 11, "2020-01-01")  # Hire Date
        data_sheet.cell(row_idx, 12, "3-5 years")  # Tenure
        data_sheet.cell(row_idx, 13, "2 years")  # Time in Job
        data_sheet.cell(row_idx, 14, performances[i % 3])  # Performance
        data_sheet.cell(row_idx, 15, potentials[i % 3])  # Potential
        data_sheet.cell(row_idx, 16, (i % 9) + 1)  # Grid Position
        data_sheet.cell(row_idx, 17, f"[{performances[i % 3][0]},{potentials[i % 3][0]}]")

    workbook.save(file_path)
    workbook.close()  # Explicitly close to prevent openpyxl state pollution
    return file_path


class TestSessionRestorePerformance:
    """Performance benchmarks for session restore operations."""

    def test_session_restore_when_100_employees_then_completes_within_limit(
        self,
        benchmark: pytest.fixture,
        test_client: TestClient,
        large_dataset: Path,
    ) -> None:
        """Benchmark session restore with 100 employees.

        Target: <100ms
        """
        # Upload large dataset
        with open(large_dataset, "rb") as f:  # noqa: PTH123
            files = {
                "file": (
                    "large.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            response = test_client.post("/api/session/upload", files=files)

        assert response.status_code == 200
        assert response.json()["employee_count"] == 100

        # Clear cache before benchmarking
        get_session_manager.cache_clear()

        def restore_session() -> SessionManager:
            return SessionManager()

        # Benchmark the restore operation
        new_manager = benchmark(restore_session)

        # Verify session restored correctly
        assert "local-user" in new_manager.sessions, "Test session not found"
        restored_session = new_manager.sessions["local-user"]
        assert len(restored_session.original_employees) == 100


class TestSessionPersistPerformance:
    """Performance benchmarks for session persistence operations."""

    def test_session_persist_when_employee_move_then_completes_within_limit(
        self,
        benchmark: pytest.fixture,
        test_client: TestClient,
        sample_excel_file: Path,
    ) -> None:
        """Benchmark session persist after employee move.

        Target: <200ms (includes move logic + persistence)
        """
        # Upload file
        with open(sample_excel_file, "rb") as f:  # noqa: PTH123
            files = {
                "file": (
                    "test.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            test_client.post("/api/session/upload", files=files)

        # Get employee
        response = test_client.get("/api/employees")
        employee_id = response.json()["employees"][0]["employee_id"]

        def move_and_persist() -> None:
            response = test_client.patch(
                f"/api/employees/{employee_id}/move",
                json={"performance": "High", "potential": "High"},
            )
            assert response.status_code == 200

        # Benchmark the move + persist operation
        benchmark(move_and_persist)


class TestSerializationPerformance:
    """Performance benchmarks for session serialization."""

    def test_serialize_when_1000_employees_then_completes_within_limit(
        self,
        benchmark: pytest.fixture,
    ) -> None:
        """Benchmark serialization/deserialization of 1000 employees.

        Target: <500ms total
        """
        from datetime import date, datetime, timezone

        # Create 1000 employees
        employees = []
        for i in range(1000):
            employee = Employee(
                employee_id=i + 1,
                name=f"Employee {i + 1}",
                business_title=f"Title {i % 10}",
                job_title="Software Engineer",
                job_profile="EngineeringUSA",
                job_level=f"MT{(i % 5) + 1}",
                job_function="Other",
                location="USA",
                direct_manager=f"Manager {i % 20}",
                management_chain_04=f"Manager {i % 20}",
                management_chain_05="Director",
                management_chain_06=None,
                hire_date=date(2020, 1, 1),
                tenure_category="3-5 years",
                time_in_job_profile="2 years",
                performance=PerformanceLevel(["Low", "Medium", "High"][i % 3]),
                potential=PotentialLevel(["Low", "Medium", "High"][i % 3]),
                grid_position=(i % 9) + 1,
                position_label=f"Label {i % 9}",
                talent_indicator="Solid Contributor",
                ratings_history=[],
                development_focus=None,
                development_action=None,
                notes=None,
                promotion_status=None,
                modified_in_session=False,
            )
            employees.append(employee)

        # Create session
        session = SessionState(
            session_id="test_session",
            user_id="test_user",
            created_at=datetime.now(timezone.utc),
            original_employees=employees,
            current_employees=employees.copy(),
            original_filename="large.xlsx",
            original_file_path="/tmp/large.xlsx",
            sheet_name="Employee Data",
            sheet_index=0,
            job_function_config=None,
            changes=[],
        )

        def serialize_and_deserialize() -> SessionState:
            serialized = SessionSerializer.serialize(session)
            return SessionSerializer.deserialize(serialized)

        # Benchmark the serialize + deserialize operation
        deserialized = benchmark(serialize_and_deserialize)

        # Verify correctness
        assert len(deserialized.original_employees) == 1000
        assert len(deserialized.current_employees) == 1000
