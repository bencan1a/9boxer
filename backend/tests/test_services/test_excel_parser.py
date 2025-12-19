"""Tests for Excel parser service."""

from pathlib import Path

import openpyxl
import pytest

from ninebox.models.employee import PerformanceLevel, PotentialLevel
from ninebox.models.grid_positions import calculate_grid_position, get_position_label
from ninebox.services.excel_parser import ExcelParser


def test_parse_when_valid_file_then_returns_employees(sample_excel_file: Path) -> None:
    """Test parsing a valid Excel file."""
    parser = ExcelParser()
    result = parser.parse(sample_excel_file)

    assert len(result.employees) == 5
    assert all(emp.employee_id > 0 for emp in result.employees)
    assert all(emp.name for emp in result.employees)
    # Verify metadata is populated
    assert result.metadata.sheet_name is not None
    assert result.metadata.parsed_rows == 5
    assert result.metadata.total_rows >= 5


def test_parse_when_valid_file_then_extracts_employee_data_correctly(
    sample_excel_file: Path,
) -> None:
    """Test that employee data is extracted correctly."""
    parser = ExcelParser()
    result = parser.parse(sample_excel_file)

    # Check first employee
    emp = result.employees[0]
    assert emp.employee_id == 1
    assert emp.name == "Alice Smith"
    assert emp.business_title == "Senior Engineer"
    assert emp.job_level == "MT4"
    assert emp.manager == "Bob Manager"
    assert emp.performance == PerformanceLevel.HIGH
    assert emp.potential == PotentialLevel.HIGH


def test_parse_when_valid_file_then_handles_historical_ratings(sample_excel_file: Path) -> None:
    """Test that historical ratings are parsed correctly."""
    parser = ExcelParser()
    result = parser.parse(sample_excel_file)
    employees = result.employees

    # First employee has 2 ratings
    emp = employees[0]
    assert len(emp.ratings_history) == 2
    assert emp.ratings_history[0].year == 2023
    assert emp.ratings_history[0].rating == "Strong"
    assert emp.ratings_history[1].year == 2024
    assert emp.ratings_history[1].rating == "Leading"

    # Second employee has 1 rating
    emp = employees[1]
    assert len(emp.ratings_history) == 1
    assert emp.ratings_history[0].year == 2024


def test_parse_when_valid_file_then_calculates_grid_positions_correctly(
    sample_excel_file: Path,
) -> None:
    """Test grid position calculation."""
    parser = ExcelParser()
    result = parser.parse(sample_excel_file)
    employees = result.employees

    # H,H = High Performance (3), High Potential (6) = 9
    assert employees[0].grid_position == 9
    # M,M = Medium Performance (2), Medium Potential (3) = 5
    assert employees[1].grid_position == 5
    # L,H = Low Performance (1), High Potential (6) = 7
    assert employees[2].grid_position == 7
    # H,M = High Performance (3), Medium Potential (3) = 6
    assert employees[3].grid_position == 6
    # M,H = Medium Performance (2), High Potential (6) = 8
    assert employees[4].grid_position == 8


def test_parse_when_valid_file_then_handles_optional_fields_gracefully(tmp_path: Path) -> None:
    """Test handling of missing optional fields."""
    # Create minimal Excel file
    file_path = tmp_path / "minimal.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Summary"
    data_sheet = workbook.create_sheet("Employee Data")

    headers = ["Employee ID", "Worker", "Business Title", "Job Level - Primary Position"]
    for col_idx, header in enumerate(headers, start=1):
        data_sheet.cell(1, col_idx, header)

    # Minimal employee data
    data_sheet.cell(2, 1, 1)
    data_sheet.cell(2, 2, "Test User")
    data_sheet.cell(2, 3, "Test Title")
    data_sheet.cell(2, 4, "MT1")

    workbook.save(file_path)

    parser = ExcelParser()
    result = parser.parse(file_path)

    assert len(result.employees) == 1
    emp = result.employees[0]
    assert emp.employee_id == 1
    assert emp.name == "Test User"
    assert emp.development_focus is None
    assert emp.notes is None
    # Verify defaulted fields tracking
    assert result.metadata.defaulted_fields.get("Performance", 0) == 1
    assert result.metadata.defaulted_fields.get("Potential", 0) == 1


def test_parse_when_invalid_file_format_then_raises_error(tmp_path: Path) -> None:
    """Test error on invalid file format."""
    # Create a text file instead of Excel
    file_path = tmp_path / "invalid.txt"
    file_path.write_text("This is not an Excel file")

    parser = ExcelParser()
    with pytest.raises(ValueError, match="Failed to read Excel file"):
        parser.parse(file_path)


def test_parse_when_missing_required_columns_then_raises_error(tmp_path: Path) -> None:
    """Test error when required columns are missing."""
    file_path = tmp_path / "missing_columns.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Summary"
    data_sheet = workbook.create_sheet("Employee Data")

    # Only add some headers
    data_sheet.cell(1, 1, "Employee ID")
    data_sheet.cell(1, 2, "Worker")
    # Missing "Business Title" and "Job Level - Primary Position"

    workbook.save(file_path)

    parser = ExcelParser()
    with pytest.raises(ValueError, match="Missing required columns"):
        parser.parse(file_path)


def test_parse_when_no_valid_employees_then_raises_error(tmp_path: Path) -> None:
    """Test error when no valid employees can be parsed."""
    file_path = tmp_path / "no_employees.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Summary"
    data_sheet = workbook.create_sheet("Employee Data")

    # Add headers but no data
    headers = ["Employee ID", "Worker", "Business Title", "Job Level - Primary Position"]
    for col_idx, header in enumerate(headers, start=1):
        data_sheet.cell(1, col_idx, header)

    workbook.save(file_path)

    parser = ExcelParser()
    with pytest.raises(ValueError, match="No valid employees found"):
        parser.parse(file_path)


def test_calculate_position_when_all_combinations_then_returns_correct_positions() -> None:
    """Test all 9 position calculations.

    Standard 9-box grid layout:
        Performance (columns): Low=1, Medium=2, High=3
        Potential (rows): Low=1-3, Medium=4-6, High=7-9
    """
    # Low performance (column 1)
    assert calculate_grid_position(PerformanceLevel.LOW, PotentialLevel.LOW) == 1
    assert calculate_grid_position(PerformanceLevel.LOW, PotentialLevel.MEDIUM) == 4
    assert calculate_grid_position(PerformanceLevel.LOW, PotentialLevel.HIGH) == 7

    # Medium performance (column 2)
    assert calculate_grid_position(PerformanceLevel.MEDIUM, PotentialLevel.LOW) == 2
    assert calculate_grid_position(PerformanceLevel.MEDIUM, PotentialLevel.MEDIUM) == 5
    assert calculate_grid_position(PerformanceLevel.MEDIUM, PotentialLevel.HIGH) == 8

    # High performance (column 3)
    assert calculate_grid_position(PerformanceLevel.HIGH, PotentialLevel.LOW) == 3
    assert calculate_grid_position(PerformanceLevel.HIGH, PotentialLevel.MEDIUM) == 6
    assert calculate_grid_position(PerformanceLevel.HIGH, PotentialLevel.HIGH) == 9


def test_get_position_label_when_all_combinations_then_returns_correct_labels() -> None:
    """Test position label generation."""
    assert get_position_label(PerformanceLevel.HIGH, PotentialLevel.HIGH) == "Star [H,H]"
    assert (
        get_position_label(PerformanceLevel.HIGH, PotentialLevel.MEDIUM) == "High Impact [H,M]"
    )
    assert get_position_label(PerformanceLevel.HIGH, PotentialLevel.LOW) == "Workhorse [H,L]"
    assert get_position_label(PerformanceLevel.MEDIUM, PotentialLevel.HIGH) == "Growth [M,H]"
    assert (
        get_position_label(PerformanceLevel.MEDIUM, PotentialLevel.MEDIUM)
        == "Core Talent [M,M]"
    )
    assert (
        get_position_label(PerformanceLevel.MEDIUM, PotentialLevel.LOW)
        == "Effective Pro [M,L]"
    )
    assert get_position_label(PerformanceLevel.LOW, PotentialLevel.HIGH) == "Enigma [L,H]"
    assert (
        get_position_label(PerformanceLevel.LOW, PotentialLevel.MEDIUM) == "Inconsistent [L,M]"
    )
    assert (
        get_position_label(PerformanceLevel.LOW, PotentialLevel.LOW) == "Underperformer [L,L]"
    )
