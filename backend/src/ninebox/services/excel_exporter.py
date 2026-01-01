"""Excel file exporter service."""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

from ninebox.models.employee import Employee
from ninebox.models.grid_positions import get_position_label, get_position_label_by_number

if TYPE_CHECKING:
    import openpyxl
    from openpyxl.worksheet.worksheet import Worksheet

    from ninebox.models.session import SessionState


class ExcelExporter:
    """Export modified employee data back to Excel."""

    def export(  # Complexity acceptable for data export logic
        self,
        original_file: str | Path,
        employees: list[Employee],
        output_path: str | Path,
        sheet_index: int = 1,
        session: SessionState | None = None,
    ) -> None:
        """
        Create new Excel file with updated ratings.

        Args:
            original_file: Path to original Excel file (can be empty for sample data)
            employees: List of employees with current data
            output_path: Path to save modified Excel file
            sheet_index: Index of the sheet to export to (default: 1 for backward compatibility)
            session: Optional session state for accessing change notes
        """
        # Lazy import openpyxl to reduce startup time
        import openpyxl

        # Check if original file exists
        original_file_path = Path(original_file) if original_file else None
        has_original = (
            original_file_path and original_file_path.exists() and str(original_file).strip()
        )

        # Create or load workbook
        if has_original:
            # Read original file to preserve formatting - use try/finally to ensure cleanup
            workbook = openpyxl.load_workbook(original_file)
        else:
            # Create new workbook from scratch for sample data
            workbook = self._create_workbook_from_employees(employees, session)
        try:
            # Work with specified sheet
            if len(workbook.worksheets) <= sheet_index:
                raise ValueError(
                    f"Excel file must have at least {sheet_index + 1} sheets. "
                    f"Only {len(workbook.worksheets)} sheets found."
                )

            sheet = workbook.worksheets[sheet_index]

            # Find column indices
            perf_col = self._find_column(sheet, "Aug 2025 Talent Assessment Performance")
            pot_col = self._find_column(sheet, "Aug 2025  Talent Assessment Potential")
            box_col = self._find_column(sheet, "Aug 2025 Talent Assessment 9-Box Label")
            label_col = self._find_column(sheet, "Talent Mapping Position")
            promotion_readiness_col = self._find_column(sheet, "Promotion Readiness")

            # Add "Modified" columns if they don't exist
            max_col = sheet.max_column or 1
            modified_col = self._find_column(sheet, "Modified in Session", create=True)
            assert modified_col is not None, "modified_col should not be None when create=True"  # nosec B101  # Type narrowing
            if modified_col > max_col:
                sheet.cell(1, modified_col, "Modified in Session")
                sheet.cell(1, modified_col + 1, "Modification Date")
                sheet.cell(1, modified_col + 2, "9Boxer Change Description")
                sheet.cell(1, modified_col + 3, "9Boxer Change Notes")
                sheet.cell(1, modified_col + 4, "Donut Exercise Position")
                sheet.cell(1, modified_col + 5, "Donut Exercise Label")
                sheet.cell(1, modified_col + 6, "Donut Exercise Change Description")
                sheet.cell(1, modified_col + 7, "Donut Exercise Notes")
                sheet.cell(1, modified_col + 8, "Flags")
                sheet.cell(1, modified_col + 9, "Original Performance")
                sheet.cell(1, modified_col + 10, "Original Potential")

            # Ensure "Original Performance" and "Original Potential" columns exist even for
            # workbooks exported with older versions that only had "Modified in Session".
            original_perf_col = self._find_column(sheet, "Original Performance")
            if original_perf_col is None:
                original_perf_col = self._find_column(sheet, "Original Performance", create=True)
                assert (
                    original_perf_col is not None
                ), "original_perf_col should not be None when create=True"  # nosec B101
                sheet.cell(1, original_perf_col, "Original Performance")

            original_pot_col = self._find_column(sheet, "Original Potential")
            if original_pot_col is None:
                original_pot_col = self._find_column(sheet, "Original Potential", create=True)
                assert (
                    original_pot_col is not None
                ), "original_pot_col should not be None when create=True"  # nosec B101
                sheet.cell(1, original_pot_col, "Original Potential")
            # Create employee lookup by ID
            employee_map = {e.employee_id: e for e in employees}

            # Create original employee lookup by ID (for tracking original values)
            original_employee_map = {}
            if session:
                original_employee_map = {e.employee_id: e for e in session.original_employees}

            # Create change notes and descriptions lookup by employee ID
            change_notes_map = {}
            change_description_map = {}
            if session:
                for event in session.events:
                    # Only process grid move events
                    if event.event_type == "grid_move":
                        if event.notes:
                            change_notes_map[event.employee_id] = event.notes
                        # Create movement description
                        old_label = get_position_label(event.old_performance, event.old_potential)
                        new_label = get_position_label(event.new_performance, event.new_potential)
                        change_description_map[
                            event.employee_id
                        ] = f"Moved from {old_label} to {new_label}"

            # Create donut change descriptions lookup by employee ID
            donut_change_description_map = {}
            if session:
                for event in session.donut_events:
                    # Only process donut move events
                    if event.event_type == "donut_move":
                        # Create donut movement description
                        old_label = get_position_label(event.old_performance, event.old_potential)
                        new_label = get_position_label(event.new_performance, event.new_potential)
                        donut_change_description_map[
                            event.employee_id
                        ] = f"Donut: Moved from {old_label} to {new_label}"

            # Update rows with modified data
            for row_idx in range(2, sheet.max_row + 1):
                emp_id_cell = sheet.cell(row_idx, 1).value
                if emp_id_cell is None:
                    continue

                try:
                    # Convert cell value to int, handling various types
                    emp_id = int(str(emp_id_cell))
                except (ValueError, TypeError):
                    continue

                if emp_id in employee_map:
                    emp = employee_map[emp_id]

                    # Update performance data
                    if perf_col:
                        sheet.cell(row_idx, perf_col, emp.performance.value)
                    if pot_col:
                        sheet.cell(row_idx, pot_col, emp.potential.value)
                    if box_col:
                        sheet.cell(row_idx, box_col, emp.grid_position)
                    if label_col:
                        sheet.cell(
                            row_idx, label_col, get_position_label_by_number(emp.grid_position)
                        )

                    # Update promotion readiness
                    if promotion_readiness_col and emp.promotion_readiness is not None:
                        sheet.cell(
                            row_idx,
                            promotion_readiness_col,
                            "Yes" if emp.promotion_readiness else "No",
                        )

                    # Mark as modified
                    sheet.cell(row_idx, modified_col, "Yes" if emp.modified_in_session else "No")
                    if emp.modified_in_session and emp.last_modified:
                        sheet.cell(row_idx, modified_col + 1, emp.last_modified.isoformat())

                    # Add change description if available
                    description_value = change_description_map.get(emp_id, "")
                    sheet.cell(row_idx, modified_col + 2, description_value)

                    # Add change notes if available
                    notes_value = change_notes_map.get(emp_id, "")
                    sheet.cell(row_idx, modified_col + 3, notes_value)

                    # Add donut exercise data if employee was modified in donut mode
                    if emp.donut_modified:
                        # Donut Exercise Position
                        sheet.cell(row_idx, modified_col + 4, emp.donut_position)
                        # Donut Exercise Label
                        sheet.cell(
                            row_idx,
                            modified_col + 5,
                            get_position_label_by_number(emp.donut_position)
                            if emp.donut_position
                            else "",
                        )
                        # Donut Exercise Change Description
                        donut_description = donut_change_description_map.get(emp_id, "")
                        sheet.cell(row_idx, modified_col + 6, donut_description)
                        # Donut Exercise Notes
                        sheet.cell(row_idx, modified_col + 7, emp.donut_notes or "")
                    else:
                        # Leave donut columns empty if not modified in donut mode
                        sheet.cell(row_idx, modified_col + 4, "")
                        sheet.cell(row_idx, modified_col + 5, "")
                        sheet.cell(row_idx, modified_col + 6, "")
                        sheet.cell(row_idx, modified_col + 7, "")

                    # Add flags (comma-separated list)
                    flags_value = ", ".join(emp.flags) if emp.flags else ""
                    sheet.cell(row_idx, modified_col + 8, flags_value)

                    # Add original Performance/Potential values if employee was modified
                    if emp.modified_in_session and emp_id in original_employee_map:
                        original_emp = original_employee_map[emp_id]
                        sheet.cell(row_idx, modified_col + 9, original_emp.performance.value)
                        sheet.cell(row_idx, modified_col + 10, original_emp.potential.value)
                    else:
                        # Leave original value columns empty if not modified
                        sheet.cell(row_idx, modified_col + 9, "")
                        sheet.cell(row_idx, modified_col + 10, "")

            # Save modified workbook
            workbook.save(output_path)

        finally:
            # Always close workbook to release file handles, even if exception occurred
            workbook.close()

    def _create_workbook_from_employees(
        self, employees: list[Employee], session: SessionState | None = None
    ) -> openpyxl.Workbook:
        """
        Create a new Excel workbook from employee data.

        Used when exporting sample data that has no original file.

        Args:
            employees: List of employees to include
            session: Optional session state for metadata

        Returns:
            New openpyxl Workbook with employee data
        """
        from datetime import datetime, timezone

        import openpyxl

        workbook = openpyxl.Workbook()

        # Summary sheet
        summary_sheet = workbook.active
        assert summary_sheet is not None  # nosec B101  # Type narrowing
        summary_sheet.title = "Summary"
        summary_sheet["A1"] = "9-Box Talent Review"
        summary_sheet["A2"] = f"Generated: {datetime.now(timezone.utc).isoformat()}"
        if session:
            summary_sheet["A3"] = f"Source: {session.original_filename or 'Sample Data'}"
            summary_sheet["A4"] = f"Employees: {len(employees)}"

        # Employee data sheet
        data_sheet = workbook.create_sheet("Employee Data")

        # Headers - matching the expected format from ExcelParser
        headers = [
            "Employee ID",
            "Worker",
            "Business Title",
            "Job Title",
            "Job Profile",
            "Job Level - Primary Position",
            "Worker's Manager",
            "Management Chain - Level 04",
            "Management Chain - Level 05",
            "Management Chain - Level 06",
            "Hire Date",
            "Tenure Category (Months)",
            "Time in Job Profile",
            "Aug 2025 Talent Assessment Performance",
            "Aug 2025  Talent Assessment Potential",
            "Aug 2025 Talent Assessment 9-Box Label",
            "Talent Mapping Position",
            "FY25 Talent Indicator",
            "2023 Completed Performance Rating",
            "2024 Completed Performance Rating",
            "Development Focus",
            "Development Action",
            "Notes",
            "Promotion Readiness",
        ]

        for col_idx, header in enumerate(headers, start=1):
            data_sheet.cell(1, col_idx, header)

        # Data rows
        for row_idx, emp in enumerate(employees, start=2):
            data_sheet.cell(row_idx, 1, emp.employee_id)
            data_sheet.cell(row_idx, 2, emp.name)
            data_sheet.cell(row_idx, 3, emp.business_title)
            data_sheet.cell(row_idx, 4, emp.job_title)
            data_sheet.cell(row_idx, 5, emp.job_profile)
            data_sheet.cell(row_idx, 6, emp.job_level)
            data_sheet.cell(row_idx, 7, emp.direct_manager)
            data_sheet.cell(row_idx, 8, emp.management_chain_04)
            data_sheet.cell(row_idx, 9, emp.management_chain_05)
            data_sheet.cell(row_idx, 10, emp.management_chain_06)
            data_sheet.cell(row_idx, 11, emp.hire_date.isoformat() if emp.hire_date else None)
            data_sheet.cell(row_idx, 12, emp.tenure_category)
            data_sheet.cell(row_idx, 13, emp.time_in_job_profile)
            data_sheet.cell(row_idx, 14, emp.performance.value)
            data_sheet.cell(row_idx, 15, emp.potential.value)
            data_sheet.cell(row_idx, 16, emp.grid_position)
            data_sheet.cell(row_idx, 17, get_position_label_by_number(emp.grid_position))
            data_sheet.cell(row_idx, 18, emp.talent_indicator)

            # Historical ratings
            if emp.ratings_history:
                for rating in emp.ratings_history:
                    if rating.year == 2023:
                        data_sheet.cell(row_idx, 19, rating.rating)
                    elif rating.year == 2024:
                        data_sheet.cell(row_idx, 20, rating.rating)

            data_sheet.cell(row_idx, 21, emp.development_focus)
            data_sheet.cell(row_idx, 22, emp.development_action)
            data_sheet.cell(row_idx, 23, emp.notes)
            data_sheet.cell(row_idx, 24, emp.promotion_status)

        return workbook

    def _find_column(self, sheet: Worksheet, col_name: str, create: bool = False) -> int | None:
        """Find column index by name."""
        for col_idx in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, col_idx).value
            if cell_value and col_name in str(cell_value):
                return col_idx

        # If not found and create=True, return next available column
        if create:
            return (sheet.max_column or 0) + 1

        return None
